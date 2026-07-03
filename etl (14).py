"""
DHC Working automation — core ETL pipeline.

Replicates, in pandas, the manual VLOOKUP/COUNTIFS/pivot workflow found inside
DHC_Working_-_Jun_26.xlsx. See ROADMAP.md for the full reverse-engineered logic
map and the open questions that need confirming with the process owner.
"""
import io
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Static mapping tables copied verbatim from the workbook's "Sheet3" tab.
# These are the small reference tables mam's VLOOKUPs point at.
# ---------------------------------------------------------------------------
MODE_MAP = {
    "AIRTEL": "AIRTEL / CASH", "CASH": "AIRTEL / CASH",
    "CHEQUE": "CHQ / DD", "DD": "CHQ / DD",
    "ONLINE_PAYMENT": "ONLINE_PAYMENT", "RTGS": "RTGS",
}
STATUS_MAP = {"B": "Bounced", "C": "Cleared", "D": "Deposit", "NA": "Pending", "X": "Cxn"}
RECEIPT_SOURCE_MAP = {
    "BBPS": "BBPS", "CHOLAONE DIRECT": "CHOLAONE DIRECT",
    "CCP - BITLY": "CCP - BITLY", "CCP - QR": "CCP - QR", "CCP": "CCP - QR",
}
RECEIPT_TYPE_MAP = {
    "Part Payment": "Part Payment", "FC": "Settlement",
    "Sale/EMD receipt": "Settlement", "Settlement": "Settlement", "OD": "OD",
}
RECEIPT_TYPE_FALLBACK = {"OD": "OD", "OTHER OD": "OTHER OD"}

# Label used whenever a raw code doesn't match any of the maps above (typo,
# stray whitespace, different casing, a genuinely new code, or a blank
# cell). Rows are NEVER dropped/NaN'd because of an unmapped code anymore —
# they fall in here instead, so every count in the output can be traced
# back to 100% of the input rows.
UNMAPPED_LABEL = "Other/Unmapped"


def _map_with_fallback(series: pd.Series, mapping: dict, fallback: str = UNMAPPED_LABEL) -> pd.Series:
    """
    Same intent as pandas' .map(), except it never produces NaN for a
    matchable-but-messy input. It strips whitespace and upper-cases both
    sides before matching, so "cash", "Cash ", and "CASH" all resolve to
    the same mapped value. Anything that still doesn't match a known key
    (including genuinely blank cells) becomes `fallback` rather than NaN,
    so it can be counted and audited instead of silently vanishing.
    """
    norm_map = {str(k).strip().upper(): v for k, v in mapping.items()}
    normalized = series.astype(str).str.strip().str.upper()
    return normalized.map(norm_map).fillna(fallback)


EXCEL_DATE_COLUMNS = [
    "Date", "TXN DATE IN PL TAB", "RECEIPT ENTER DATE",
    "RECEIPTENTEREDTIME", "VALUEDATE", "TXNDATE",
]


def _fix_excel_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    pyxlsb (unlike openpyxl) returns Excel date cells as raw serial numbers,
    not datetimes. Convert the known date columns so downstream date-math
    behaves correctly.
    """
    for col in EXCEL_DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], unit="D", origin="1899-12-30", errors="coerce")
    return df


def tat_bucket(days):
    """
    Replicates the approximate-match VLOOKUP against Sheet3!A:B for rows
    that have a real Ageing figure. Rows whose TXN DATE IN PL TAB carried
    the "NA" (Pending) or "X" (Cxn) code instead of a date have no Ageing
    to bucket here — those are overlaid afterwards from TXN DATE CODE (see
    build_rtgs_tab / build_delay_in_rcpting_summary), so this function
    still returns None for them and callers must apply that overlay.
    """
    if pd.isna(days):
        return None
    if days < 5:
        return "Less then 4"
    if days < 11:
        return "5 TO 10"
    return "Great then 10"


def _apply_tat_code_overlay(df: pd.DataFrame) -> pd.DataFrame:
    """
    Routes TXN DATE IN PL TAB's "NA"/"X" codes into their own TAT buckets
    (Pending / Cxn) instead of leaving TAT as None, which is what caused
    these rows to be silently dropped by compute_zone_tat_matrix's
    dropna(subset=["ZONE", "TAT"]) — even though their ZONE was known and
    they belonged in the summary. No-op if TXN DATE CODE isn't present.
    """
    if "TXN DATE CODE" not in df.columns:
        return df
    df.loc[df["TXN DATE CODE"] == "NA", "TAT"] = "Pending"
    df.loc[df["TXN DATE CODE"] == "X", "TAT"] = "Cxn"
    return df


# ---------------------------------------------------------------------------
# Loaders — one per source file
# ---------------------------------------------------------------------------
def load_dcr(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """DCR.xlsb -> (raw receipts [Sheet1], agreement slab master [Sheet2])."""
    receipts = pd.read_excel(file, sheet_name="Sheet1", engine="pyxlsb")
    receipts = _fix_excel_dates(receipts)

    # STATUS IN TAB uses the literal code "NA" to mean Pending. TXN DATE IN
    # PL TAB uses the same trick for a different purpose: "NA" there means
    # no transaction date has posted yet (Pending), and "X" means the
    # transaction was cancelled (Cxn) — neither is a real date. Pandas'
    # default NA-value list treats the bare text "NA" as missing on every
    # read, so every Pending STATUS IN TAB row was silently turning blank
    # before STATUS_MAP ever saw it, AND every NA/X-coded TXN DATE IN PL
    # TAB row was silently turning into NaT — which zeroed out Ageing/TAT
    # and made the row vanish outright from every TAT summary matrix
    # (RTGS Summary, Delay in RCPTING Summary), rather than being counted
    # under a Pending/Cxn bucket. keep_default_na=False fixes that — but
    # applying it to the WHOLE sheet breaks numeric/date columns instead
    # (blank date cells stop being real NaN, forcing those columns to
    # object dtype and crashing pd.to_datetime). So: re-read ONLY these
    # columns with protection on, and patch them back in by row position.
    protected_cols = [c for c in ("STATUS IN TAB", "TXN DATE IN PL TAB") if c in receipts.columns]
    if protected_cols:
        if hasattr(file, "seek"):
            file.seek(0)
        protected = pd.read_excel(
            file, sheet_name="Sheet1", engine="pyxlsb",
            usecols=protected_cols, keep_default_na=False,
        )
        if "STATUS IN TAB" in protected_cols:
            status_col = protected["STATUS IN TAB"].replace("", np.nan)  # genuine blanks still count as missing
            receipts["STATUS IN TAB"] = status_col.values
        if "TXN DATE IN PL TAB" in protected_cols:
            txn_raw = protected["TXN DATE IN PL TAB"].astype(str).str.strip().str.upper()
            # Only "NA" and "X" are real codes. A genuine date reads back
            # here as a numeric serial string, and a genuine blank cell
            # reads back as "" — neither matches, so TXN DATE CODE stays
            # NaN for those and the normal Ageing/TAT date math (via
            # _fix_excel_dates below) is completely untouched for them.
            receipts["TXN DATE CODE"] = np.where(txn_raw.isin(["NA", "X"]), txn_raw.values, np.nan)

    master = pd.read_excel(file, sheet_name="Sheet2", engine="pyxlsb", usecols=[0, 1, 2, 3])
    master.columns = ["AGREEMENTNO", "OPENING_DPD", "OPNG_SLAB_TYPE", "OPENING_SLAB"]
    master = master.dropna(subset=["AGREEMENTNO"]).drop_duplicates("AGREEMENTNO", keep="last")
    return receipts, master


def _normalize_cif(series: pd.Series) -> pd.Series:
    """
    Coerce a CIF column to a consistent string form regardless of source:
    the delinquency CSV reads CIF as text, but the Disable List .xlsb reads
    it as a number (int/float) via pyxlsb. Merging a string CIF against a
    numeric CIF raises a hard pandas error, so every loader that touches
    CIF numbers must run through this before being used as a merge key.
    """
    s = series.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)  # e.g. "23174458.0" -> "23174458"
    return s


def load_disable_lists(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """To_be_Disabled.xlsb -> (cif_level_disable, agreement_level_disable)."""
    cif = pd.read_excel(file, sheet_name="CIF Level Disable", engine="pyxlsb", usecols=[0, 15])
    cif.columns = ["CIF_NO", "Status"]
    cif = cif.dropna(subset=["CIF_NO"])
    cif["CIF_NO"] = _normalize_cif(cif["CIF_NO"])
    agr = pd.read_excel(file, sheet_name="Agreement Level Disble", engine="pyxlsb", usecols=[0, 5])
    agr.columns = ["AGREEMENTNO", "Status"]
    agr = agr.dropna(subset=["AGREEMENTNO"])
    return cif, agr


def load_employee_mobiles(file) -> set[str]:
    """CIFCL_CBSL_List.xlsx -> set of employee/agent mobile numbers (as strings)."""
    emp = pd.read_excel(file, sheet_name=0, usecols=["Mobile Number"])
    nums = emp["Mobile Number"].dropna().astype(str).str.replace(r"\.0$", "", regex=True)
    return set(nums)


def load_delinquency_master(file) -> pd.DataFrame:
    """
    LAP_Delq_Dump.csv -> AGREEMENTNO -> CIF_NO / Opening Slab table.

    Replaces the old carried-forward 'Look Up' master. The delinquency dump
    is produced fresh from source every period and already carries the CIF
    number and DPD slab for every live agreement, so there's no need to
    pull forward a manually-maintained master from last month's output —
    this is a cleaner source of truth for CIF mapping than a stale carry-
    forward file.

    NB: this export has ~90k fully-blank trailing rows (a known artifact
    of the report tool's fixed-size export) — dropna() below strips them.
    """
    df = pd.read_csv(
        file,
        usecols=["AGREEMENTNO", "CIF NO", "DPD", "DPD SLAB"],
        dtype={"AGREEMENTNO": str, "CIF NO": str},
        low_memory=False,
    )
    # usecols preserves the file's own column order, not the order listed
    # above — reselect explicitly before renaming so columns can't get
    # silently swapped (e.g. CIF_NO ending up mapped to the DPD column).
    df = df[["AGREEMENTNO", "CIF NO", "DPD SLAB", "DPD"]]
    df.columns = ["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB"]
    df["AGREEMENTNO"] = df["AGREEMENTNO"].str.strip()
    df["CIF_NO"] = _normalize_cif(df["CIF_NO"])
    df = df.dropna(subset=["AGREEMENTNO"])
    df = df.drop_duplicates("AGREEMENTNO", keep="last")
    return df[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB"]]


# ---------------------------------------------------------------------------
# Transform — build the working tabs
# ---------------------------------------------------------------------------
def build_lookup_master(dcr_master: pd.DataFrame, delinquency_master: pd.DataFrame) -> pd.DataFrame:
    """
    For every agreement appearing in this period's DCR (Sheet2), pull its
    CIF number and Opening Slab straight from this period's delinquency
    dump (fresh source, not carried forward). An agreement with no match
    in the delinquency dump — e.g. fully closed, or too new to have
    appeared in the delinquency pull yet — is flagged for manual CIF
    completion instead of being silently left unmapped.
    """
    merged = dcr_master[["AGREEMENTNO"]].drop_duplicates().merge(
        delinquency_master, on="AGREEMENTNO", how="left",
    )
    merged["NEEDS_CIF_MAPPING"] = merged["CIF_NO"].isna()
    return merged[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB", "NEEDS_CIF_MAPPING"]]


def build_dcr_tab(
    receipts: pd.DataFrame,
    lookup_master: pd.DataFrame,
    agr_disable: pd.DataFrame,
    cif_disable: pd.DataFrame,
    employee_mobiles: set[str],
) -> pd.DataFrame:
    """Replicates DCR tab columns 74-84 (the 11 derived/lookup columns)."""
    df = receipts.copy()

    df = df.merge(
        lookup_master[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "NEEDS_CIF_MAPPING"]],
        on="AGREEMENTNO", how="left",
    )
    df["CIF"] = df["CIF_NO"]

    df["Unique Mob number"] = df.groupby("MOBILENO")["MOBILENO"].transform("count")
    df["Mob Num VS Emp Mob Num"] = df["MOBILENO"].astype(str).str.replace(
        r"\.0$", "", regex=True
    ).isin(employee_mobiles)

    df["Mode"] = _map_with_fallback(df["MODEOFPAYMENT"], MODE_MAP)
    df["Status"] = _map_with_fallback(df["STATUS IN TAB"], STATUS_MAP)
    df["Receipt Source"] = _map_with_fallback(df["RECEIPTSOURCE"], RECEIPT_SOURCE_MAP)

    # Audit trail: raw codes that fell through to "Other/Unmapped", so you
    # can see exactly what's driving that bucket and decide whether to add
    # it to MODE_MAP / STATUS_MAP, or whether it's genuinely junk data.
    df.attrs["unmapped_mode_codes"] = (
        df.loc[df["Mode"] == UNMAPPED_LABEL, "MODEOFPAYMENT"].value_counts().to_dict()
    )
    df.attrs["unmapped_status_codes"] = (
        df.loc[df["Status"] == UNMAPPED_LABEL, "STATUS IN TAB"].value_counts().to_dict()
    )

    # Zone / Sub Region: the raw DCR extract already carries these per-row
    # (ZONE NEW / SUB REGION columns) — no need for the CIF master here.
    # NB: "Zone" here (from ZONE NEW) is the broad 4-way field used for
    # Cash Mode Validation / RCPT CXN. The RTGS Summary and Delay in
    # RCPTING Summary pivots group by the raw "ZONE" column instead
    # (EAST/NORTH/SOUTH_1/SOUTH_2/WEST_1/WEST_2) — see
    # compute_zone_tat_matrix, confirmed against the reference workbook.
    df["Zone"] = df["ZONE NEW"]
    df["Sub Region"] = df["SUB REGION"]

    # Slab: prefer this month's fresh Sheet2 value (OPNG SLAB/SLAB already
    # in the raw extract); this matches the BT/BU formulas which re-pull
    # from DCR.xlsb's own Sheet2 rather than the stale Look Up master.
    df["Slab"] = df["SLAB"]

    df = df.merge(
        agr_disable.rename(columns={"Status": "Ag Level cash mode"}),
        on="AGREEMENTNO", how="left",
    )
    df = df.merge(
        cif_disable.rename(columns={"CIF_NO": "CIF", "Status": "CIF LEVEL"}),
        on="CIF", how="left",
    )
    return df


def build_rtgs_tab(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """Filters to RTGS-mode receipts and computes Ageing / TAT / Receipt Type."""
    rtgs = dcr_tab[dcr_tab["MODEOFPAYMENT"] == "RTGS"].copy()
    rtgs["Ageing"] = (
        pd.to_datetime(rtgs["RECEIPT ENTER DATE"]) - pd.to_datetime(rtgs["TXN DATE IN PL TAB"])
    ).dt.days
    rtgs["TAT"] = rtgs["Ageing"].apply(tat_bucket)
    rtgs = _apply_tat_code_overlay(rtgs)
    rtgs["Receipt Type"] = rtgs["RECEIPTTYPE"].map(RECEIPT_TYPE_MAP)
    rtgs["Receipt Type"] = rtgs["Receipt Type"].fillna(
        rtgs["RECEIPT CAT"].map(RECEIPT_TYPE_FALLBACK)
    )
    return rtgs


# ---------------------------------------------------------------------------
# Display constants for the pivot-style summary sheets
# ---------------------------------------------------------------------------
RECEIPT_TYPE_ORDER = ["OD", "Settlement", "Part Payment", "OTHER OD"]
RECEIPT_TYPE_DISPLAY = {
    "OD": "EMI OD/Charges",
    "Settlement": "FORECLOSURE/SETTLEMENT",
    "Part Payment": "PART PAYMENT",
    "OTHER OD": "Other OD",
}
TAT_ORDER = ["Less then 4", "5 TO 10", "Great then 10", "Pending", "Cxn"]
TAT_DISPLAY = {
    "Less then 4": "< 4 Days", "5 TO 10": "5 - 10 Days", "Great then 10": "> 10 Days",
    "Pending": "Pending", "Cxn": "Cancelled",
}
MODE_ORDER = ["AIRTEL / CASH", "CHQ / DD", "ONLINE_PAYMENT", "RTGS", UNMAPPED_LABEL]


# ---------------------------------------------------------------------------
# Summary sheet builders — each returns a plain-data structure (dicts/lists),
# kept separate from Excel writing so the numbers can be unit-tested or
# previewed in Streamlit without touching openpyxl.
# ---------------------------------------------------------------------------
def _tat_stats(df: pd.DataFrame, value_col: str = "AMOUNTPAID") -> dict:
    out = {"by_bucket": {}, "total_count": 0, "total_value": 0.0}
    for bucket in TAT_ORDER:
        sub = df[df["TAT"] == bucket]
        cnt = int(len(sub))
        val = float(sub[value_col].sum()) / 1e7
        out["by_bucket"][bucket] = {"count": cnt, "value": val}
        out["total_count"] += cnt
        out["total_value"] += val
    return out


def compute_zone_tat_matrix(df: pd.DataFrame, value_col: str = "AMOUNTPAID") -> dict:
    """
    Zone x Receipt Type x TAT bucket — the structure behind both 'RTGS
    Summary' and 'Delay in RCPTING Summary'. Zone here means the raw
    "ZONE" field (EAST, NORTH, SOUTH_1, SOUTH_2, WEST_1, WEST_2) —
    confirmed by reproducing the reference workbook's pivot row-for-row
    against real data. This is NOT the same as "Sub Zone" (over-splits
    EAST/NORTH into EAST_1/EAST_2/NORTH_1-3, which don't exist as row
    labels in the real pivot) nor "ZONE NEW" (too coarse — collapses
    SOUTH_1/SOUTH_2 and WEST_1/WEST_2 together).
    """
    g = df.dropna(subset=["ZONE", "TAT"]).copy()
    zones = sorted(g["ZONE"].unique())
    blocks = []
    grand = _tat_stats(g.iloc[0:0], value_col)  # zeroed template
    for zone in zones:
        zdf = g[g["ZONE"] == zone]
        subtotal = _tat_stats(zdf, value_col)
        breakdown = [
            (code, RECEIPT_TYPE_DISPLAY[code], _tat_stats(zdf[zdf["Receipt Type"] == code], value_col))
            for code in RECEIPT_TYPE_ORDER
        ]
        blocks.append({"zone": zone, "subtotal": subtotal, "breakdown": breakdown})
        for bucket in TAT_ORDER:
            grand["by_bucket"][bucket]["count"] += subtotal["by_bucket"][bucket]["count"]
            grand["by_bucket"][bucket]["value"] += subtotal["by_bucket"][bucket]["value"]
        grand["total_count"] += subtotal["total_count"]
        grand["total_value"] += subtotal["total_value"]
    return {"zones": blocks, "grand_total": grand}


def compute_online_receipt_source_block(df: pd.DataFrame) -> dict:
    """Simple Receipt Source count among ONLINE_PAYMENT-mode rows."""
    sub = df[df["Mode"] == "ONLINE_PAYMENT"]
    counts = sub["RECEIPTSOURCE"].value_counts(dropna=True)
    rows = [(name, int(cnt)) for name, cnt in counts.items()]
    return {"rows": rows, "total": int(counts.sum())}


def build_rtgs_summary(rtgs_tab: pd.DataFrame, dcr_tab: pd.DataFrame) -> dict:
    """
    The Zone x Receipt Type x TAT matrix comes from the RTGS-filtered tab,
    narrowed to RECEIPT STATUS == "Updated" — confirmed against the
    reference workbook (RTGS.xlsx): its pivot table has a report filter
    reading "RECEIPT STATUS: Updated", so Pending/Bounced-or-Cancelled RTGS
    receipts are excluded from turn-around-time performance. rtgs_tab
    itself is left untouched (still all RTGS-mode receipts, any status) —
    only this summary's matrix is narrowed, so the raw "RTGS receipts"
    count elsewhere in the app doesn't silently change.
    The 'Online Payment — Receipt Source' mini-block is a secondary KPI on
    the same dashboard and is computed from the full month's DCR data (it
    would always be empty if computed from rtgs_tab, since Mode there is
    always 'RTGS').
    """
    updated_only = rtgs_tab[rtgs_tab["RECEIPT STATUS"] == "Updated"]
    return {
        "matrix": compute_zone_tat_matrix(updated_only),
        "online_source_block": compute_online_receipt_source_block(dcr_tab),
    }


def build_delay_in_rcpting_summary(dcr_tab: pd.DataFrame) -> dict:
    full = dcr_tab.copy()
    full["Ageing"] = (
        pd.to_datetime(full["RECEIPT ENTER DATE"]) - pd.to_datetime(full["TXN DATE IN PL TAB"])
    ).dt.days
    full["TAT"] = full["Ageing"].apply(tat_bucket)
    full = _apply_tat_code_overlay(full)
    full["Receipt Type"] = full["RECEIPTTYPE"].map(RECEIPT_TYPE_MAP)
    full["Receipt Type"] = full["Receipt Type"].fillna(
        full["RECEIPT CAT"].map(RECEIPT_TYPE_FALLBACK)
    )
    return {
        "matrix": compute_zone_tat_matrix(full),
        "online_source_block": compute_online_receipt_source_block(full),
    }


def _status_table(df: pd.DataFrame, status_groups: list[str], status_cols: list[str]) -> dict:
    """
    RECEIPT STATUS (row group) x Mode (row) x Status (column) counts,
    used for both halves of 'Receipt made summary'.
    """
    groups = []
    grand = {c: 0 for c in status_cols}
    grand_total = 0
    for status in status_groups:
        sdf = df[df["RECEIPT STATUS"] == status]
        if sdf.empty:
            continue
        modes = [m for m in MODE_ORDER if m in sdf["Mode"].unique()]
        rows = []
        group_totals = {c: 0 for c in status_cols}
        for mode in modes:
            mdf = sdf[sdf["Mode"] == mode]
            counts = {c: int((mdf["Status"] == c).sum()) for c in status_cols}
            row_total = sum(counts.values())
            rows.append({"mode": mode, "counts": counts, "total": row_total})
            for c in status_cols:
                group_totals[c] += counts[c]
                grand[c] += counts[c]
            grand_total += row_total
        groups.append({"status": status, "rows": rows, "totals": group_totals})
    return {"groups": groups, "grand_totals": grand, "grand_total": grand_total}


def build_receipt_made_summary(dcr_tab: pd.DataFrame) -> dict:
    """
    Two side-by-side tables, both grouped from the same RECEIPT STATUS field:
    left = Updated/Pending x (Cleared, Deposited, Pending);
    right = Updated/Bounced-or-Cancelled x (Cleared, Deposited, Bounced, Cxn).
    """
    status_relabel = {
        "Updated": "UPDATED",
        "Updation Pending": "PENDING",
        "Bounced-or-Cancelled": "BOUNCED/CANCELLED",
    }
    d = dcr_tab.copy()
    d["RECEIPT STATUS"] = d["RECEIPT STATUS"].map(status_relabel).fillna(d["RECEIPT STATUS"])
    left = _status_table(d, ["UPDATED", "PENDING"], ["Cleared", "Deposit", "Pending"])
    right = _status_table(d, ["UPDATED", "BOUNCED/CANCELLED"], ["Cleared", "Deposit", "Bounced", "Cxn"])
    return {"left": left, "right": right}


# Friendlier display names for the 3 codes that actually appear in STATUS
# IN TAB day to day. "Bounced"/"Deposit" are kept as-is if they ever show
# up — they're real STATUS_MAP outputs too, just rarer in practice.
STATUS_DISPLAY_RENAME = {"Cxn": "Cancelled"}
STATUS_DISPLAY_ORDER = ["Cleared", "Pending", "Cancelled", "Bounced", "Deposit", UNMAPPED_LABEL]


def build_receipt_made_summary_by_status(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Clean Mode x Status breakdown, sourced directly from STATUS IN TAB
    (C -> Cleared, NA -> Pending, X -> Cancelled), independent of the
    separate 'RECEIPT STATUS' field the legacy two-table replica above
    splits on. Every receipt is counted exactly once: each Mode row's
    Total always equals the true row count for that mode, and the Grand
    Total always equals len(dcr_tab) — so this always reconciles against
    a plain filter on the raw data, with nothing silently dropped.
    """
    d = dcr_tab.copy()
    d["Status Display"] = d["Status"].replace(STATUS_DISPLAY_RENAME)

    modes = [m for m in MODE_ORDER if m in d["Mode"].unique()]
    modes += [m for m in d["Mode"].unique() if m not in modes]  # safety net for any surprise value

    statuses = [s for s in STATUS_DISPLAY_ORDER if s in d["Status Display"].unique()]
    statuses += [s for s in d["Status Display"].unique() if s not in statuses]  # same safety net

    rows = []
    for mode in modes:
        mdf = d[d["Mode"] == mode]
        row = {"Mode": mode}
        for s in statuses:
            row[s] = int((mdf["Status Display"] == s).sum())
        row["Total"] = len(mdf)
        rows.append(row)

    out = pd.DataFrame(rows, columns=["Mode"] + statuses + ["Total"])
    grand = {"Mode": "Grand Total"}
    for s in statuses:
        grand[s] = int(out[s].sum())
    grand["Total"] = int(out["Total"].sum())
    out = pd.concat([out, pd.DataFrame([grand])], ignore_index=True)
    return out


def receipt_made_summary_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    """Single-sheet, formatted .xlsx containing just this one summary table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt Made Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B7C0C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(df.columns)
    ws.append(["Receipt Made Summary — by Payment Mode & Status"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}  ·  Source: STATUS IN TAB (C / NA / X)"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(list(df.columns))
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for _, r in df.iterrows():
        ws.append(list(r))

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        is_total = row[0].value == "Grand Total"
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.column_letter != "A" else "left")
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(col)) + 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def receipt_made_summary_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Dashboard-style PNG: a row of KPI stat boxes up top (grand totals),
    then one soft-shadowed card per payment mode, each containing its own
    colour-coded status boxes. Deliberately built as a card dashboard
    rather than a row-per-line table, so it reads as a professional
    summary graphic rather than a spreadsheet screenshot.
    """
    from PIL import Image, ImageDraw, ImageFont, ImageFilter

    def font(size: int) -> ImageFont.ImageFont:
        try:
            return ImageFont.load_default(size=size)
        except TypeError:
            return ImageFont.load_default()

    status_colors = {
        "Cleared": (39, 174, 96),
        "Pending": (243, 156, 18),
        "Cancelled": (231, 76, 60),
        "Bounced": (192, 57, 43),
        "Deposit": (41, 128, 185),
        UNMAPPED_LABEL: (127, 140, 141),
    }
    mode_colors = {
        "AIRTEL / CASH": (26, 188, 156),
        "CHQ / DD": (142, 68, 173),
        "ONLINE_PAYMENT": (52, 152, 219),
        "RTGS": (44, 62, 80),
        UNMAPPED_LABEL: (127, 140, 141),
    }
    navy = (15, 32, 60)

    cols = list(df.columns)
    status_cols = cols[1:-1]
    mode_rows = df[df["Mode"] != "Grand Total"].reset_index(drop=True)
    grand = df[df["Mode"] == "Grand Total"].iloc[0]
    grand_total = int(grand["Total"]) or 1

    width = 1120
    pad = 40
    header_h = 130
    gap = 30
    kpi_h = 130
    card_h = 130
    card_gap = 18
    height = header_h + gap + kpi_h + gap + 34 + len(mode_rows) * (card_h + card_gap) + 50

    base = Image.new("RGB", (width, height), (243, 245, 248))

    def shadowed_card(box, radius=16, fill="white"):
        """Draws a soft drop-shadow behind a rounded rect, then the rect itself."""
        x0, y0, x1, y1 = box
        shadow = Image.new("RGBA", base.size, (0, 0, 0, 0))
        sd = ImageDraw.Draw(shadow)
        sd.rounded_rectangle([x0, y0 + 6, x1, y1 + 6], radius=radius, fill=(15, 32, 60, 55))
        shadow = shadow.filter(ImageFilter.GaussianBlur(7))
        composited = Image.alpha_composite(base.convert("RGBA"), shadow).convert("RGB")
        base.paste(composited, (0, 0))
        ImageDraw.Draw(base).rounded_rectangle(box, radius=radius, fill=fill)

    draw = ImageDraw.Draw(base)

    # ── Header banner ──
    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 26), "RECEIPT MADE SUMMARY", font=font(30), fill="white")
    draw.text((pad, 68), "By Payment Mode & Receipt Status", font=font(14), fill=(180, 197, 222))
    draw.text(
        (pad, 92),
        f"Source: STATUS IN TAB (C/NA/X)  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        font=font(12), fill=(140, 160, 190),
    )

    # ── KPI boxes (grand totals) ──
    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", grand_total, navy)]
    kpi_items += [(sc.upper(), int(grand[sc]), status_colors.get(sc, (100, 100, 100))) for sc in status_cols]
    n_kpi = len(kpi_items)
    kpi_gap = 18
    kpi_w = (width - 2 * pad - kpi_gap * (n_kpi - 1)) // n_kpi

    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        shadowed_card(box, radius=14, fill="white")
        d = ImageDraw.Draw(base)
        d.rounded_rectangle([x + 16, y + 16, x + 42, y + 42], radius=7, fill=color)
        d.text((x + 16, y + 56), f"{val:,}", font=font(26), fill=(25, 25, 25))
        d.text((x + 16, y + 94), label, font=font(11), fill=(120, 120, 120))
        pct = "of total" if label == "TOTAL RECEIPTS" else f"{val / grand_total * 100:.1f}%"
        d.text((x + kpi_w - 62, y + 20), pct, font=font(12), fill=color)
        x += kpi_w + kpi_gap

    # ── Section label ──
    y += kpi_h + gap
    ImageDraw.Draw(base).text((pad, y), "BREAKDOWN BY PAYMENT MODE", font=font(15), fill=(70, 70, 70))
    y += 34

    # ── Per-mode cards ──
    for _, r in mode_rows.iterrows():
        mode = str(r["Mode"])
        accent = mode_colors.get(mode, (100, 100, 100))
        box = [pad, y, width - pad, y + card_h]
        shadowed_card(box, radius=16, fill="white")
        d = ImageDraw.Draw(base)

        d.rounded_rectangle([pad, y, pad + 8, y + card_h], radius=4, fill=accent)
        d.text((pad + 30, y + 18), mode, font=font(20), fill=(25, 25, 25))

        total_val = int(r["Total"])
        pill_text = f"TOTAL {total_val:,}"
        pill_w = 26 + len(pill_text) * 9
        d.rounded_rectangle([width - pad - pill_w - 20, y + 16, width - pad - 20, y + 46], radius=15, fill=navy)
        d.text((width - pad - pill_w - 6, y + 23), pill_text, font=font(13), fill="white")

        cx, cy = pad + 30, y + 62
        chip_gap = 14
        chip_w = max(140, (width - 2 * pad - 60 - chip_gap * (len(status_cols) - 1)) // len(status_cols))
        for sc in status_cols:
            val = int(r[sc])
            color = status_colors.get(sc, (100, 100, 100))
            d.rounded_rectangle([cx, cy, cx + chip_w, cy + 48], radius=10, fill=color)
            d.text((cx + 14, cy + 7), f"{val:,}", font=font(18), fill="white")
            d.text((cx + 14, cy + 29), str(sc), font=font(11), fill=(255, 255, 255))
            cx += chip_w + chip_gap

        y += card_h + card_gap

    ImageDraw.Draw(base).text(
        (pad, height - 32), "Generated automatically — DHC Working Automation", font=font(11), fill=(150, 150, 150)
    )

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def _img_font(size: int):
    """Shared helper for the PNG summary builders below: default PIL font at the given size."""
    from PIL import ImageFont
    try:
        return ImageFont.load_default(size=size)
    except TypeError:
        return ImageFont.load_default()


_TAT_COLORS = {
    "Less then 4": (39, 174, 96),
    "5 TO 10": (243, 156, 18),
    "Great then 10": (231, 76, 60),
    "Pending": (149, 165, 166),
    "Cxn": (127, 140, 141),
}


def rtgs_summary_to_image_bytes(summary: dict) -> io.BytesIO:
    """
    RTGS Intelligence dashboard PNG: navy/gold 'control tower' theme. A KPI
    strip up top (count totals per TAT bucket + total value), then zones
    ranked by volume as horizontal stacked bars (green/amber/red = TAT mix),
    and a navy leaderboard panel for online payment channels. Deliberately
    bar-chart-led — visually distinct from the card-grid used for Receipt
    Made and the heatmap used for Delay, so each export reads as its own
    dashboard rather than a re-skin of the others.
    """
    from PIL import Image, ImageDraw

    navy = (13, 27, 51)
    gold = (197, 160, 89)

    matrix = summary["matrix"]
    zones = sorted(matrix["zones"], key=lambda b: b["subtotal"]["total_count"], reverse=True)
    grand = matrix["grand_total"]
    online = summary.get("online_source_block", {"rows": [], "total": 0})

    MAX_ZONES = 10
    shown_zones = zones[:MAX_ZONES]
    extra = len(zones) - len(shown_zones)

    width = 1160
    pad = 40
    header_h = 130
    kpi_h = 110
    gap = 26
    row_h = 46
    row_gap = 10
    online_rows = online["rows"] or [("No online receipts this period", 0)]
    online_h = 50 + len(online_rows) * 32 + 16

    height = (
        header_h + gap + kpi_h + gap + 30
        + len(shown_zones) * (row_h + row_gap)
        + (26 if extra else 0)
        + 34 + gap + online_h + 60
    )

    base = Image.new("RGB", (width, height), (244, 246, 249))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 24), "RTGS INTELLIGENCE DASHBOARD", font=_img_font(30), fill="white")
    draw.text((pad, 66), "Zone-wise Turn-Around-Time Performance", font=_img_font(14), fill=(200, 210, 225))
    draw.text(
        (pad, 92),
        f"{datetime.now().strftime('%d-%b-%Y %H:%M')}  ·  {len(zones)} zones",
        font=_img_font(12), fill=gold,
    )

    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", f"{grand['total_count']:,}", navy)]
    kpi_items += [(TAT_DISPLAY[b].upper(), f"{grand['by_bucket'][b]['count']:,}", _TAT_COLORS[b]) for b in TAT_ORDER]
    kpi_items.append(("TOTAL VALUE (CR)", f"{grand['total_value']:.1f}", gold))
    n = len(kpi_items)
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (n - 1)) // n
    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        draw.rounded_rectangle(box, radius=12, fill="white", outline=(226, 232, 240))
        draw.rectangle([x, y, x + 6, y + kpi_h], fill=color)
        draw.text((x + 18, y + 18), str(val), font=_img_font(22), fill=(20, 20, 20))
        draw.text((x + 18, y + 56), label, font=_img_font(10), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    draw.text((pad, y), "ZONE PERFORMANCE — RANKED BY VOLUME", font=_img_font(14), fill=(70, 70, 70))
    y += 30

    max_zone_count = max((b["subtotal"]["total_count"] for b in shown_zones), default=1) or 1
    label_w = 190
    bar_x0 = pad + label_w
    bar_max_w = width - pad - bar_x0 - 90

    for block in shown_zones:
        zname = block["zone"]
        st = block["subtotal"]
        total = st["total_count"]
        draw.text((pad, y + 12), str(zname), font=_img_font(13), fill=(30, 30, 30))

        bar_w = int(bar_max_w * (total / max_zone_count))
        by0, by1 = y + 6, y + row_h - 6
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + bar_max_w, by1], radius=6, fill=(232, 236, 241))

        if total:
            label_font = _img_font(11)
            counts = {b: st["by_bucket"][b]["count"] for b in TAT_ORDER}
            # Proportional width first, but every non-zero bucket is floored
            # to whatever width its own number needs to actually fit —
            # otherwise a small-but-real Yellow/Red count just disappears.
            raw_w = {b: bar_w * (counts[b] / total) for b in TAT_ORDER}
            min_w = {
                b: (draw.textlength(f"{counts[b]:,}", font=label_font) + 12) if counts[b] > 0 else 0
                for b in TAT_ORDER
            }
            seg_w = {b: max(raw_w[b], min_w[b]) for b in TAT_ORDER}

            # If the floors push total width past the bar, claw the excess
            # back from whichever segment(s) have the most room to spare
            # (almost always the dominant "< 4 Days" segment), never below
            # that segment's own text-minimum.
            overflow = sum(seg_w.values()) - bar_w
            for b in sorted(TAT_ORDER, key=lambda k: seg_w[k], reverse=True):
                if overflow <= 0:
                    break
                spare = seg_w[b] - min_w[b]
                take = min(spare, overflow)
                seg_w[b] -= take
                overflow -= take

            seg_x = bar_x0
            for b in TAT_ORDER:
                w = seg_w[b]
                if w > 0:
                    draw.rectangle([seg_x, by0, seg_x + w, by1], fill=_TAT_COLORS[b])
                    draw.text(
                        (seg_x + w / 2, (by0 + by1) / 2), f"{counts[b]:,}",
                        font=label_font, fill="white", anchor="mm",
                    )
                    seg_x += w
        draw.text((bar_x0 + bar_max_w + 12, (by0 + by1) / 2), f"{total:,}", font=_img_font(14), fill=(20, 20, 20), anchor="lm")
        y += row_h + row_gap

    if extra:
        draw.text((pad, y), f"+ {extra} more zone(s) not shown", font=_img_font(11), fill=(140, 140, 140))
        y += 26

    lx = pad
    for b in TAT_ORDER:
        draw.rectangle([lx, y, lx + 14, y + 14], fill=_TAT_COLORS[b])
        draw.text((lx + 20, y - 2), TAT_DISPLAY[b], font=_img_font(11), fill=(70, 70, 70))
        lx += 150
    y += 34

    y += gap
    draw.rounded_rectangle([pad, y, width - pad, y + online_h], radius=14, fill=navy)
    draw.text((pad + 20, y + 16), "ONLINE PAYMENT CHANNELS", font=_img_font(15), fill=gold)
    ry = y + 50
    max_val = max((c for _, c in online_rows), default=1) or 1
    for name, cnt in online_rows:
        draw.text((pad + 20, ry), str(name), font=_img_font(12), fill="white")
        bw = int(240 * (cnt / max_val)) if max_val and cnt else 0
        if bw > 0:
            draw.rounded_rectangle([pad + 300, ry + 2, pad + 300 + max(bw, 2), ry + 16], radius=6, fill=gold)
        draw.text((pad + 300 + 250, ry), f"{cnt:,}", font=_img_font(12), fill="white")
        ry += 32

    draw.text((pad, height - 32), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def delay_summary_to_image_bytes(summary: dict) -> io.BytesIO:
    """
    Delay-in-RCPTING dashboard PNG: a slate/teal aging-heatmap grid (zones
    x TAT buckets, cell shade = that bucket's share of the zone's volume)
    rather than the bar-chart used for RTGS — same underlying zone x TAT
    matrix, deliberately re-visualised so the two exports don't look
    interchangeable.
    """
    from PIL import Image, ImageDraw

    slate = (30, 41, 59)
    teal = (13, 148, 136)
    matrix = summary["matrix"]
    zones = sorted(matrix["zones"], key=lambda b: b["subtotal"]["total_count"], reverse=True)
    grand = matrix["grand_total"]

    MAX_ZONES = 14
    shown = zones[:MAX_ZONES]
    extra = len(zones) - len(shown)

    width = 1160
    pad = 40
    header_h = 120
    kpi_h = 100
    gap = 26
    col_w = 220
    row_label_w = 190
    zone_total_w = 110
    row_h = 42
    header_row_h = 42
    card_pad = 18
    grid_x0 = pad + card_pad
    grid_w = row_label_w + col_w * len(TAT_ORDER) + zone_total_w

    height = (
        header_h + gap + kpi_h + gap
        + card_pad + header_row_h + len(shown) * row_h
        + (22 if extra else 0) + card_pad
        + gap + 60 + 40
    )

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=slate)
    draw.text((pad, 22), "DELAY IN RCPTING — AGING HEATMAP", font=_img_font(28), fill="white")
    draw.text((pad, 64), "Receipt processing time, Zone x TAT bucket", font=_img_font(14), fill=(180, 210, 205))
    draw.text((pad, 88), datetime.now().strftime("%d-%b-%Y %H:%M"), font=_img_font(12), fill=teal)

    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", f"{grand['total_count']:,}")]
    kpi_items += [(TAT_DISPLAY[b].upper(), f"{grand['by_bucket'][b]['count']:,}") for b in TAT_ORDER]
    n = len(kpi_items)
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (n - 1)) // n
    x = pad
    for label, val in kpi_items:
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=12, fill="white", outline=(220, 226, 232))
        draw.text((x + 16, y + 18), val, font=_img_font(22), fill=slate)
        draw.text((x + 16, y + 56), label, font=_img_font(10), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    y += kpi_h + gap

    # Card container behind the whole grid: flat shadow + white bordered panel
    card_box = [pad, y, pad + grid_w + 2 * card_pad, height - gap - 60]
    draw.rounded_rectangle(
        [card_box[0] + 3, card_box[1] + 5, card_box[2] + 3, card_box[3] + 5], radius=14, fill=(224, 227, 231)
    )
    draw.rounded_rectangle(card_box, radius=14, fill="white", outline=(210, 216, 224), width=1)

    y += card_pad

    # Header row (bordered)
    draw.rectangle([grid_x0, y, grid_x0 + grid_w, y + header_row_h], fill=(238, 241, 245))
    draw.rectangle([grid_x0, y, grid_x0 + grid_w, y + header_row_h], outline=(205, 212, 220), width=1)
    draw.text((grid_x0 + 10, y + 13), "ZONE", font=_img_font(12), fill=slate)
    cx = grid_x0 + row_label_w
    for b in TAT_ORDER:
        draw.line([cx, y, cx, y + header_row_h], fill=(205, 212, 220), width=1)
        draw.text((cx + 12, y + 13), TAT_DISPLAY[b], font=_img_font(12), fill=slate)
        cx += col_w
    draw.line([cx, y, cx, y + header_row_h], fill=(205, 212, 220), width=1)
    draw.text((cx + 12, y + 13), "ZONE TOTAL", font=_img_font(12), fill=slate)
    y += header_row_h

    def heat_color(share):
        r = int(232 - share * (232 - 13))
        g = int(240 - share * (240 - 90))
        b_ = int(240 - share * (240 - 90))
        return (r, g, b_)

    for idx, block in enumerate(shown):
        zname = block["zone"]
        st = block["subtotal"]
        total = st["total_count"] or 1
        row_bg = (250, 251, 252) if idx % 2 == 0 else "white"
        draw.rectangle([grid_x0, y, grid_x0 + row_label_w, y + row_h], fill=row_bg)
        draw.text((grid_x0 + 10, y + row_h / 2), str(zname), font=_img_font(12), fill=(30, 30, 30), anchor="lm")
        cx = grid_x0 + row_label_w
        for b in TAT_ORDER:
            cnt = st["by_bucket"][b]["count"]
            share = cnt / total
            color = heat_color(share)
            cell_box = [cx + 3, y + 3, cx + col_w - 3, y + row_h - 3]
            draw.rounded_rectangle(cell_box, radius=5, fill=color, outline=(255, 255, 255), width=2)
            txt_color = "white" if share > 0.55 else (30, 30, 30)
            draw.text(((cell_box[0] + cell_box[2]) / 2, (cell_box[1] + cell_box[3]) / 2), f"{cnt:,} ({share * 100:.0f}%)", font=_img_font(11), fill=txt_color, anchor="mm")
            cx += col_w
        draw.rectangle([cx, y, cx + zone_total_w, y + row_h], fill=(238, 241, 245))
        draw.text((cx + zone_total_w / 2, y + row_h / 2), f"{st['total_count']:,}", font=_img_font(13), fill=slate, anchor="mm")
        draw.line([grid_x0, y + row_h, grid_x0 + grid_w, y + row_h], fill=(225, 229, 234), width=1)
        y += row_h

    draw.rectangle([grid_x0, header_h + gap + kpi_h + gap + card_pad, grid_x0 + grid_w, y], outline=(210, 216, 224), width=1)

    if extra:
        draw.text((grid_x0, y + 4), f"+ {extra} more zone(s) not shown", font=_img_font(11), fill=(140, 140, 140))
        y += 22

    y = card_box[3] + gap
    draw.text((pad, y), "SHARE WITHIN ZONE:", font=_img_font(11), fill=(70, 70, 70))
    gx = pad + 150
    for i in range(40):
        share = i / 39
        draw.rectangle([gx + i * 6, y - 2, gx + i * 6 + 6, y + 14], fill=heat_color(share))
    draw.text((gx, y + 18), "LOW", font=_img_font(10), fill=(120, 120, 120))
    draw.text((gx + 40 * 6 - 30, y + 18), "HIGH", font=_img_font(10), fill=(120, 120, 120))

    draw.text((pad, height - 26), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def cash_mode_validation_summary_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Compliance alert PNG: red/amber 'violation ticket' theme when breaches
    exist — each row rendered as a stub-style ticket with a perforated
    left edge — or a clean green compliance seal when the sheet is empty.
    Deliberately unrelated visually to the other exports: this one reads
    as an alert, not a metrics dashboard.
    """
    from PIL import Image, ImageDraw

    red = (192, 57, 43)
    amber = (243, 156, 18)
    dark = (30, 30, 30)
    width = 1000

    if df.empty:
        height = 320
        base = Image.new("RGB", (width, height), (240, 248, 244))
        draw = ImageDraw.Draw(base)
        draw.rounded_rectangle([40, 40, width - 40, height - 40], radius=20, outline=(39, 174, 96), width=4)
        draw.ellipse([width / 2 - 60, 80, width / 2 + 60, 200], outline=(39, 174, 96), width=6)
        draw.line([width / 2 - 26, 140, width / 2 - 6, 162], fill=(39, 174, 96), width=6)
        draw.line([width / 2 - 6, 162, width / 2 + 34, 112], fill=(39, 174, 96), width=6)
        draw.text((width / 2, 222), "COMPLIANT", font=_img_font(28), fill=(39, 174, 96), anchor="mm")
        draw.text((width / 2, 256), "No cash-mode violations detected this period", font=_img_font(13), fill=(90, 110, 100), anchor="mm")
        draw.text((40, height - 28), f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(11), fill=(150, 150, 150))
        buf = io.BytesIO()
        base.save(buf, format="PNG")
        buf.seek(0)
        return buf

    MAX_ROWS = 12
    rows = df.head(MAX_ROWS)
    extra = len(df) - len(rows)
    total_amount = float(df["Grand Total"].sum())

    pad = 36
    header_h = 130
    kpi_h = 90
    gap = 20
    ticket_h = 62
    ticket_gap = 12
    height = header_h + gap + kpi_h + gap + 30 + len(rows) * (ticket_h + ticket_gap) + (26 if extra else 0) + 60

    base = Image.new("RGB", (width, height), (253, 245, 244))

    stripe = Image.new("RGBA", (width, header_h), red + (255,))
    sd = ImageDraw.Draw(stripe)
    for i in range(-header_h, width, 26):
        sd.line([(i, header_h), (i + header_h, 0)], fill=(255, 255, 255, 40), width=8)
    base.paste(stripe.convert("RGB"), (0, 0))
    draw = ImageDraw.Draw(base)

    draw.text((pad, 24), "⚠ CASH MODE COMPLIANCE ALERT", font=_img_font(26), fill="white")
    draw.text((pad, 64), "Customers on the disable list who paid via restricted mode", font=_img_font(13), fill=(255, 224, 220))
    draw.text((pad, 88), datetime.now().strftime("%d-%b-%Y %H:%M"), font=_img_font(11), fill=(255, 204, 198))

    y = header_h + gap
    kpi_items = [("VIOLATIONS FOUND", f"{len(df):,}", red), ("TOTAL AMOUNT", f"{total_amount:,.0f}", amber)]
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap) // 2
    x = pad
    for label, val, color in kpi_items:
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=12, fill="white", outline=(230, 200, 198))
        draw.rectangle([x, y, x + 6, y + kpi_h], fill=color)
        draw.text((x + 20, y + 18), val, font=_img_font(24), fill=dark)
        draw.text((x + 20, y + 58), label, font=_img_font(11), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    draw.text((pad, y), "VIOLATION TICKETS", font=_img_font(14), fill=(70, 40, 38))
    y += 30

    for _, r in rows.iterrows():
        box = [pad, y, width - pad, y + ticket_h]
        draw.rounded_rectangle(box, radius=10, fill="white", outline=(235, 205, 200))
        stub_x = pad + 90
        for cy in range(y, y + ticket_h, 12):
            draw.ellipse([stub_x - 4, cy - 4, stub_x + 4, cy + 4], fill=(253, 245, 244))
        draw.rectangle([pad, y, pad + 6, y + ticket_h], fill=red)
        draw.text((pad + 16, y + 10), "CIF", font=_img_font(9), fill=(150, 150, 150))
        draw.text((pad + 16, y + 24), str(r.get("CIF", "")), font=_img_font(13), fill=dark)
        draw.text((stub_x + 20, y + 8), f"{r.get('Zone2', '')} / {r.get('Sub Region2', '')}", font=_img_font(12), fill=dark)
        draw.text((stub_x + 20, y + 32), f"Slab: {r.get('Slab', '')}", font=_img_font(11), fill=(110, 110, 110))
        amt_text = f"{float(r.get('Grand Total', 0)):,.0f}"
        draw.text((width - pad - 20, y + ticket_h / 2), amt_text, font=_img_font(16), fill=red, anchor="rm")
        y += ticket_h + ticket_gap

    if extra:
        draw.text((pad, y), f"+ {extra} more violation(s) not shown — see full export", font=_img_font(11), fill=(140, 100, 98))
        y += 26

    draw.text((pad, height - 30), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def rcpt_cxn_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Cancelled-receipts register PNG: cream 'ledger' theme with a rotated
    red stamp watermark and striped register rows — styled like a formal
    paper register rather than a digital dashboard, distinct from the
    other three summary exports.
    """
    from PIL import Image, ImageDraw

    cream = (250, 246, 235)
    ink = (40, 35, 30)
    red = (176, 42, 42)
    width = 1100

    if df.empty:
        height = 320
        base = Image.new("RGB", (width, height), cream)
        draw = ImageDraw.Draw(base)
        draw.rectangle([0, 0, width, 10], fill=red)
        draw.text((width / 2, 130), "CLEAN REGISTER", font=_img_font(28), fill=(39, 120, 96), anchor="mm")
        draw.text((width / 2, 168), "No cancelled receipts this period", font=_img_font(13), fill=(100, 95, 85), anchor="mm")
        draw.text((40, height - 28), f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(11), fill=(150, 145, 135))
        buf = io.BytesIO()
        base.save(buf, format="PNG")
        buf.seek(0)
        return buf

    MAX_ROWS = 18
    rows = df.head(MAX_ROWS)
    extra = len(df) - len(rows)

    pad = 36
    header_h = 110
    row_h = 34
    col_header_h = 32
    height = header_h + col_header_h + len(rows) * row_h + (24 if extra else 0) + 70

    base = Image.new("RGB", (width, height), cream)
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=(58, 47, 38))
    draw.text((pad, 22), "CANCELLED RECEIPTS REGISTER", font=_img_font(26), fill=cream)
    draw.text((pad, 62), f"{len(df)} entries  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(13), fill=(210, 195, 175))

    stamp = Image.new("RGBA", (220, 90), (0, 0, 0, 0))
    sd = ImageDraw.Draw(stamp)
    sd.rounded_rectangle([4, 4, 216, 86], radius=10, outline=red, width=5)
    sd.text((110, 45), "CANCELLED", font=_img_font(22), fill=red, anchor="mm")
    stamp = stamp.rotate(-14, expand=True)
    base.paste(stamp, (width - pad - stamp.width + 30, 12), stamp)

    cols = ["ReceiptNo", "ReceiptDate", "AgreementNo", "CustomerName", "PaymentMode", "Zone", "Amount"]
    col_w = [110, 100, 150, 200, 130, 130, 120]
    x0 = pad
    y = header_h
    draw.rectangle([0, y, width, y + col_header_h], fill=(222, 210, 188))
    cx = x0
    for c, w in zip(cols, col_w):
        draw.text((cx + 8, y + 8), c, font=_img_font(11), fill=(70, 58, 40))
        cx += w
    y += col_header_h

    for i, (_, r) in enumerate(rows.iterrows()):
        if i % 2 == 0:
            draw.rectangle([0, y, width, y + row_h], fill=(255, 253, 246))
        cx = x0
        vals = [
            str(r.get("ReceiptNo", "")), str(r.get("ReceiptDate", "")), str(r.get("AgreementNo", "")),
            str(r.get("CustomerName", ""))[:22], str(r.get("PaymentMode", "")), str(r.get("Zone", "")),
            f"{float(r.get('Amount', 0)):,.0f}",
        ]
        for v, w in zip(vals, col_w):
            draw.text((cx + 8, y + 9), v, font=_img_font(11), fill=ink)
            cx += w
        draw.line([0, y + row_h, width, y + row_h], fill=(222, 210, 188), width=1)
        y += row_h

    if extra:
        draw.text((pad, y + 6), f"+ {extra} more entr{'y' if extra == 1 else 'ies'} not shown — see full export", font=_img_font(11), fill=(120, 110, 95))
        y += 26

    draw.text((pad, height - 30), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 145, 135))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def build_cash_mode_validation_summary(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Compliance check: customers flagged on either disable list who still
    paid in cash/Airtel, broken out by CIF and by day.
    """
    flagged = dcr_tab[
        (dcr_tab["Mode"] == "AIRTEL / CASH")
        & (dcr_tab["Ag Level cash mode"].notna() | dcr_tab["CIF LEVEL"].notna())
    ].copy()
    cols = ["CIF", "Zone2", "Sub Region2", "Slab", "Grand Total"]
    if flagged.empty:
        return pd.DataFrame(columns=cols)
    flagged["Receipt Date"] = pd.to_datetime(flagged["RECEIPT ENTER DATE"]).dt.normalize()
    pivot = pd.pivot_table(
        flagged, index=["CIF", "Zone", "Sub Region", "Slab"],
        columns="Receipt Date", values="AMOUNTPAID", aggfunc="sum",
    )
    pivot["Grand Total"] = pivot.sum(axis=1, skipna=True)
    pivot = pivot.reset_index().rename(columns={"Zone": "Zone2", "Sub Region": "Sub Region2"})
    return pivot


def build_rcpt_cxn(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Cancelled-receipt register. Structural columns are auto-filled; the
    'Remarks' column is left blank for mam to fill in — this sheet is a
    judgment log, not a formula output (see ROADMAP.md).
    """
    cxn = dcr_tab[dcr_tab["Status"] == "Cxn"].copy()
    out = pd.DataFrame({
        "ReceiptNo": cxn.get("Receipt No"),
        "ReceiptDate": pd.to_datetime(cxn["RECEIPT ENTER DATE"]).dt.strftime("%d/%m/%Y"),
        "Amount": cxn["AMOUNTPAID"],
        "ReceiptStatus": "Cancelled",
        "ReceiptType": cxn.get("Receipt Type", cxn.get("RECEIPTTYPE")),
        "PaymentMode": cxn["Mode"],
        "Zone": cxn["Zone"],
        "AgreementNo": cxn["AGREEMENTNO"],
        "CustomerName": cxn.get("PAYERNAME"),
        "ReceiptCreatedDate": pd.to_datetime(cxn["RECEIPT ENTER DATE"]).dt.strftime("%d/%m/%Y"),
        "Status": "Duplicate Receipt",
        "Remarks": "",
    })
    return out


def _stats_row(label: str, stats: dict) -> list:
    """Helper to flatten a stats dict into a row for preview DataFrame."""
    row = [label]
    for bucket in TAT_ORDER:
        row += [stats["by_bucket"][bucket]["count"], round(stats["by_bucket"][bucket]["value"], 2)]
    row += [stats["total_count"], round(stats["total_value"], 2)]
    return row


def zone_tat_matrix_to_dataframe(matrix: dict) -> pd.DataFrame:
    """Flattens the Zone x Receipt Type x TAT structure into a preview-friendly DataFrame."""
    cols = ["Zone / Receipt Type"]
    for bucket in TAT_ORDER:
        cols += [f"{TAT_DISPLAY[bucket]} Count", f"{TAT_DISPLAY[bucket]} Value"]
    cols += ["Total Count", "Total Value (Cr)"]
    rows = []
    for block in matrix["zones"]:
        rows.append(_stats_row(block["zone"], block["subtotal"]))
        for code, label, stats in block["breakdown"]:
            rows.append(_stats_row(f"   {label}", stats))
    rows.append(_stats_row("GRAND TOTAL", matrix["grand_total"]))
    return pd.DataFrame(rows, columns=cols)


def receipt_made_table_to_dataframe(table: dict, status_cols: list[str]) -> pd.DataFrame:
    """Flattens the receipt status x mode table into a preview DataFrame."""
    cols = ["Status", "Mode"] + [c.upper() for c in status_cols] + ["Grand Total"]
    rows = []
    for group in table["groups"]:
        for i, r in enumerate(group["rows"]):
            label = group["status"] if i == 0 else ""
            rows.append([label, r["mode"]] + [r["counts"][c] for c in status_cols] + [r["total"]])
    rows.append(["GRAND TOTAL", ""] + [table["grand_totals"][c] for c in status_cols] + [table["grand_total"]])
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
SUBTOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
GRANDTOTAL_FILL = PatternFill("solid", start_color="2E5395")
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=NAVY)
SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
GRANDTOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="B7C5D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COUNT_FMT = '#,##0;-#,##0;"-"'
VALUE_FMT = '#,##0.00;-#,##0.00;"-"'
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_INDENT = Alignment(horizontal="left", indent=1)


def _set(ws, row, col, value, font=BODY_FONT, fill=None, fmt=None, align=None, border=BOX):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    for j, col in enumerate(df.columns, start=1):
        col_label = col.strftime("%d-%b") if isinstance(col, pd.Timestamp) else str(col)
        _set(ws, start_row, j, col_label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        width = max(10, min(28, len(col_label) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            fmt = None
            if isinstance(val, (np.integer,)):
                val = int(val)
                fmt = COUNT_FMT
            elif isinstance(val, (np.floating, float)):
                if pd.isna(val):
                    val = None
                else:
                    val = float(val)
                    fmt = VALUE_FMT
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
                fmt = "dd-mmm-yyyy"
            _set(ws, i, j, val, fmt=fmt)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _write_zone_tat_sheet(ws, title: str, summary: dict):
    """Writes the merged-header Zone x Receipt Type x TAT pivot, with an
    online-receipt-source mini block on the left, exactly mirroring the
    layout found in the original 'RTGS Summary' / 'Delay in RCPTING
    Summary' tabs.

    Column layout is entirely derived from TAT_ORDER's length (was
    hardcoded for exactly 3 buckets, which broke — MergedCell write error
    — the moment TAT_ORDER grew to 5 with the Pending/Cxn buckets, because
    the fixed "Total Count"/"Total Value" columns (K/L) landed on top of
    cells the bucket loop had already merged)."""
    matrix = summary["matrix"]
    source_block = summary["online_source_block"]

    zone_col = 4                                   # D: "ZONE / RECEIPT TOWARDS"
    first_bucket_col = zone_col + 1                 # E
    n_buckets = len(TAT_ORDER)
    total_count_col = first_bucket_col + n_buckets * 2   # right after the last bucket's 2 cols
    total_value_col = total_count_col + 1
    last_col = total_value_col

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _set(ws, 1, 1, title, font=TITLE_FONT, fill=None, align=Alignment(horizontal="left"), border=None)

    # --- online receipt source mini-block (cols A:B) ---
    ws.merge_cells("A2:B2")
    _set(ws, 2, 1, "Online Payment — Receipt Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 1, "Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 2, "Receipt Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    r = 4
    for name, cnt in source_block["rows"]:
        _set(ws, r, 1, name, align=LEFT_INDENT)
        _set(ws, r, 2, cnt, fmt=COUNT_FMT)
        r += 1
    _set(ws, r, 1, "GRAND TOTAL", font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL)
    _set(ws, r, 2, source_block["total"], font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL, fmt=COUNT_FMT)

    # --- main Zone x Receipt Type x TAT matrix ---
    ws.merge_cells(start_row=2, start_column=zone_col, end_row=3, end_column=zone_col)
    _set(ws, 2, zone_col, "ZONE / RECEIPT TOWARDS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    col = first_bucket_col
    for bucket in TAT_ORDER:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        _set(ws, 2, col, TAT_DISPLAY[bucket], font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, col + 1, None, fill=HEADER_FILL)
        _set(ws, 3, col, "Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 3, col + 1, "Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        col += 2
    ws.merge_cells(start_row=2, start_column=total_count_col, end_row=3, end_column=total_count_col)
    ws.merge_cells(start_row=2, start_column=total_value_col, end_row=3, end_column=total_value_col)
    _set(ws, 2, total_count_col, "Total Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 2, total_value_col, "Total Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    def _write_stat_row(row, label_col, label, stats, font, fill, indent=False):
        align = LEFT_INDENT if indent else Alignment(horizontal="left")
        _set(ws, row, label_col, label, font=font, fill=fill, align=align)
        col = label_col + 1
        for bucket in TAT_ORDER:
            _set(ws, row, col, stats["by_bucket"][bucket]["count"], font=font, fill=fill, fmt=COUNT_FMT)
            _set(ws, row, col + 1, stats["by_bucket"][bucket]["value"], font=font, fill=fill, fmt=VALUE_FMT)
            col += 2
        _set(ws, row, col, stats["total_count"], font=font, fill=fill, fmt=COUNT_FMT)
        _set(ws, row, col + 1, stats["total_value"], font=font, fill=fill, fmt=VALUE_FMT)

    row = 4
    for block in matrix["zones"]:
        _write_stat_row(row, zone_col, block["zone"], block["subtotal"], SUBTOTAL_FONT, SUBTOTAL_FILL)
        row += 1
        for code, display, stats in block["breakdown"]:
            _write_stat_row(row, zone_col, display, stats, BODY_FONT, None, indent=True)
            row += 1
    _write_stat_row(row, zone_col, "GRAND TOTAL", matrix["grand_total"], GRANDTOTAL_FONT, GRANDTOTAL_FILL)

    fixed_widths = {1: 20, 2: 14, 3: 3, 4: 26, total_count_col: 12, total_value_col: 14}  # A,B,C(spacer),D,Total Count,Total Value
    for c in range(1, last_col + 1):
        if c in fixed_widths:
            width = fixed_widths[c]
        else:
            width = 9 if (c - first_bucket_col) % 2 == 0 else 12  # bucket pairs: Count(9) then Value(12)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = ws.cell(row=4, column=first_bucket_col).coordinate


def _write_receipt_made_summary_sheet(ws, summary: dict):
    def _write_table(start_col, title, status_cols, table):
        n = len(status_cols)
        last_col = start_col + 2 + n  # label cols (2) + status cols (n) + grand total (1)
        from openpyxl.utils import get_column_letter as gcl
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=last_col)
        _set(ws, 1, start_col, title, font=TITLE_FONT, align=Alignment(horizontal="left"), border=None)
        _set(ws, 2, start_col, "RECEIPT STATUS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, start_col + 1, "PAYMENT MODE", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        for k, c in enumerate(status_cols):
            _set(ws, 2, start_col + 2 + k, c.upper(), font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, last_col, "GRAND TOTAL", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

        row = 3
        for group in table["groups"]:
            first_row = row
            for r in group["rows"]:
                _set(ws, row, start_col, None)
                _set(ws, row, start_col + 1, r["mode"], align=LEFT_INDENT)
                for k, c in enumerate(status_cols):
                    _set(ws, row, start_col + 2 + k, r["counts"][c], fmt=COUNT_FMT)
                _set(ws, row, last_col, r["total"], font=SUBTOTAL_FONT, fmt=COUNT_FMT)
                row += 1
            if row > first_row:
                _set(ws, first_row, start_col, group["status"], font=SUBTOTAL_FONT, align=Alignment(horizontal="left"))
                if row - 1 > first_row:
                    ws.merge_cells(start_row=first_row, start_column=start_col, end_row=row - 1, end_column=start_col)
        _set(ws, row, start_col, "GRAND TOTAL", font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL)
        _set(ws, row, start_col + 1, None, fill=GRANDTOTAL_FILL)
        for k, c in enumerate(status_cols):
            _set(ws, row, start_col + 2 + k, table["grand_totals"][c], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        _set(ws, row, last_col, table["grand_total"], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        return last_col

    last = _write_table(1, "Receipt Made Summary — Updated / Pending",
                         ["Cleared", "Deposit", "Pending"], summary["left"])
    _write_table(last + 2, "Receipt Made Summary — Updated / Bounced or Cancelled",
                 ["Cleared", "Deposit", "Bounced", "Cxn"], summary["right"])

    for letter, width in zip("ABCDEFGHIJKLMN", [20, 16, 11, 11, 11, 13, 3, 20, 16, 11, 11, 11, 11, 13]):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A3"


def write_output_workbook(
    rtgs_summary: dict,
    delay_summary: dict,
    receipt_made_summary: dict,
    cash_mode_validation_summary: pd.DataFrame,
    rcpt_cxn: pd.DataFrame,
    extra_tabs: dict[str, pd.DataFrame] | None = None,
) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Receipt made summary")
    _write_receipt_made_summary_sheet(ws, receipt_made_summary)

    ws = wb.create_sheet("RTGS Summary")
    _write_zone_tat_sheet(ws, "RTGS Summary", rtgs_summary)

    ws = wb.create_sheet("Cash Mode Validat Summary")
    _write_df(ws, cash_mode_validation_summary)

    ws = wb.create_sheet("Delay in RCPTING Summary")
    _write_zone_tat_sheet(ws, "Delay in RCPTING Summary", delay_summary)

    ws = wb.create_sheet("RCPT CXN")
    _write_df(ws, rcpt_cxn)

    for name, df in (extra_tabs or {}).items():
        ws = wb.create_sheet(name[:31])
        _write_df(ws, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# Single-sheet Excel exports for the "Export a Specific Summary" picker —
# each reuses the same sheet-writing logic as the full workbook so the
# numbers always match, just packaged as a standalone one-sheet .xlsx.
# ---------------------------------------------------------------------------
def rtgs_summary_to_excel_bytes(summary: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RTGS Summary")
    _write_zone_tat_sheet(ws, "RTGS Summary", summary)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def delay_summary_to_excel_bytes(summary: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Delay in RCPTING Summary")
    _write_zone_tat_sheet(ws, "Delay in RCPTING Summary", summary)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def cash_mode_validation_summary_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Cash Mode Validation")
    _write_df(ws, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def rcpt_cxn_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RCPT CXN")
    _write_df(ws, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
