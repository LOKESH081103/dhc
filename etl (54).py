"""
Cash / Cheque / DD — TAT Validation & Collection Dashboards.

Standalone project, separate from the main DHC Working Automation app.
Same DCR source file, but a narrower job: flag CASH receipts not
deposited (seal challan uploaded) within 1 working day, and CHEQUE/DD
receipts not deposited within 3 working days.
"""
import html
import io
import re
from datetime import datetime

import pandas as pd

MODE_TAT = {"Cash": 1, "Cheque": 3, "DD": 3}
MODE_MATCH = {
    "Cash": ["CASH"],
    "Cheque": ["CHEQUE", "CHQ"],
    "DD": ["DD"],
}

# (display column name in the output/mail table, candidate raw-file column names to look for)
OUTPUT_COLUMNS = [
    ("RECEIPT ENTER DATE", ["RECEIPT ENTER DATE"]),
    ("Receipt No", ["Receipt No", "RECEIPT NO", "RECEIPTNO", "Receipt Number"]),
    ("AGREEMENTNO", ["AGREEMENTNO", "AGREEMENT NO", "Agreement No"]),
    ("PAYERNAME", ["PAYERNAME", "PAYER NAME"]),
    ("AMOUNTPAID", ["AMOUNTPAID", "AMOUNT PAID"]),
    ("COLLECTIONAGENTID", ["COLLECTIONAGENTID", "COLLECTION AGENT ID"]),
    ("COLLECTIONAGENTNAME", ["COLLECTIONAGENTNAME", "COLLECTION AGENT NAME"]),
    ("MODEOFPAYMENT", ["MODEOFPAYMENT", "MODE OF PAYMENT"]),
    ("RECEIPTTYPE", ["RECEIPTTYPE", "RECEIPT TYPE"]),
    ("BRANCH NAME", ["BRANCH NAME", "BRANCHNAME"]),
    ("NEW AREA", ["NEW AREA", "AREA"]),
    ("SUB REGION", ["SUB REGION", "Sub Region"]),
    ("MAIN REGION", ["MAIN REGION", "Region"]),
    ("Sub Zone", ["Sub Zone", "SUB ZONE"]),
    ("ZONE NEW", ["ZONE NEW", "ZONE", "Zone"]),
    ("SLAB", ["SLAB", "Slab"]),
]
DISPLAY_COLUMNS = ["PENDING DAYS"] + [c for c, _ in OUTPUT_COLUMNS]


def _normalize(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip()).upper()


def _find_col(df: pd.DataFrame, *candidates: str) -> str | None:
    """Case/whitespace-insensitive column lookup, with a loose contains-match fallback."""
    norm_map = {_normalize(c): c for c in df.columns}
    for cand in candidates:
        key = _normalize(cand)
        if key in norm_map:
            return norm_map[key]
    for cand in candidates:
        key = _normalize(cand)
        for norm, orig in norm_map.items():
            if key in norm or norm in key:
                return orig
    return None


def _parse_dates(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    from_serial = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    from_text = pd.to_datetime(series, errors="coerce", dayfirst=True)
    result = from_serial.where(numeric.notna(), from_text)
    return result.dt.normalize()


def load_dcr_raw(file) -> pd.DataFrame:
    """Reads Sheet1 of the uploaded DCR .xlsb, exactly as-is (no ETL transforms)."""
    df = pd.read_excel(file, sheet_name="Sheet1", engine="pyxlsb")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def filter_mode_exceeding(df: pd.DataFrame, mode: str) -> tuple[pd.DataFrame, dict]:
    """
    Filters df to the selected payment mode and "Updation Pending" status, 
    computes PENDING DAYS, and returns only rows exceeding that mode's TAT 
    — sorted worst (most overdue) first.
    """
    mop_col = _find_col(df, "MODEOFPAYMENT", "MODE OF PAYMENT")
    if mop_col is None:
        raise ValueError("Couldn't find a MODEOFPAYMENT column in this file.")
        
    receipt_enter_col = _find_col(df, "RECEIPT ENTER DATE")
    if receipt_enter_col is None:
        raise ValueError("Couldn't find a RECEIPT ENTER DATE column in this file.")

    status_col = _find_col(df, "RECEIPT STATUS", "RECEIPTSTATUS", "STATUS")
    if status_col is None:
        raise ValueError("Couldn't find a RECEIPT STATUS column in this file.")

    date_col = df.columns[0]
    if _normalize(date_col) != "DATE":
        found = _find_col(df, "Date")
        if found is not None:
            date_col = found

    match_values = [v.upper() for v in MODE_MATCH[mode]]
    mode_mask = df[mop_col].astype(str).str.strip().str.upper().isin(match_values)
    
    status_mask = df[status_col].astype(str).str.strip().str.upper() == "UPDATION PENDING"
    
    sub = df.loc[mode_mask & status_mask].copy()
    total_matched = int(len(sub))

    receipt_norm = _parse_dates(sub[receipt_enter_col])
    present_norm = _parse_dates(sub[date_col])
    present_source = f'the "{date_col}" column (report date)'

    valid = receipt_norm.notna() & present_norm.notna()
    unparseable = int((~valid).sum())
    sub = sub.loc[valid].copy()
    receipt_norm = receipt_norm.loc[valid]
    present_norm = present_norm.loc[valid]

    sub["_PENDING_DAYS"] = (present_norm - receipt_norm).dt.days
    sub[receipt_enter_col] = receipt_norm

    threshold = MODE_TAT[mode]
    exceeding = sub[sub["_PENDING_DAYS"] > threshold].copy()
    exceeding = exceeding.sort_values(["_PENDING_DAYS", receipt_enter_col], ascending=[False, True])

    missing: list[str] = []
    out = pd.DataFrame(index=exceeding.index)
    out["PENDING DAYS"] = exceeding["_PENDING_DAYS"].astype(int)
    for display_name, candidates in OUTPUT_COLUMNS:
        col = _find_col(df, *candidates)
        if col is None:
            missing.append(display_name)
            out[display_name] = "N/A"
        else:
            out[display_name] = exceeding[col]

    out["RECEIPT ENTER DATE"] = pd.to_datetime(out["RECEIPT ENTER DATE"], errors="coerce").dt.strftime("%d-%b-%y")
    out = out.reset_index(drop=True)

    meta = {
        "mode": mode,
        "threshold": threshold,
        "present_source": present_source,
        "missing_columns": missing,
        "total_matched": total_matched,
        "total_exceeding": int(len(out)),
        "unparseable_dates": unparseable,
    }
    return out, meta


def _mode_wording(mode: str) -> dict:
    if mode == "Cash":
        return {
            "item": "Cash Payments",
            "intro": (
                "Please find the below mentioned Cash Payments that have been delayed in deposit "
                "/uploading the seal challan in the system for more than one working day."
            ),
            "note1": (
                "As per Audit Concerns, it is mandatory that any cash payment is deposited and the "
                "corresponding seal challan is uploaded into the system within one working day, and "
                "these cases will be sent to FCU by Operations for review, so please act on top priority"
            ),
        }
    item = f"{mode} Payments"
    return {
        "item": item,
        "intro": (
            f"Please find the below mentioned {item} that have been delayed in deposit / updation "
            "in the system for more than three working days."
        ),
        "note1": (
            f"As per Audit Concerns, it is mandatory that any {mode.lower()} payment is deposited and "
            "updated into the system within three working days, and these cases will be sent to FCU by "
            "Operations for review, so please act on top priority"
        ),
    }


NOTE2 = "Kindly ensure that we comply with this guideline moving forward to avoid any audit discrepancies."
SIGNOFF = "Regards,\nShalini R\nLAP Collection Support \u2013 Chennai HO\n9600145482"


def build_tat_mail_html(mode: str, exceeding_df: pd.DataFrame) -> tuple[str, str]:
    e = html.escape
    w = _mode_wording(mode)
    subject = f"{w['item']} Delayed Beyond TAT \u2014 {datetime.now().strftime('%d-%b-%Y')}"

    header_html = "".join(
        f'<th style="padding:6px 10px;border:1px solid #d1d5db;background:#1e293b;'
        f'color:#ffffff;font-size:12px;text-align:left;white-space:nowrap;">{e(c)}</th>'
        for c in DISPLAY_COLUMNS
    )
    rows_html = ""
    for i, (_, row) in enumerate(exceeding_df.iterrows()):
        bg = "#ffffff" if i % 2 == 0 else "#f3f4f6"
        cells = "".join(
            f'<td style="padding:5px 10px;border:1px solid #e5e7eb;font-size:12px;'
            f'color:#111827;white-space:nowrap;">{e(str(row[c]))}</td>'
            for c in DISPLAY_COLUMNS
        )
        rows_html += f'<tr style="background:{bg};">{cells}</tr>'

    table_html = (
        '<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        f'font-family:Segoe UI,Arial,sans-serif;"><tr>{header_html}</tr>{rows_html}</table>'
    )

    body = f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111827;line-height:1.6;">
      <p>Dear Team,</p>
      <p>{e(w['intro'])}</p>
      <p>Note: {e(w['note1'])}<br/>{e(NOTE2)}</p>
      {table_html}
      <p style="margin-top:20px;">{e(SIGNOFF).replace(chr(10), '<br/>')}</p>
    </div>
    """
    return subject, body


def build_tat_mail_text(mode: str, exceeding_df: pd.DataFrame, max_rows: int = 20) -> tuple[str, str]:
    w = _mode_wording(mode)
    subject = f"{w['item']} Delayed Beyond TAT \u2014 {datetime.now().strftime('%d-%b-%Y')}"

    lines = ["Dear Team,", "", w["intro"], "", f"Note: {w['note1']}", NOTE2, ""]
    shown = exceeding_df.head(max_rows)
    for _, row in shown.iterrows():
        lines.append(
            f"- {row['Receipt No']} | {row['RECEIPT ENTER DATE']} | Pending {row['PENDING DAYS']}d | "
            f"{row['PAYERNAME']} | Rs {row['AMOUNTPAID']}"
        )
    if len(exceeding_df) > max_rows:
        lines.append(f"... and {len(exceeding_df) - max_rows} more (use 'Compose in Outlook' for the full list).")
    lines += ["", SIGNOFF]
    return subject, "\n".join(lines)


def compose_outlook_mail_html(subject: str, html_body: str) -> None:
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Display()
    finally:
        pythoncom.CoUninitialize()

# ==========================================
# NEW FEATURE: AIRTEL GATEWAY DASHBOARD
# ==========================================
# ==========================================
# NEW FEATURE: AIRTEL GATEWAY DASHBOARD
# ==========================================
# ==========================================
# NEW FEATURE: AIRTEL GATEWAY DASHBOARD
# ==========================================
# ==========================================
# NEW FEATURE: AIRTEL GATEWAY DASHBOARD
# ==========================================
def _build_pivot(df: pd.DataFrame, index_cols: list, index_names: list) -> pd.DataFrame:
    """Helper function to build a MultiIndex pivot table with Counts, Values, and Percentages."""
    pivot = pd.pivot_table(
        df,
        index=index_cols,
        columns='Dash_MOP',
        values='Amount_Num',
        aggfunc=['count', 'sum'],
        fill_value=0
    )
    
    data = {}
    for mode in ['Deposited - Airtel', 'Deposited - Bank']:
        if 'count' in pivot.columns and mode in pivot['count'].columns:
            data[(mode, 'Count')] = pivot['count'][mode]
            data[(mode, 'Value')] = pivot['sum'][mode] / 100000.0
        else:
            data[(mode, 'Count')] = pd.Series(0, index=pivot.index)
            data[(mode, 'Value')] = pd.Series(0.0, index=pivot.index)
            
    new_df = pd.DataFrame(data, index=pivot.index)
    
    new_df[('Overall', 'Count')] = new_df[('Deposited - Airtel', 'Count')] + new_df[('Deposited - Bank', 'Count')]
    new_df[('Overall', 'Value')] = new_df[('Deposited - Airtel', 'Value')] + new_df[('Deposited - Bank', 'Value')]
    
    for mode in ['Deposited - Airtel', 'Deposited - Bank']:
        new_df[(mode, 'Count %')] = (new_df[(mode, 'Count')] / new_df[('Overall', 'Count')].replace(0, 1)) * 100
        new_df[(mode, 'Value %')] = (new_df[(mode, 'Value')] / new_df[('Overall', 'Value')].replace(0, 1)) * 100
        
    ordered_cols = [
        ('Deposited - Airtel', 'Count'), ('Deposited - Airtel', 'Value'), 
        ('Deposited - Airtel', 'Count %'), ('Deposited - Airtel', 'Value %'),
        ('Deposited - Bank', 'Count'), ('Deposited - Bank', 'Value'), 
        ('Deposited - Bank', 'Count %'), ('Deposited - Bank', 'Value %'),
        ('Overall', 'Count'), ('Overall', 'Value')
    ]
    new_df = new_df[ordered_cols]
    return new_df, ordered_cols


def _grand_total_row(df: pd.DataFrame, ordered_cols: list, label_cols: dict) -> pd.DataFrame:
    """Builds a single Grand-Total row (sums + recomputed %) matching df's column structure."""
    gt = df[ordered_cols].sum()
    gt_full = pd.DataFrame([gt], columns=pd.MultiIndex.from_tuples(ordered_cols))
    for mode in ['Deposited - Airtel', 'Deposited - Bank']:
        overall = gt_full[('Overall', 'Count')].iloc[0]
        overall_v = gt_full[('Overall', 'Value')].iloc[0]
        gt_full[(mode, 'Count %')] = (gt_full[(mode, 'Count')] / overall * 100) if overall else 0
        gt_full[(mode, 'Value %')] = (gt_full[(mode, 'Value')] / overall_v * 100) if overall_v else 0
    gt_full = gt_full[ordered_cols]
    for col, val in label_cols.items():
        gt_full[col] = val
    return gt_full


# Canonical business hierarchy: which regions belong under which zone,
# and in what order they should be displayed — NOT alphabetical. Any
# region value found in the data that isn't listed here for its zone is
# still shown (appended after the known ones), never silently dropped;
# any zone value not in this list at all gets its own "Other Zones"
# group at the end, same reasoning.
ZONE_REGION_ORDER = {
    "EAST": ["West Bengal", "Chhattisgarh"],
    "NORTH": ["Delhi", "PHC", "Rajasthan", "Uttar_Pradesh", "Uttarakhand"],
    "SOUTH_1": ["Andhra_Pradesh", "Kerala", "Tamil_Nadu"],
    "SOUTH_2": ["Karnataka", "Telangana"],
    "WEST_1": ["Gujarat", "Madhya_Pradesh", "Maharashtra"],
    "WEST_2": ["Goa", "Mumbai"],
}
ZONE_ORDER = list(ZONE_REGION_ORDER.keys())

# Some zones' raw MAIN REGION values in the DCR aren't human-readable
# state names — East's region codes are recorded as "EAST_1"/"EAST_2"
# rather than the actual state names. Rename those specific known codes
# for display; every other region value passes through unchanged.
REGION_DISPLAY_RENAME = {
    "EAST_1": "West Bengal",
    "EAST_2": "Chhattisgarh",
}

# --- Region-wise table: grouped by Sub Zone (not the 4-value ZONE NEW
# column), since that's the actual granularity MAIN REGION values sit
# under in the DCR — e.g. Sub Zone "EAST_1" covers West Bengal / Odisha /
# Chhattisgarh, "EAST_2" covers Bihar / Jharkhand / North East, both
# under the same broad "EAST" ZONE NEW value. Raw MAIN REGION codes are
# inconsistent (some plain state names, some with a "_REGION" suffix, a
# couple of non-state codes like "PHC"/"ROMH", one state split into
# "RAJASTHAN_1"/"RAJASTHAN_2") so they're cleaned up for display here.
# Anything not listed — a new/unexpected code — still shows up as a
# leftover row rather than being dropped; see generate_airtel_dashboard.
SUBZONE_REGION_DISPLAY_RENAME = {
    "CHATTISGARH": "Chhattisgarh",
    "ODISHA_REGION": "Odisha",
    "WB_REGION": "West Bengal",
    "BIHAR_REGION": "Bihar",
    "JHARKHAND_REGION": "Jharkhand",
    "NE_REGION": "North East",
    "DELHI": "Delhi",
    "UTTAR_PRADESH": "Uttar Pradesh",
    "PHC": "PHC",
    "UK_REGION": "Uttarakhand",
    "RAJASTHAN_1": "Rajasthan 1",
    "RAJASTHAN_2": "Rajasthan 2",
    "ANDHRA_PRADESH": "Andhra Pradesh",
    "KERALA_REGION": "Kerala",
    "TAMIL_NADU": "Tamil Nadu",
    "KARNATAKA": "Karnataka",
    "TELANGANA": "Telangana",
    "GUJARAT": "Gujarat",
    "ROMH": "ROMH",
    "MADHYA_PRADESH": "Madhya Pradesh",
    "MUMBAI & GOA": "Mumbai & Goa",
}
SUBZONE_REGION_ORDER = {
    "EAST_1": ["West Bengal", "Odisha", "Chhattisgarh"],
    "EAST_2": ["Bihar", "Jharkhand", "North East"],
    "NORTH_1": ["Delhi", "Uttar Pradesh"],
    "NORTH_2": ["PHC", "Uttarakhand"],
    "NORTH_3": ["Rajasthan 1", "Rajasthan 2"],
    "SOUTH_1": ["Andhra Pradesh", "Kerala", "Tamil Nadu"],
    "SOUTH_2": ["Karnataka", "Telangana"],
    "WEST_1": ["Gujarat", "ROMH"],
    "WEST_2": ["Madhya Pradesh", "Mumbai & Goa"],
}
SUBZONE_ORDER = list(SUBZONE_REGION_ORDER.keys())


def generate_airtel_dashboard(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    zone_col = _find_col(df, "ZONE NEW", "ZONE", "Zone")
    region_col = _find_col(df, "MAIN REGION", "Region", "REGION")
    subzone_col = _find_col(df, "Sub Zone", "SUB ZONE")
    mop_col = _find_col(df, "MODEOFPAYMENT", "MODE OF PAYMENT")
    amount_col = _find_col(df, "AMOUNTPAID", "AMOUNT PAID")
    status_col = _find_col(df, "RECEIPT STATUS")

    if not all([zone_col, region_col, subzone_col, mop_col, amount_col, status_col]):
        raise ValueError(
            "Missing required columns for Dashboard (Zone, Sub Zone, Region, Mode of Payment, "
            "Amount, or Receipt Status)."
        )

    df_dash = df.copy()
    df_dash['Amount_Num'] = pd.to_numeric(df_dash[amount_col], errors='coerce').fillna(0)

    def categorize_mop(m):
        m_upper = str(m).strip().upper()
        if 'AIRTEL' in m_upper:
            return 'Deposited - Airtel'
        elif 'CASH' in m_upper:
            return 'Deposited - Bank'
        return 'Other'

    df_dash['Dash_MOP'] = df_dash[mop_col].apply(categorize_mop)
    is_updated = df_dash[status_col].astype(str).str.strip().str.upper() == 'UPDATED'
    df_filtered = df_dash[df_dash['Dash_MOP'].isin(['Deposited - Airtel', 'Deposited - Bank']) & is_updated].copy()

    # Clean zone/region labels: zone normalized to the canonical 6-code
    # form (matches ZONE_ORDER exactly) — ZONE NEW itself only has 4
    # broad values (EAST/NORTH/SOUTH/WEST), so for SOUTH and WEST
    # specifically, the Sub Zone code (SOUTH_1/SOUTH_2/WEST_1/WEST_2) is
    # used instead, splitting those two into separate rows; EAST/NORTH
    # pass through unchanged. Region renamed per REGION_DISPLAY_RENAME
    # where applicable, left as-is otherwise.
    df_filtered['Zone_Clean'] = df_filtered[zone_col].astype(str).str.strip().str.upper()
    raw_subzone_for_zone = df_filtered[subzone_col].astype(str).str.strip().str.upper()
    _SPLIT_SUBZONES = {"SOUTH_1", "SOUTH_2", "WEST_1", "WEST_2"}
    df_filtered['Zone_Clean'] = raw_subzone_for_zone.where(
        raw_subzone_for_zone.isin(_SPLIT_SUBZONES), df_filtered['Zone_Clean']
    )
    raw_region = df_filtered[region_col].astype(str).str.strip()
    df_filtered['Region_Clean'] = raw_region.str.upper().map(REGION_DISPLAY_RENAME).fillna(raw_region)

    zone_pivot, zone_cols = _build_pivot(df_filtered, ['Zone_Clean'], ["Zone"])

    # --- Zone-wise table: canonical order first, then any zone that
    # appears in the data but isn't in ZONE_ORDER (never dropped — just
    # appended after the known ones, before Grand Total) ---
    zones_in_data = list(zone_pivot.index)
    zone_order_full = [z for z in ZONE_ORDER if z in zones_in_data] + [z for z in zones_in_data if z not in ZONE_ORDER]
    zone_pivot = zone_pivot.reindex(zone_order_full)
    zone_pivot = zone_pivot.reset_index().rename(columns={'Zone_Clean': 'Zone'})
    zone_gt = _grand_total_row(zone_pivot, zone_cols, {'Zone': 'Grand Total'})
    zone_df = pd.concat([zone_pivot, zone_gt], ignore_index=True)

    # --- Region-wise table: grouped by Sub Zone (not ZONE NEW) — Sub
    # Zone header row first, then its regions in the specified order,
    # then any leftover regions for that sub zone, then the next sub
    # zone. The Sub Zone header row is its own subtotal pulled straight
    # from subzone_pivot, which is built from the exact same df_filtered
    # as region_pivot, so the header always sums exactly to its regions
    # underneath — not a separately-sourced number that could drift.
    df_filtered['SubZone_Clean'] = df_filtered[subzone_col].astype(str).str.strip().str.upper()
    raw_subzone_region = df_filtered[region_col].astype(str).str.strip()
    df_filtered['SubZoneRegion_Clean'] = (
        raw_subzone_region.str.upper().map(SUBZONE_REGION_DISPLAY_RENAME).fillna(raw_subzone_region)
    )

    subzone_pivot, subzone_cols = _build_pivot(df_filtered, ['SubZone_Clean'], ["Sub Zone"])
    region_pivot, region_cols = _build_pivot(df_filtered, ['SubZone_Clean', 'SubZoneRegion_Clean'], ["Sub Zone", "Region"])

    subzone_flat = subzone_pivot.reset_index().rename(columns={'SubZone_Clean': 'Sub Zone'})
    region_flat = region_pivot.reset_index().rename(columns={'SubZone_Clean': 'Sub Zone', 'SubZoneRegion_Clean': 'Region'})

    blocks = []
    subzones_present = list(dict.fromkeys(region_flat['Sub Zone']))  # preserves first-seen order as a fallback
    ordered_subzones = [sz for sz in SUBZONE_ORDER if sz in subzones_present] + [sz for sz in subzones_present if sz not in SUBZONE_ORDER]
    for subzone in ordered_subzones:
        szdf = region_flat[region_flat['Sub Zone'] == subzone]
        # Sub Zone header row = that sub zone's own subtotal from the subzone-wise pivot
        header = subzone_flat[subzone_flat['Sub Zone'] == subzone].copy()
        if not header.empty:
            header['Region'] = ''
            blocks.append(header)
        known = SUBZONE_REGION_ORDER.get(subzone, [])
        placed = set()
        for region in known:
            match = szdf[szdf['Region'].str.upper() == region.upper()]
            if not match.empty:
                m = match.copy()
                m['Region'] = region  # canonical display spelling/case
                blocks.append(m)
                placed.update(match.index)
        leftover = szdf[~szdf.index.isin(placed)]
        if not leftover.empty:
            blocks.append(leftover)
    region_df = pd.concat(blocks, ignore_index=True) if blocks else region_flat.iloc[0:0]
    # Reuse the already-correct subzone-level grand total rather than
    # summing region_df directly — region_df intentionally repeats each
    # sub zone's subtotal as its own header row above that sub zone's
    # regions, so a straight sum of region_df would double-count every
    # sub zone.
    subzone_gt = _grand_total_row(subzone_flat, subzone_cols, {'Sub Zone': 'PAN INDIA'})
    subzone_gt['Region'] = ''
    region_df = pd.concat([region_df, subzone_gt], ignore_index=True)

    return zone_df, region_df


_DASH_GREEN = "#4b5320"   # olive-green header bars + Grand Total / PAN INDIA row
_DASH_NAVY = "#1f3864"    # "Overall" column-group header
_DASH_SUBHEAD = "#dde3ee"  # light lavender sub-header row (Count / Value in lacs / ...)
_DASH_SUBHEAD_OVERALL = "#c9d6e8"  # slightly bluer sub-header under the Overall group
_DASH_PEACH = "#fbe0c4"   # region sub-row background
_DASH_BROWN = "#7a4a1e"   # region sub-row text
_DASH_BORDER = "#c9a227"  # gold border, matches the reference dashboard


def _fmt_count(v) -> str:
    return f"{v:,.0f}"


def _fmt_value(v) -> str:
    return f"{v:,.2f}"


def _fmt_pct(v, denom, decimals: int = 2) -> str:
    """'-' when the denominator is 0 (e.g. a zone with zero receipts), matching the reference dashboard."""
    return "-" if not denom else f"{v:.{decimals}f}%"


def _dash_th(text: str, colspan: int = 1, rowspan: int = 1, bg: str = _DASH_GREEN) -> str:
    rs = f' rowspan="{rowspan}"' if rowspan > 1 else ""
    cs = f' colspan="{colspan}"' if colspan > 1 else ""
    return (
        f'<th{rs}{cs} style="background:{bg};color:#ffffff;padding:8px 12px;'
        f'border:1px solid {_DASH_BORDER};font-size:12px;white-space:nowrap;">{html.escape(text)}</th>'
    )


def _dash_subhead(text: str, bg: str = _DASH_SUBHEAD) -> str:
    return (
        f'<th style="background:{bg};color:#1f2937;padding:6px 10px;border:1px solid {_DASH_BORDER};'
        f'font-size:11px;font-weight:600;white-space:nowrap;">{html.escape(text)}</th>'
    )


def _dash_metric_cells(row: pd.Series, pct_decimals: int) -> str:
    """The 10 Airtel/Bank/Overall metric <td>s shared by every data row in both tables."""
    a_count, a_val = row[("Deposited - Airtel", "Count")], row[("Deposited - Airtel", "Value")]
    a_cpct, a_vpct = row[("Deposited - Airtel", "Count %")], row[("Deposited - Airtel", "Value %")]
    b_count, b_val = row[("Deposited - Bank", "Count")], row[("Deposited - Bank", "Value")]
    b_cpct, b_vpct = row[("Deposited - Bank", "Count %")], row[("Deposited - Bank", "Value %")]
    o_count, o_val = row[("Overall", "Count")], row[("Overall", "Value")]
    return [
        _fmt_count(a_count), _fmt_value(a_val),
        _fmt_pct(a_cpct, o_count, pct_decimals), _fmt_pct(a_vpct, o_val, pct_decimals),
        _fmt_count(b_count), _fmt_value(b_val),
        _fmt_pct(b_cpct, o_count, pct_decimals), _fmt_pct(b_vpct, o_val, pct_decimals),
        _fmt_count(o_count), _fmt_value(o_val),
    ]


def _dash_td(text: str, bg: str, color: str, weight: str, align: str) -> str:
    return (
        f'<td style="background:{bg};color:{color};font-weight:{weight};text-align:{align};'
        f'padding:6px 10px;border:1px solid {_DASH_BORDER};font-size:12px;white-space:nowrap;">'
        f'{html.escape(str(text))}</td>'
    )


def _zone_table_html(zone_df: pd.DataFrame) -> str:
    """Table 1 of the reference dashboard: one row per zone, Grand Total last."""
    header_row1 = (
        _dash_th("ZONE", rowspan=2)
        + _dash_th("DEPOSITED - AIRTEL", colspan=4)
        + _dash_th("DEPOSITED - BANK", colspan=4)
        + _dash_th("OVERALL", colspan=2, bg=_DASH_NAVY)
    )
    header_row2 = "".join([
        _dash_subhead("Count"), _dash_subhead("Value in lacs"), _dash_subhead("Count%"), _dash_subhead("Value%"),
        _dash_subhead("Count"), _dash_subhead("Value in lacs"), _dash_subhead("Count%"), _dash_subhead("Value%"),
        _dash_subhead("Count", bg=_DASH_SUBHEAD_OVERALL), _dash_subhead("Value in lacs", bg=_DASH_SUBHEAD_OVERALL),
    ])

    rows_html = ""
    for _, row in zone_df.iterrows():
        zone = str(row[("Zone", "")])
        is_total = zone.strip().upper() == "GRAND TOTAL"
        bg = _DASH_GREEN if is_total else "#ffffff"
        color = "#ffffff" if is_total else "#111827"
        cells = [_dash_td(zone, bg, color, "700", "center")]
        cells += [_dash_td(v, bg, color, "700", "center") for v in _dash_metric_cells(row, pct_decimals=1)]
        rows_html += "<tr>" + "".join(cells) + "</tr>"

    return (
        f'<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        f'font-family:Segoe UI,Arial,sans-serif;border:2px solid {_DASH_BORDER};">'
        f'<tr>{header_row1}</tr><tr>{header_row2}</tr>{rows_html}</table>'
    )


def _region_table_html(region_df: pd.DataFrame) -> str:
    """
    Table 2 of the reference dashboard: each Sub Zone's own subtotal as a
    bold white header row, its regions listed (indented, peach
    background) directly underneath, then the next Sub Zone — ending in
    a dark-green PAN INDIA total.
    """
    header_row1 = (
        _dash_th("SUB ZONE/REGION", rowspan=2)
        + _dash_th("Deposited - Airtel", colspan=4)
        + _dash_th("Deposited - Bank", colspan=4)
        + _dash_th("Overall Cash Deposited", colspan=2, bg=_DASH_NAVY)
    )
    header_row2 = "".join([
        _dash_subhead("Count"), _dash_subhead("Value in lacs"), _dash_subhead("Count %"), _dash_subhead("Value %"),
        _dash_subhead("Count"), _dash_subhead("Value in lacs"), _dash_subhead("Count %"), _dash_subhead("Value %"),
        _dash_subhead("Count", bg=_DASH_SUBHEAD_OVERALL), _dash_subhead("Value in lacs", bg=_DASH_SUBHEAD_OVERALL),
    ])

    rows_html = ""
    for _, row in region_df.iterrows():
        subzone = str(row[("Sub Zone", "")])
        region = str(row[("Region", "")])
        is_total = subzone.strip().upper() == "PAN INDIA"
        is_subzone_header = (not is_total) and (region.strip() == "")

        if is_total:
            bg, color, label, align = _DASH_GREEN, "#ffffff", subzone, "center"
        elif is_subzone_header:
            bg, color, label, align = "#ffffff", "#111827", subzone, "center"
        else:
            bg, color, label, align = _DASH_PEACH, _DASH_BROWN, "\u00a0\u00a0\u00a0" + region, "left"

        cells = [_dash_td(label, bg, color, "700", align)]
        cells += [_dash_td(v, bg, color, "700", "center") for v in _dash_metric_cells(row, pct_decimals=2)]
        rows_html += "<tr>" + "".join(cells) + "</tr>"

    return (
        f'<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        f'font-family:Segoe UI,Arial,sans-serif;border:2px solid {_DASH_BORDER};">'
        f'<tr>{header_row1}</tr><tr>{header_row2}</tr>{rows_html}</table>'
    )


def build_airtel_mail_html(zone_df: pd.DataFrame, region_df: pd.DataFrame, bank_pct: float) -> tuple[str, str]:
    subject = f"Airtel Money Gateway Collection \u2014 {datetime.now().strftime('%d-%b-%Y')}"

    body = f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#111827;line-height:1.6;">
      <p>Dear Team,</p>
      <p>Please find the below Zone / region wise Cash Collection done by CFEs through CCP Mobile Application.<br>
      Deposition is Airtel Outlet Vs Bank details is shown below.<br>
      Ensure that 100% cash is being deposited only in Airtel payment bank.</p>

      <p><b>{bank_pct:.2f}%</b> of the Cash got Deposited in Bank, Please reduce the same.</p>
      <br>
      <b>Zone-wise Dashboard</b><br><br>
      {_zone_table_html(zone_df)}
      <br><br>
      <b>Region-wise Dashboard</b><br><br>
      {_region_table_html(region_df)}
      <p style="margin-top:20px;">{html.escape(SIGNOFF).replace(chr(10), '<br/>')}</p>
    </div>
    """
    return subject, body

# =====================================================================
# BRS STATUS PROCESSING — Cheque/DD and Airtel
# =====================================================================

# --- Configuration ---------------------------------------------------

# Cheque/DD: TAT buckets per screenshot (0-1, 2-4, 5-6, Above 6)
CHQ_TAT_BUCKETS   = ["0-1", "2-4", "5-6", "Above 6"]
CHQ_TAT_LABEL_COL = "CHEQUE/DD BRS Status"   # header title in Excel output

# Airtel: TAT buckets per screenshot (0-1, 2-3, 4-5, Above 5)
AIR_TAT_BUCKETS   = ["0-1", "2-3", "4-5", "Above 5"]
AIR_TAT_LABEL_COL = "Airtel Cash BRS Status"

# OPS BRS STATUS groups (top-level rows, used for merged-cell grouping)
OPS_BRS_ORDER = ["Credit Not Received", "Not Tally", "Tally"]

# Receipt-status sub-rows within each OPS BRS group
CHQ_RECEIPT_ROWS = {
    "Credit Not Received": ["Credit Not Received"],
    "Not Tally": ["Waiting for Re-Credit", "Amount Mismatch", "Challan No Mismatch"],
    "Tally": ["Delay in Deposition", "Delay in Receipting", "Nil Query"],
}
AIR_RECEIPT_ROWS = {
    "Tally": ["Delay In Receipting", "Delay In Deposition", "Nil Query"],
    "Credit Not Received": [],   # shown as single row, no sub-rows
    "Not Tally": [],
}


def _aging_to_bucket_chq(aging) -> str:
    """Map raw AGING (int) → Cheque/DD TAT bucket label."""
    try:
        a = int(float(aging))
    except (TypeError, ValueError):
        return "0-1"
    if a <= 1:
        return "0-1"
    elif a <= 4:
        return "2-4"
    elif a <= 6:
        return "5-6"
    else:
        return "Above 6"


def _aging_to_bucket_air(aging) -> str:
    """Map raw AGING (int) → Airtel TAT bucket label."""
    try:
        a = int(float(aging))
    except (TypeError, ValueError):
        return "0-1"
    if a <= 1:
        return "0-1"
    elif a <= 3:
        return "2-3"
    elif a <= 5:
        return "4-5"
    else:
        return "Above 5"


def _col(df, *candidates):
    """Case-insensitive column finder (reuses etl._find_col logic inline)."""
    return _find_col(df, *candidates)


# --- Core pivot builder ----------------------------------------------

def _brs_pivot(df: pd.DataFrame, bucket_fn, buckets: list[str]) -> dict:
    """
    Returns a nested dict:
        { ops_brs_status: { receipt_status: { bucket: count } } }
    Plus a grand-total dict { bucket: count }.
    Uses AGING column for the bucket calculation.
    """
    ops_col    = _col(df, "OPS BRS STATUS", "OPS_BRS_STATUS")
    tally_col  = _col(df, "RECEIPT STATUS\n(Tally)", "RECEIPT STAUS (Tally)",
                       "RECEIPT STATUS (Tally)", "RECEIPT STATUS")
    aging_col  = _col(df, "AGING", "Aging", "AGEING")

    if not all([ops_col, tally_col, aging_col]):
        raise ValueError(
            f"Missing required columns. Found: ops={ops_col}, tally={tally_col}, aging={aging_col}"
        )

    df = df.copy()
    df["_OPS"]    = df[ops_col].astype(str).str.strip()
    df["_TALLY"]  = df[tally_col].astype(str).str.strip()
    df["_BUCKET"] = df[aging_col].apply(bucket_fn)

    result: dict = {}
    for ops in df["_OPS"].unique():
        sub = df[df["_OPS"] == ops]
        result[ops] = {}
        for tally in sub["_TALLY"].unique():
            tsub = sub[sub["_TALLY"] == tally]
            result[ops][tally] = {b: int((tsub["_BUCKET"] == b).sum()) for b in buckets}

    grand = {b: int((df["_BUCKET"] == b).sum()) for b in buckets}
    return result, grand


# --- Excel output builder --------------------------------------------

def _brs_to_excel(pivot: dict, grand: dict, buckets: list[str],
                  ops_rows: dict, title: str, report_date: str) -> bytes:
    """
    Builds the Excel output matching the screenshot layout exactly:
    - Title row (merged, dark navy header)
    - Column headers (OPS BRS STATUS | RECEIPT STATUS (Tally) | bucket cols | Grand Total)
    - OPS BRS STATUS merged vertically across its sub-rows
    - Grand Total row at bottom (dark navy, bold)
    Returns raw bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, PatternFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "BRS Status"

    # --- Style definitions ------------------------------------------
    NAVY   = "1F3864"
    LIGHT  = "BDD7EE"
    WHITE  = "FFFFFF"
    GRAY   = "D9D9D9"
    BOLD_F = Font(bold=True, name="Arial", size=10)
    WHITE_F = Font(bold=True, name="Arial", size=10, color=WHITE)
    REG_F   = Font(name="Arial", size=10)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _cell(row, col, val="", font=None, fill_=None, align=None, border_=True):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font   = font
        if fill_:  c.fill   = fill_(fill_) if callable(fill_) else fill_
        if align:  c.alignment = align
        if border_: c.border = border
        return c

    n_buckets = len(buckets)
    total_cols = 2 + n_buckets + 1   # OPS | Tally | buckets... | Grand Total
    last_col   = total_cols

    # Row 1 — Title
    title_text = f"{title} as on {report_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font      = WHITE_F
    c.fill      = fill(NAVY)
    c.alignment = center
    c.border    = border

    # Row 2 — Column headers
    headers = ["OPS BRS STATUS", "RECEIPT STATUS\n(Tally)"] + buckets + ["Grand Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = BOLD_F
        c.fill      = fill(LIGHT)
        c.alignment = center
        c.border    = border

    # Data rows — group by OPS BRS order
    current_row = 3
    grand_row_buckets = {b: 0 for b in buckets}

    for ops_label in OPS_BRS_ORDER:
        tally_rows = ops_rows.get(ops_label, [])
        ops_data   = pivot.get(ops_label, {})

        if not tally_rows:
            # Single-row OPS with no sub-rows (Credit Not Received → direct)
            tally_rows = list(ops_data.keys()) or [ops_label]

        row_start = current_row

        for tally_label in tally_rows:
            bucket_counts = ops_data.get(tally_label, {b: 0 for b in buckets})
            row_total     = sum(bucket_counts.values())

            # RECEIPT STATUS (Tally) cell
            _cell(current_row, 2, tally_label, font=REG_F, align=center)

            # Bucket cells
            for bi, b in enumerate(buckets, 3):
                cnt = bucket_counts.get(b, 0)
                _cell(current_row, bi, cnt if cnt > 0 else "-", font=REG_F, align=center)
                if cnt > 0:
                    grand_row_buckets[b] += cnt

            # Row Grand Total
            _cell(current_row, last_col, row_total if row_total > 0 else "-",
                  font=BOLD_F, align=center)

            current_row += 1

        row_end = current_row - 1

        # OPS BRS STATUS — merged cell
        if row_end >= row_start:
            if row_end > row_start:
                ws.merge_cells(start_row=row_start, start_column=1,
                               end_row=row_end, end_column=1)
            c = ws.cell(row=row_start, column=1, value=ops_label)
            c.font      = BOLD_F
            c.fill      = fill(GRAY)
            c.alignment = center
            c.border    = border

    # Grand Total row
    grand_total = sum(grand_row_buckets.values())
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    c = ws.cell(row=current_row, column=1, value="Grand Total")
    c.font = WHITE_F; c.fill = fill(NAVY); c.alignment = center; c.border = border
    ws.cell(row=current_row, column=2).border = border

    for bi, b in enumerate(buckets, 3):
        cnt = grand.get(b, 0)
        c = ws.cell(row=current_row, column=bi, value=cnt if cnt > 0 else "-")
        c.font = WHITE_F; c.fill = fill(NAVY); c.alignment = center; c.border = border

    c = ws.cell(row=current_row, column=last_col, value=grand_total)
    c.font = WHITE_F; c.fill = fill(NAVY); c.alignment = center; c.border = border

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    for ci in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# --- Chart image builder (used in mail) ------------------------------

def _brs_chart_image(pivot: dict, grand: dict, buckets: list[str],
                      ops_rows: dict, title: str) -> bytes:
    """
    Returns a PNG (bytes) with two charts side-by-side:
    Left  — Stacked bar: OPS BRS Status breakdown (Tally / Not Tally / Credit Not Received)
    Right — Pie: TAT bucket distribution across all receipts
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
    fig.patch.set_facecolor("#f8fafc")

    # --- Left: grouped bar — receipt status totals by OPS BRS ---
    bar_colors = {
        "Credit Not Received": "#e74c3c",
        "Not Tally":           "#e67e22",
        "Tally":               "#27ae60",
    }
    ops_totals   = {}
    for ops in OPS_BRS_ORDER:
        tally_rows  = ops_rows.get(ops, [])
        ops_data    = pivot.get(ops, {})
        all_rows    = tally_rows or list(ops_data.keys())
        ops_totals[ops] = sum(
            sum(ops_data.get(r, {}).values()) for r in all_rows
        )

    labels  = [o for o in OPS_BRS_ORDER if ops_totals.get(o, 0) > 0]
    values  = [ops_totals[o] for o in labels]
    colors  = [bar_colors.get(o, "#3498db") for o in labels]
    bars    = ax1.bar(labels, values, color=colors, width=0.5, edgecolor="white", linewidth=1.2)

    for bar, val in zip(bars, values):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                 str(val), ha="center", va="bottom", fontsize=11, fontweight="bold")

    ax1.set_title(f"{title}\nOPS BRS Status Breakdown",
                  fontsize=12, fontweight="bold", pad=12, color="#1f2937")
    ax1.set_ylabel("Receipt Count", fontsize=10)
    ax1.set_facecolor("#ffffff")
    ax1.spines[["top", "right"]].set_visible(False)
    ax1.tick_params(axis="x", labelsize=9)
    ax1.yaxis.set_major_locator(plt.MaxNLocator(integer=True))

    # --- Right: pie — TAT bucket distribution ---
    pie_colors = ["#27ae60", "#f39c12", "#e74c3c", "#8e44ad",
                  "#2980b9", "#16a085", "#d35400"]
    pie_vals   = [grand.get(b, 0) for b in buckets]
    pie_labels = [f"{b}\n({v})" for b, v in zip(buckets, pie_vals)]
    non_zero   = [(l, v, c) for l, v, c in zip(pie_labels, pie_vals, pie_colors) if v > 0]

    if non_zero:
        pl, pv, pc = zip(*non_zero)
        wedges, texts, autotexts = ax2.pie(
            pv, labels=pl, colors=pc,
            autopct=lambda p: f"{p:.1f}%" if p > 2 else "",
            startangle=90, pctdistance=0.78,
            wedgeprops={"edgecolor": "white", "linewidth": 1.5},
        )
        for t in texts:
            t.set_fontsize(9)
        for at in autotexts:
            at.set_fontsize(8)
            at.set_fontweight("bold")
    else:
        ax2.text(0.5, 0.5, "No data", ha="center", va="center",
                 transform=ax2.transAxes, fontsize=12)

    ax2.set_title(f"TAT Bucket Distribution\n(Total: {sum(grand.values())})",
                  fontsize=12, fontweight="bold", pad=12, color="#1f2937")
    ax2.set_facecolor("#f8fafc")

    plt.tight_layout(pad=2.5)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


# --- Mail builder ----------------------------------------------------

def build_brs_mail_html(mode: str, pivot: dict, grand: dict,
                         buckets: list[str], ops_rows: dict,
                         report_date: str, chart_cid: str = "brs_chart") -> tuple[str, str]:
    """
    Returns (subject, HTML body) with:
    - A short intro paragraph
    - Embedded pie/bar chart (cid: reference)
    - Quick summary table (OPS BRS status x Grand Total only — the full
      detail is in the attached Excel, this is the at-a-glance view)
    """
    e = html.escape
    title = "CHEQUE/DD BRS Status" if mode == "Cheque" else "Airtel Cash BRS Status"
    subject = f"{title} as on {report_date}"

    # Quick summary table
    header = (
        f'<th style="padding:7px 14px;background:#1f3864;color:#fff;'
        f'border:1px solid #bbb;font-size:12px;">OPS BRS Status</th>'
        f'<th style="padding:7px 14px;background:#1f3864;color:#fff;'
        f'border:1px solid #bbb;font-size:12px;">Total Receipts</th>'
    )
    rows_html = ""
    grand_total = sum(grand.values())
    for ops in OPS_BRS_ORDER:
        tally_rows = ops_rows.get(ops, [])
        ops_data   = pivot.get(ops, {})
        all_rows   = tally_rows or list(ops_data.keys())
        total      = sum(sum(ops_data.get(r, {}).values()) for r in all_rows)
        pct        = f"{total/grand_total*100:.1f}%" if grand_total else "-"
        bg = "#fff" if ops != "Credit Not Received" else "#fde8e8"
        rows_html += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:6px 14px;border:1px solid #e5e7eb;font-size:12px;">'
            f'{e(ops)}</td>'
            f'<td style="padding:6px 14px;border:1px solid #e5e7eb;font-size:12px;'
            f'text-align:center;">{total} ({pct})</td></tr>'
        )
    summary_table = (
        f'<table cellspacing="0" cellpadding="0" style="border-collapse:collapse;'
        f'font-family:Segoe UI,Arial,sans-serif;margin-bottom:16px;">'
        f'<tr>{header}</tr>{rows_html}'
        f'<tr style="background:#1f3864;">'
        f'<td style="padding:6px 14px;border:1px solid #bbb;color:#fff;'
        f'font-weight:700;font-size:12px;">Grand Total</td>'
        f'<td style="padding:6px 14px;border:1px solid #bbb;color:#fff;'
        f'font-weight:700;font-size:12px;text-align:center;">{grand_total}</td></tr>'
        f'</table>'
    )

    body = f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;
                color:#111827;line-height:1.6;">
      <p>Dear Team,</p>
      <p>Please find below the <b>{e(title)} as on {e(report_date)}</b>.
         The detailed breakdown is attached as an Excel file.</p>

      {summary_table}

      <p><b>Visual Summary:</b></p>
      <img src="cid:{chart_cid}" width="780"
           style="max-width:100%;border:1px solid #e5e7eb;border-radius:8px;
                  display:block;margin-bottom:16px;"
           alt="BRS Status Chart" />

      <p style="margin-top:20px;">
        {e(SIGNOFF).replace(chr(10), '<br/>')}
      </p>
    </div>
    """
    return subject, body


# --- Public API -------------------------------------------------------

def process_brs(df: pd.DataFrame, mode: str, report_date: str) -> dict:
    """
    Entry point called from app.py.
    mode = "Cheque" or "Airtel"
    Returns dict with: pivot, grand, buckets, ops_rows, excel_bytes,
                       chart_bytes, title, report_date
    """
    if mode == "Cheque":
        pivot, grand = _brs_pivot(df, _aging_to_bucket_chq, CHQ_TAT_BUCKETS)
        buckets  = CHQ_TAT_BUCKETS
        ops_rows = CHQ_RECEIPT_ROWS
        title    = "CHEQUE/DD BRS Status"
    else:
        pivot, grand = _brs_pivot(df, _aging_to_bucket_air, AIR_TAT_BUCKETS)
        buckets  = AIR_TAT_BUCKETS
        ops_rows = AIR_RECEIPT_ROWS
        title    = "Airtel Cash BRS Status"

    excel_bytes = _brs_to_excel(pivot, grand, buckets, ops_rows, title, report_date)
    chart_bytes = _brs_chart_image(pivot, grand, buckets, ops_rows, title)

    return {
        "pivot":       pivot,
        "grand":       grand,
        "buckets":     buckets,
        "ops_rows":    ops_rows,
        "excel_bytes": excel_bytes,
        "chart_bytes": chart_bytes,
        "title":       title,
        "report_date": report_date,
    }


def load_brs_raw(file) -> pd.DataFrame:
    """Reads the BRS MIS .xlsx, stripping column whitespace."""
    df = pd.read_excel(file)
    df.columns = [str(c).strip() for c in df.columns]
    return df
