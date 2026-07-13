"""
DHC Working automation — core ETL pipeline.

Replicates, in pandas, the manual VLOOKUP/COUNTIFS/pivot workflow found inside
DHC_Working_-_Jun_26.xlsx. See ROADMAP.md for the full reverse-engineered logic
map and the open questions that need confirming with the process owner.
"""
import html
import io
import re
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Static mapping tables copied verbatim from the workbook's "Sheet3" tab.
# These are the small reference tables mam's VLOOKUPs point at.
# ---------------------------------------------------------------------------
MODE_MAP = {
    "AIRTEL": "AIRTEL", "CASH": "CASH",
    "CHEQUE": "CHEQUE", "DD": "DD",
    "ONLINE_PAYMENT": "ONLINE_PAYMENT", "RTGS": "RTGS",
}
STATUS_MAP = {"B": "Bounced", "C": "Cleared", "D": "Deposit", "NA": "Pending", "X": "Cxn"}
RECEIPT_SOURCE_MAP = {
    "BBPS": "BBPS", "CHOLAONE DIRECT": "CHOLAONE DIRECT",
    "CCP - BITLY": "CCP - BITLY", "CCP - QR": "CCP - QR", "CCP": "CCP - QR",
}
RECEIPT_TYPE_MAP = {
    "Part Payment": "Part Payment", "FC": "Settlement",
    "Sale/EMD receipt": "Settlement", "Settlement": "Settlement", "OD": "OD",
}
RECEIPT_TYPE_FALLBACK = {"OD": "OD", "OTHER OD": "OTHER OD"}

# Label used whenever a raw code doesn't match any of the maps above (typo,
# stray whitespace, different casing, a genuinely new code, or a blank
# cell). Rows are NEVER dropped/NaN'd because of an unmapped code anymore —
# they fall in here instead, so every count in the output can be traced
# back to 100% of the input rows.
UNMAPPED_LABEL = "Other/Unmapped"


def _map_with_fallback(series: pd.Series, mapping: dict, fallback: str = UNMAPPED_LABEL) -> pd.Series:
    """
    Same intent as pandas' .map(), except it never produces NaN for a
    matchable-but-messy input. It strips whitespace and upper-cases both
    sides before matching, so "cash", "Cash ", and "CASH" all resolve to
    the same mapped value. Anything that still doesn't match a known key
    (including genuinely blank cells) becomes `fallback` rather than NaN,
    so it can be counted and audited instead of silently vanishing.
    """
    norm_map = {str(k).strip().upper(): v for k, v in mapping.items()}
    normalized = series.astype(str).str.strip().str.upper()
    return normalized.map(norm_map).fillna(fallback)


EXCEL_DATE_COLUMNS = [
    "Date", "TXN DATE IN PL TAB", "RECEIPT ENTER DATE",
    "RECEIPTENTEREDTIME", "VALUEDATE", "TXNDATE",
]


def _fix_excel_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    pyxlsb (unlike openpyxl) returns Excel date cells as raw serial numbers,
    not datetimes. Convert the known date columns so downstream date-math
    behaves correctly.
    """
    for col in EXCEL_DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], unit="D", origin="1899-12-30", errors="coerce")
    return df


def tat_bucket(days):
    """
    Replicates the approximate-match VLOOKUP against Sheet3!A:B for rows
    that have a real Ageing figure. Rows whose TXN DATE IN PL TAB carried
    the "NA" (Pending) or "X" (Cxn) code instead of a date have no Ageing
    to bucket here — those are overlaid afterwards from TXN DATE CODE (see
    build_rtgs_tab / build_delay_in_rcpting_summary), so this function
    still returns None for them and callers must apply that overlay.
    """
    if pd.isna(days):
        return None
    if days < 5:
        return "Less then 4"
    if days < 11:
        return "5 TO 10"
    return "Great then 10"


def _apply_tat_code_overlay(df: pd.DataFrame) -> pd.DataFrame:
    """
    Routes TXN DATE IN PL TAB's "NA"/"X" codes into their own TAT buckets
    (Pending / Cxn) instead of leaving TAT as None, which is what caused
    these rows to be silently dropped by compute_zone_tat_matrix's
    dropna(subset=["ZONE", "TAT"]) — even though their ZONE was known and
    they belonged in the summary. No-op if TXN DATE CODE isn't present.
    """
    if "TXN DATE CODE" not in df.columns:
        return df
    df.loc[df["TXN DATE CODE"] == "NA", "TAT"] = "Pending"
    df.loc[df["TXN DATE CODE"] == "X", "TAT"] = "Cxn"
    return df


# ---------------------------------------------------------------------------
# Loaders — one per source file
# ---------------------------------------------------------------------------
def load_dcr(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """DCR.xlsb -> (raw receipts [Sheet1], agreement slab master [Sheet2])."""
    receipts = pd.read_excel(file, sheet_name="Sheet1", engine="pyxlsb")
    receipts = _fix_excel_dates(receipts)

    # STATUS IN TAB uses the literal code "NA" to mean Pending. TXN DATE IN
    # PL TAB uses the same trick for a different purpose: "NA" there means
    # no transaction date has posted yet (Pending), and "X" means the
    # transaction was cancelled (Cxn) — neither is a real date. Pandas'
    # default NA-value list treats the bare text "NA" as missing on every
    # read, so every Pending STATUS IN TAB row was silently turning blank
    # before STATUS_MAP ever saw it, AND every NA/X-coded TXN DATE IN PL
    # TAB row was silently turning into NaT — which zeroed out Ageing/TAT
    # and made the row vanish outright from every TAT summary matrix
    # (RTGS Summary, Delay in RCPTING Summary), rather than being counted
    # under a Pending/Cxn bucket. keep_default_na=False fixes that — but
    # applying it to the WHOLE sheet breaks numeric/date columns instead
    # (blank date cells stop being real NaN, forcing those columns to
    # object dtype and crashing pd.to_datetime). So: re-read ONLY these
    # columns with protection on, and patch them back in by row position.
    protected_cols = [c for c in ("STATUS IN TAB", "TXN DATE IN PL TAB") if c in receipts.columns]
    if protected_cols:
        if hasattr(file, "seek"):
            file.seek(0)
        protected = pd.read_excel(
            file, sheet_name="Sheet1", engine="pyxlsb",
            usecols=protected_cols, keep_default_na=False,
        )
        if "STATUS IN TAB" in protected_cols:
            status_col = protected["STATUS IN TAB"].replace("", np.nan)  # genuine blanks still count as missing
            receipts["STATUS IN TAB"] = status_col.values
        if "TXN DATE IN PL TAB" in protected_cols:
            txn_raw = protected["TXN DATE IN PL TAB"].astype(str).str.strip().str.upper()
            # Only "NA" and "X" are real codes. A genuine date reads back
            # here as a numeric serial string, and a genuine blank cell
            # reads back as "" — neither matches, so TXN DATE CODE stays
            # NaN for those and the normal Ageing/TAT date math (via
            # _fix_excel_dates below) is completely untouched for them.
            receipts["TXN DATE CODE"] = np.where(txn_raw.isin(["NA", "X"]), txn_raw.values, np.nan)

    master = pd.read_excel(file, sheet_name="Sheet2", engine="pyxlsb", usecols=[0, 1, 2, 3])
    master.columns = ["AGREEMENTNO", "OPENING_DPD", "OPNG_SLAB_TYPE", "OPENING_SLAB"]
    master = master.dropna(subset=["AGREEMENTNO"]).drop_duplicates("AGREEMENTNO", keep="last")
    return receipts, master


def _normalize_cif(series: pd.Series) -> pd.Series:
    """
    Coerce a CIF column to a consistent string form regardless of source:
    the Collection MIS Base reads CIF_NO as a number (int), but the
    Disable List .xlsb also reads it as a number (int/float) via pyxlsb —
    and either can come through as a differently-typed number. Merging a
    string CIF against a numeric CIF raises a hard pandas error, so every
    loader that touches CIF numbers must run through this before being
    used as a merge key.
    """
    s = series.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)  # e.g. "23174458.0" -> "23174458"
    return s


def load_disable_lists(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """To_be_Disabled.xlsb -> (cif_level_disable, agreement_level_disable).

    CIF Level Disable sheet: the Status column carries two values —
    "To be Disabled" (an actual cash restriction) and "Not Required" (the
    CIF is listed but is NOT restricted). Only "To be Disabled" rows are
    real violations; a CIF showing "Not Required" must NOT be flagged just
    because it appears in the sheet. So we filter to that status here, at
    the source — everything downstream (the CIF LEVEL merge in
    build_dcr_tab, and the notna() check in
    build_cash_mode_validation_summary) then works unchanged, since a
    "Not Required" CIF simply won't be present to match against anymore.

    Agreement Level Disble sheet: every row here IS a real restriction
    (Status is always some "> 1.95L" reason category, never a "not
    required" equivalent) — so presence alone is correctly the flag,
    no status filtering needed on this side.
    """
    cif = pd.read_excel(file, sheet_name="CIF Level Disable", engine="pyxlsb", usecols=[0, 15])
    cif.columns = ["CIF_NO", "Status"]
    cif = cif.dropna(subset=["CIF_NO"])
    cif["CIF_NO"] = _normalize_cif(cif["CIF_NO"])
    cif = cif[cif["Status"].astype(str).str.strip().str.lower() == "to be disabled"]
    agr = pd.read_excel(file, sheet_name="Agreement Level Disble", engine="pyxlsb", usecols=[0, 5])
    agr.columns = ["AGREEMENTNO", "Status"]
    agr = agr.dropna(subset=["AGREEMENTNO"])
    return cif, agr


def load_employee_mobiles(file) -> set[str]:
    """CIFCL_CBSL_List.xlsx -> set of employee/agent mobile numbers (as strings)."""
    emp = pd.read_excel(file, sheet_name=0, usecols=["Mobile Number"])
    nums = emp["Mobile Number"].dropna().astype(str).str.replace(r"\.0$", "", regex=True)
    return set(nums)


def load_collection_mis_base(file) -> pd.DataFrame:
    """
    Collection MIS Base Working (.xlsb, sheet 'Overall Base') -> AGREEMENTNO
    -> CIF_NO / Opening Slab table.

    CIF mapping source, replacing the old LAP_Delq_Dump.csv: this MIS base
    already carries CIF_NO, Opening DPD and Opening DPD SLAB for every live
    agreement, refreshed each period, so there's no need for a separate
    delinquency dump — this is now the single source of truth for CIF
    mapping.
    """
    df = pd.read_excel(
        file, sheet_name="Overall Base", engine="pyxlsb",
        usecols=["AGREEMENTNO", "CIF_NO", "Opening DPD SLAB", "Opening DPD"],
    )
    # reselect explicitly (usecols doesn't guarantee this order) before
    # renaming, so columns can't get silently swapped (e.g. CIF_NO ending
    # up mapped to the DPD column).
    df = df[["AGREEMENTNO", "CIF_NO", "Opening DPD SLAB", "Opening DPD"]]
    df.columns = ["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB"]
    df["AGREEMENTNO"] = df["AGREEMENTNO"].astype(str).str.strip()
    df["CIF_NO"] = _normalize_cif(df["CIF_NO"])
    df = df.dropna(subset=["AGREEMENTNO"])
    df = df.drop_duplicates("AGREEMENTNO", keep="last")
    return df[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB"]]


def load_receipt_cancellation_report(file) -> pd.DataFrame:
    """
    Receipt Cancellation Report (.xlsx, sheet 'Report') — the dedicated
    cancellation export, replacing the old approach of deriving cancelled
    receipts from the DCR extract's Status == 'Cxn' rows. This report
    already carries every cancelled receipt with its own real PaymentMode
    (e.g. Cashfree, Payment Gateway, QR Code, CashfreeQR, RTGS...), which
    is far more granular than the DCR's coarse Mode bucket.
    """
    df = pd.read_excel(file, sheet_name="Report")
    cols = ["ReceiptNo", "ReceiptDate", "Amount", "ReceiptStatus", "ReceiptType",
            "PaymentMode", "Zone", "AgreementNo", "CustomerName", "ReceiptCreatedDate"]
    df = df[[c for c in cols if c in df.columns]].copy()
    for date_col in ("ReceiptDate", "ReceiptCreatedDate"):
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
    return df


# ---------------------------------------------------------------------------
# Transform — build the working tabs
# ---------------------------------------------------------------------------
def build_lookup_master(dcr_master: pd.DataFrame, collection_mis_base: pd.DataFrame) -> pd.DataFrame:
    """
    For every agreement appearing in this period's DCR (Sheet2), pull its
    CIF number and Opening Slab straight from this period's Collection MIS
    Base (fresh source, not carried forward). An agreement with no match
    in the MIS base — e.g. fully closed, or too new to have appeared in
    the MIS base pull yet — is flagged for manual CIF completion instead
    of being silently left unmapped.
    """
    merged = dcr_master[["AGREEMENTNO"]].drop_duplicates().merge(
        collection_mis_base, on="AGREEMENTNO", how="left",
    )
    merged["NEEDS_CIF_MAPPING"] = merged["CIF_NO"].isna()
    return merged[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "OPENING_SLAB", "NEEDS_CIF_MAPPING"]]


def build_dcr_tab(
    receipts: pd.DataFrame,
    lookup_master: pd.DataFrame,
    agr_disable: pd.DataFrame,
    cif_disable: pd.DataFrame,
    employee_mobiles: set[str],
) -> pd.DataFrame:
    """Replicates DCR tab columns 74-84 (the 11 derived/lookup columns)."""
    df = receipts.copy()

    df = df.merge(
        lookup_master[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "NEEDS_CIF_MAPPING"]],
        on="AGREEMENTNO", how="left",
    )
    df["CIF"] = df["CIF_NO"]

    df["Unique Mob number"] = df.groupby("MOBILENO")["MOBILENO"].transform("count")
    df["Mob Num VS Emp Mob Num"] = df["MOBILENO"].astype(str).str.replace(
        r"\.0$", "", regex=True
    ).isin(employee_mobiles)

    df["Mode"] = _map_with_fallback(df["MODEOFPAYMENT"], MODE_MAP)
    df["Status"] = _map_with_fallback(df["STATUS IN TAB"], STATUS_MAP)
    df["Receipt Source"] = _map_with_fallback(df["RECEIPTSOURCE"], RECEIPT_SOURCE_MAP)

    # Audit trail: raw codes that fell through to "Other/Unmapped", so you
    # can see exactly what's driving that bucket and decide whether to add
    # it to MODE_MAP / STATUS_MAP, or whether it's genuinely junk data.
    df.attrs["unmapped_mode_codes"] = (
        df.loc[df["Mode"] == UNMAPPED_LABEL, "MODEOFPAYMENT"].value_counts().to_dict()
    )
    df.attrs["unmapped_status_codes"] = (
        df.loc[df["Status"] == UNMAPPED_LABEL, "STATUS IN TAB"].value_counts().to_dict()
    )

    # Zone / Sub Region: the raw DCR extract already carries these per-row
    # (ZONE NEW / SUB REGION columns) — no need for the CIF master here.
    # NB: "Zone" here (from ZONE NEW) is the broad 4-way field used for
    # Cash Mode Validation / RCPT CXN. The RTGS Summary and Delay in
    # RCPTING Summary pivots group by the raw "ZONE" column instead
    # (EAST/NORTH/SOUTH_1/SOUTH_2/WEST_1/WEST_2) — see
    # compute_zone_tat_matrix, confirmed against the reference workbook.
    df["Zone"] = df["ZONE NEW"]
    df["Sub Region"] = df["SUB REGION"]

    # Slab: prefer this month's fresh Sheet2 value (OPNG SLAB/SLAB already
    # in the raw extract); this matches the BT/BU formulas which re-pull
    # from DCR.xlsb's own Sheet2 rather than the stale Look Up master.
    df["Slab"] = df["SLAB"]

    df = df.merge(
        agr_disable.rename(columns={"Status": "Ag Level cash mode"}),
        on="AGREEMENTNO", how="left",
    )
    df = df.merge(
        cif_disable.rename(columns={"CIF_NO": "CIF", "Status": "CIF LEVEL"}),
        on="CIF", how="left",
    )
    return df


def build_rtgs_tab(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """Filters to RTGS-mode receipts and computes Ageing / TAT / Receipt Type."""
    rtgs = dcr_tab[dcr_tab["MODEOFPAYMENT"] == "RTGS"].copy()
    rtgs["Ageing"] = (
        pd.to_datetime(rtgs["RECEIPT ENTER DATE"]) - pd.to_datetime(rtgs["TXN DATE IN PL TAB"])
    ).dt.days
    rtgs["TAT"] = rtgs["Ageing"].apply(tat_bucket)
    rtgs = _apply_tat_code_overlay(rtgs)
    rtgs["Receipt Type"] = rtgs["RECEIPTTYPE"].map(RECEIPT_TYPE_MAP)
    rtgs["Receipt Type"] = rtgs["Receipt Type"].fillna(
        rtgs["RECEIPT CAT"].map(RECEIPT_TYPE_FALLBACK)
    )
    return rtgs


# ---------------------------------------------------------------------------
# Display constants for the pivot-style summary sheets
# ---------------------------------------------------------------------------
RECEIPT_TYPE_ORDER = ["OD", "Settlement", "Part Payment", "OTHER OD"]
RECEIPT_TYPE_DISPLAY = {
    "OD": "EMI OD/Charges",
    "Settlement": "FORECLOSURE/SETTLEMENT",
    "Part Payment": "PART PAYMENT",
    "OTHER OD": "Other OD",
}
TAT_ORDER = ["Less then 4", "5 TO 10", "Great then 10", "Pending", "Cxn"]
TAT_DISPLAY = {
    "Less then 4": "< 4 Days", "5 TO 10": "5 - 10 Days", "Great then 10": "> 10 Days",
    "Pending": "Pending", "Cxn": "Cancelled",
}
MODE_ORDER = ["AIRTEL", "CASH", "CHEQUE", "DD", "ONLINE_PAYMENT", "RTGS", UNMAPPED_LABEL]
MODE_DISPLAY = {
    "AIRTEL": "Airtel",
    "CASH": "Cash",
    "CHEQUE": "Cheque",
    "DD": "DD",
    "ONLINE_PAYMENT": "Online Payment",
    "RTGS": "RTGS",
    UNMAPPED_LABEL: UNMAPPED_LABEL,
}
# TAT limit (days) each mode is allowed before a receipt counts as
# "TAT Exceeded": Cash / Airtel / Online Payment / RTGS must be receipted
# same day or next day (Ageing <= 1); Cheque / DD get 3 days (Ageing <= 3).
TAT_LIMIT_DAYS = {
    "AIRTEL": 1, "CASH": 1, "ONLINE_PAYMENT": 1, "RTGS": 1,
    "CHEQUE": 3, "DD": 3,
}


# ---------------------------------------------------------------------------
# Summary sheet builders — each returns a plain-data structure (dicts/lists),
# kept separate from Excel writing so the numbers can be unit-tested or
# previewed in Streamlit without touching openpyxl.
# ---------------------------------------------------------------------------
def _tat_stats(df: pd.DataFrame, value_col: str = "AMOUNTPAID") -> dict:
    out = {"by_bucket": {}, "total_count": 0, "total_value": 0.0}
    for bucket in TAT_ORDER:
        sub = df[df["TAT"] == bucket]
        cnt = int(len(sub))
        val = float(sub[value_col].sum()) / 1e7
        out["by_bucket"][bucket] = {"count": cnt, "value": val}
        out["total_count"] += cnt
        out["total_value"] += val
    return out


def compute_zone_tat_matrix(
    df: pd.DataFrame,
    value_col: str = "AMOUNTPAID",
    breakdown_col: str = "Receipt Type",
    breakdown_order: list[str] = RECEIPT_TYPE_ORDER,
    breakdown_display: dict = RECEIPT_TYPE_DISPLAY,
) -> dict:
    """
    Zone x <breakdown> x TAT bucket — the structure behind both 'RTGS
    Summary' (breakdown_col='Receipt Type', the default — unchanged,
    confirmed against the reference workbook) and 'Delay in RCPTING
    Summary' (breakdown_col='Mode' — Cash/Cheque-DD/Online/RTGS instead of
    Receipt Type, since Delay Summary covers every payment mode, unlike
    RTGS Summary which is already RTGS-only).

    Zone here means the raw "ZONE" field (EAST, NORTH, SOUTH_1, SOUTH_2,
    WEST_1, WEST_2) — confirmed by reproducing the reference workbook's
    pivot row-for-row against real data. This is NOT the same as "Sub
    Zone" (over-splits EAST/NORTH into EAST_1/EAST_2/NORTH_1-3, which
    don't exist as row labels in the real pivot) nor "ZONE NEW" (too
    coarse — collapses SOUTH_1/SOUTH_2 and WEST_1/WEST_2 together).
    """
    g = df.dropna(subset=["ZONE", "TAT"]).copy()
    zones = sorted(g["ZONE"].unique())
    blocks = []
    grand = _tat_stats(g.iloc[0:0], value_col)  # zeroed template
    for zone in zones:
        zdf = g[g["ZONE"] == zone]
        subtotal = _tat_stats(zdf, value_col)
        breakdown = [
            (code, breakdown_display[code], _tat_stats(zdf[zdf[breakdown_col] == code], value_col))
            for code in breakdown_order
        ]
        blocks.append({"zone": zone, "subtotal": subtotal, "breakdown": breakdown})
        for bucket in TAT_ORDER:
            grand["by_bucket"][bucket]["count"] += subtotal["by_bucket"][bucket]["count"]
            grand["by_bucket"][bucket]["value"] += subtotal["by_bucket"][bucket]["value"]
        grand["total_count"] += subtotal["total_count"]
        grand["total_value"] += subtotal["total_value"]
    return {"zones": blocks, "grand_total": grand}


def compute_online_receipt_source_block(df: pd.DataFrame) -> dict:
    """
    Online payment (Mode == ONLINE_PAYMENT) Receipt Source leaderboard,
    narrowed to RECEIPT STATUS == "Updated" only — Bounced/Cancelled/
    Pending online receipts are excluded from the count, mirroring the
    same "Updated only" filter RTGS Summary applies. Applies to every
    source (BBPS included), not a special case for one of them.

    Receipt Source values are normalised via RECEIPT_SOURCE_MAP through
    _map_with_fallback (case/whitespace-insensitive), so raw variants
    collapse together: "CCP" and "CCP - QR" both become "CCP - QR", and
    "CCP - Bitly" / "CCP - BITLY" (any casing) both become "CCP - BITLY".
    Previously this counted the raw RECEIPTSOURCE column directly, so
    those variants were being counted as separate rows instead of one.
    """
    sub = df[(df["Mode"] == "ONLINE_PAYMENT") & (df["RECEIPT STATUS"] == "Updated")]
    normalized_source = _map_with_fallback(sub["RECEIPTSOURCE"], RECEIPT_SOURCE_MAP)
    counts = normalized_source.value_counts(dropna=True)
    rows = [(name, int(cnt)) for name, cnt in counts.items()]
    return {"rows": rows, "total": int(counts.sum())}


# The one collection-agent login the RTGS auto-receipting bot posts under.
# Every other value (a real agent ID, "customer", "lapintegrationuser",
# blank, etc.) means a person keyed the receipt in by hand.
AUTO_RECEIPT_AGENT = "indus-auto-receipt-user"


def compute_auto_manual_block(dcr_tab: pd.DataFrame) -> dict:
    """
    Every receipt (any payment mode), split by COLLECTIONAGENTNAME into
    Auto Receipt vs Manual — same matching rule as
    compute_rtgs_auto_manual_block, just not narrowed to RTGS. This is
    what actually keys the receipt in, so it applies just as much to
    Cash/Airtel/Cheque/DD/Online as it does to RTGS.
    """
    agent = dcr_tab["COLLECTIONAGENTNAME"].astype(str).str.strip().str.upper()
    is_auto = agent == AUTO_RECEIPT_AGENT.upper()
    auto_count = int(is_auto.sum())
    manual_count = int(len(dcr_tab)) - auto_count
    return {
        "rows": [("Auto Receipt", auto_count), ("Manual", manual_count)],
        "total": int(len(dcr_tab)),
    }


def compute_rtgs_auto_manual_block(dcr_tab: pd.DataFrame) -> dict:
    """RTGS-only version of compute_auto_manual_block (MODEOFPAYMENT == "RTGS")."""
    return compute_auto_manual_block(dcr_tab[dcr_tab["MODEOFPAYMENT"] == "RTGS"])


# ---------------------------------------------------------------------------
# Zone filter — lets the mail be drafted for one broad zone (North / South /
# East / West) instead of the whole DCR. Raw ZONE has messier sub-values
# than that (SOUTH_1, SOUTH_2, a stray "SOUTH 2" typo, WEST_1, WEST_2, ...),
# so broad_zone() collapses anything starting with a known broad-zone name
# down to that name, whitespace/underscore/casing variants and all.
#
# Separately, some rows (seen in the BRS MIS files) carry a "HE" (Head
# Office) tag alongside the zone code itself, e.g. "SOUTH_1 HE", "NORTH2
# HE" — that's not a 5th geographic zone, it's an orthogonal flag on top
# of the real zone, so broad_zone() deliberately leaves it alone (a "HE"
# suffix doesn't stop the prefix match — "SOUTH_1 HE" still resolves to
# "SOUTH"). Head Office rows still show up under their normal geographic
# zone. "Head Office" is offered as its own, separate, non-exclusive
# filter choice instead — see is_head_office() / filter_by_broad_zone().
# ---------------------------------------------------------------------------
BROAD_ZONES = ["NORTH", "SOUTH", "EAST", "WEST"]


def broad_zone(raw_zone) -> str:
    if pd.isna(raw_zone):
        return "UNKNOWN"
    s = str(raw_zone).strip().upper()
    for b in BROAD_ZONES:
        if s.startswith(b):
            return b
    return s


def is_head_office(raw_zone) -> bool:
    """True for a Head Office-tagged zone value: "Head Office" / "HO" / "HE"
    on their own, or a zone code with a " HE" (or "_HE"/"-HE") suffix like
    "SOUTH_1 HE". Whitespace/casing-insensitive."""
    if pd.isna(raw_zone):
        return False
    s = str(raw_zone).strip().upper()
    if s in ("HEAD OFFICE", "HO", "HE"):
        return True
    return bool(re.search(r"[\s_-]HE$", s))


def filter_by_broad_zone(df: pd.DataFrame, zone_col: str, zone: str) -> pd.DataFrame:
    """
    zone='Overall' (or falsy) returns df unchanged. zone='Head Office'
    filters to Head Office-tagged rows specifically (see is_head_office) —
    a cross-cutting filter that can include rows from any geographic zone,
    not a 5th entry in BROAD_ZONES. Otherwise filters to that broad zone
    (North/South/East/West), Head Office-tagged rows included, since the
    "HE" tag doesn't change which geographic zone a row belongs to.
    """
    if not zone or str(zone).strip().upper() == "OVERALL":
        return df
    if zone_col not in df.columns:
        return df
    if str(zone).strip().upper() == "HEAD OFFICE":
        mask = df[zone_col].apply(is_head_office)
        return df[mask].copy()
    mask = df[zone_col].apply(broad_zone) == str(zone).strip().upper()
    return df[mask].copy()


# ---------------------------------------------------------------------------
# Receipt Sync Status — RECEIPT_STATUS (Approved / Batched / Pending /
# Realized, the LMS processing stage) crossed with LMS Sync Details (Sync
# Success vs still Awaiting), for the mail's "Receipt Sync Status" image.
# ---------------------------------------------------------------------------
def compute_receipt_status_sync_block(dcr_tab: pd.DataFrame) -> dict:
    """
    Cross-tabulates RECEIPT_STATUS against LMS Sync Details. Any LMS Sync
    Details value other than "Sync Success" (Awaiting, Submitted, or
    anything else) counts as still-awaiting, since a receipt is either
    confirmed synced or it isn't yet — both fields are matched
    case/whitespace-insensitively since the raw data has variants like
    "SYNC SUCCESS" vs "Sync Success".
    """
    status_order = ["APPROVED", "BATCHED", "PENDING", "REALIZED"]
    status_raw = dcr_tab["RECEIPT_STATUS"].astype(str).str.strip().str.upper()
    sync_raw = dcr_tab["LMS Sync Details"].astype(str).str.strip().str.upper()
    is_success = sync_raw == "SYNC SUCCESS"

    totals, success, awaiting = {}, {}, {}
    for s in status_order:
        mask = status_raw == s
        totals[s] = int(mask.sum())
        success[s] = int((mask & is_success).sum())
        awaiting[s] = int((mask & ~is_success).sum())

    return {
        "status_order": status_order,
        "totals": totals,
        "success": success,
        "awaiting": awaiting,
        "success_total": sum(success.values()),
        "awaiting_total": sum(awaiting.values()),
        "grand_total": int(len(dcr_tab)),
    }


# ---------------------------------------------------------------------------
# Receipts Overview — No of Receipts / Auto / Manual (all modes) plus two
# pies: RECEIPT STATUS (Updated/Pending/Bounced-CXN) and LMS Sync Details
# (Success/Awaited) — for the mail's headline "Receipts Overview" image,
# shown first.
# ---------------------------------------------------------------------------
def compute_receipts_overview_block(dcr_tab: pd.DataFrame) -> dict:
    auto_manual = compute_auto_manual_block(dcr_tab)
    auto_count = auto_manual["rows"][0][1]
    manual_count = auto_manual["rows"][1][1]

    status_display = {"Updated": "UPDATED", "Updation Pending": "PENDING", "Bounced-or-Cancelled": "BOUNCED/CXN"}
    status_counts = dcr_tab["RECEIPT STATUS"].map(status_display).fillna(dcr_tab["RECEIPT STATUS"]).value_counts()
    receipt_status_rows = [(label, int(status_counts.get(label, 0))) for label in ["UPDATED", "PENDING", "BOUNCED/CXN"]]

    sync_raw = dcr_tab["LMS Sync Details"].astype(str).str.strip().str.upper()
    sync_success = int((sync_raw == "SYNC SUCCESS").sum())
    sync_awaited = int(len(dcr_tab)) - sync_success

    return {
        "total": int(len(dcr_tab)),
        "auto": auto_count,
        "manual": manual_count,
        "receipt_status_rows": receipt_status_rows,
        "sync_success": sync_success,
        "sync_awaited": sync_awaited,
    }


def _donut(draw, cx, cy, r_outer, r_inner, slices, colors, center_label, center_sub, font):
    """Shared multi-slice donut helper: slices=[(label,count), ...]."""
    total = sum(c for _, c in slices) or 1
    start = -90
    bbox = [cx - r_outer, cy - r_outer, cx + r_outer, cy + r_outer]
    for (label, count), color in zip(slices, colors):
        sweep = 360 * count / total
        if sweep > 0:
            draw.pieslice(bbox, start, start + sweep, fill=color)
            start += sweep
    draw.ellipse([cx - r_inner, cy - r_inner, cx + r_inner, cy + r_inner], fill="white")
    draw.text((cx, cy - 8), center_label, font=font(22), fill=(30, 30, 30), anchor="mm")
    draw.text((cx, cy + 16), center_sub, font=font(10), fill=(120, 120, 120), anchor="mm")


def receipts_overview_image_bytes(block: dict, zone_label: str = "Overall") -> io.BytesIO:
    """
    "Receipts Overview" dashboard: No of Receipts / Auto Receipts / Manual
    Receipts KPI row, then two donut charts side by side — RECEIPT STATUS
    (Updated/Pending/Bounced-CXN) and LMS Sync Details (Success/Awaited).
    Meant to be the first image in the mail, giving an at-a-glance shape
    of the day before the more detailed sections below it.
    """
    from PIL import Image, ImageDraw

    navy = (13, 27, 51)
    teal = (26, 155, 141)
    brown = (140, 84, 60)
    grey = (127, 140, 141)
    ink = (30, 30, 30)

    width = 1000
    pad = 40
    header_h = 90
    kpi_h = 90
    gap = 24
    panel_h = 300
    height = header_h + gap + kpi_h + gap + panel_h + 50

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 22), "RECEIPTS OVERVIEW", font=_img_font(24), fill="white")
    draw.text((pad, 58), f"Zone: {zone_label}  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(12), fill=(190, 200, 220))

    y = header_h + gap
    kpis = [
        ("NO OF RECEIPTS", f"{block['total']:,}", brown),
        ("AUTO RECEIPTS", f"{block['auto']:,}", teal),
        ("MANUAL RECEIPTS", f"{block['manual']:,}", grey),
    ]
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * 2) // 3
    x = pad
    for label, val, color in kpis:
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=10, fill="white", outline=(225, 225, 230))
        draw.rectangle([x, y, x + 5, y + kpi_h], fill=color)
        draw.text((x + 18, y + 16), val, font=_img_font(22), fill=ink)
        draw.text((x + 18, y + kpi_h - 26), label, font=_img_font(10), fill=(120, 120, 120))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    col_w = (width - 2 * pad - gap) // 2

    box1 = [pad, y, pad + col_w, y + panel_h]
    draw.rounded_rectangle(box1, radius=12, fill="white", outline=(225, 225, 230))
    draw.text((pad + 20, y + 18), "RECEIPT STATUS", font=_img_font(13), fill=(70, 70, 70))
    cx1, cy1 = pad + col_w // 2 - 70, y + panel_h // 2 + 15
    rows1 = block["receipt_status_rows"]
    colors1 = [teal, (230, 126, 34), (192, 57, 43)]
    _donut(draw, cx1, cy1, 85, 50, rows1, colors1, f"{block['total']:,}", "receipts", _img_font)
    lx, ly = cx1 + 85 + 30, cy1 - 46
    for (label, cnt), color in zip(rows1, colors1):
        draw.rectangle([lx, ly, lx + 14, ly + 14], fill=color)
        draw.text((lx + 22, ly - 1), label, font=_img_font(12), fill=ink)
        draw.text((lx + 22, ly + 15), f"{cnt:,}", font=_img_font(11), fill=(110, 110, 110))
        ly += 46

    box2 = [pad + col_w + gap, y, width - pad, y + panel_h]
    draw.rounded_rectangle(box2, radius=12, fill="white", outline=(225, 225, 230))
    zx0 = pad + col_w + gap + 20
    draw.text((zx0, y + 18), "LMS SYNC DETAILS", font=_img_font(13), fill=(70, 70, 70))
    cx2, cy2 = pad + col_w + gap + col_w // 2 - 70, y + panel_h // 2 + 15
    rows2 = [("Sync Success", block["sync_success"]), ("Sync Awaited", block["sync_awaited"])]
    colors2 = [teal, brown]
    _donut(draw, cx2, cy2, 85, 50, rows2, colors2, f"{block['sync_success'] + block['sync_awaited']:,}", "receipts", _img_font)
    lx, ly = cx2 + 85 + 30, cy2 - 24
    for (label, cnt), color in zip(rows2, colors2):
        draw.rectangle([lx, ly, lx + 14, ly + 14], fill=color)
        draw.text((lx + 22, ly - 1), label, font=_img_font(12), fill=ink)
        draw.text((lx + 22, ly + 15), f"{cnt:,}", font=_img_font(11), fill=(110, 110, 110))
        ly += 46

    draw.text((pad, height - 26), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def receipt_status_sync_image_bytes(block: dict, zone_label: str = "Overall") -> io.BytesIO:
    """
    "Receipt Sync Status" dashboard: Approved/Batched/Pending/Realized KPI
    row, then two panels — SYNC SUCCESS (the Approved/Batched/Realized
    rows that already synced) and SYNC AWAITING (the Batched/Pending rows
    that haven't) — each row shown as a proportion bar so it's clear how
    much of each status is still stuck.
    """
    from PIL import Image, ImageDraw

    navy = (13, 27, 51)
    status_colors = {"APPROVED": (140, 42, 42), "BATCHED": (176, 100, 60), "PENDING": (110, 100, 140), "REALIZED": (140, 120, 90)}
    success_color = (39, 139, 116)
    awaiting_color = (196, 143, 60)
    ink = (30, 30, 30)

    status_order = block["status_order"]
    width = 1000
    pad = 40
    header_h = 90
    kpi_h = 90
    gap = 24
    row_h = 40
    panel_pad = 50

    success_rows = [(s, block["success"][s]) for s in status_order if block["success"][s] > 0]
    awaiting_rows = [(s, block["awaiting"][s]) for s in status_order if block["awaiting"][s] > 0]
    panel_h = panel_pad + max(len(success_rows), len(awaiting_rows), 1) * row_h + 30

    height = header_h + gap + kpi_h + gap + panel_h + 50
    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 22), "RECEIPT SYNC STATUS", font=_img_font(24), fill="white")
    draw.text((pad, 58), f"Zone: {zone_label}  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(12), fill=(190, 200, 220))

    y = header_h + gap
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (len(status_order) - 1)) // len(status_order)
    x = pad
    for s in status_order:
        color = status_colors[s]
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=10, fill="white", outline=(225, 225, 230))
        draw.rectangle([x, y, x + 5, y + kpi_h], fill=color)
        draw.text((x + 16, y + 16), f"{block['totals'][s]:,}", font=_img_font(20), fill=ink)
        draw.text((x + 16, y + kpi_h - 24), s.title(), font=_img_font(10), fill=(120, 120, 120))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    col_w = (width - 2 * pad - gap) // 2

    def panel(x0, title, rows, total, color):
        box = [x0, y, x0 + col_w, y + panel_h]
        draw.rounded_rectangle(box, radius=12, fill="white", outline=(225, 225, 230))
        draw.rounded_rectangle([x0, y, x0 + col_w, y + 34], radius=12, fill=color)
        draw.rectangle([x0, y + 17, x0 + col_w, y + 34], fill=color)
        draw.text((x0 + col_w / 2, y + 17), f"{title}  ·  {total:,}", font=_img_font(14), fill="white", anchor="mm")
        ry = y + panel_pad
        max_cnt = max((c for _, c in rows), default=1) or 1
        for label, cnt in rows:
            draw.text((x0 + 20, ry), label.title(), font=_img_font(12), fill=ink)
            bar_x0 = x0 + 150
            bar_max_w = col_w - 150 - 90
            bar_w = int(bar_max_w * cnt / max_cnt)
            draw.rounded_rectangle([bar_x0, ry + 2, bar_x0 + bar_max_w, ry + 18], radius=6, fill=(232, 236, 241))
            if bar_w > 0:
                draw.rounded_rectangle([bar_x0, ry + 2, bar_x0 + bar_w, ry + 18], radius=6, fill=status_colors[label])
            draw.text((bar_x0 + bar_max_w + 12, ry + 2), f"{cnt:,}", font=_img_font(12), fill=ink)
            ry += row_h
        if not rows:
            draw.text((x0 + 20, ry), "Nil", font=_img_font(12), fill=(150, 150, 150))

    panel(pad, "SYNC SUCCESS", success_rows, block["success_total"], success_color)
    panel(pad + col_w + gap, "SYNC AWAITING", awaiting_rows, block["awaiting_total"], awaiting_color)

    draw.text((pad, height - 26), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf



def build_rtgs_summary(rtgs_tab: pd.DataFrame, dcr_tab: pd.DataFrame) -> dict:
    """
    The Zone x Receipt Type x TAT matrix comes from the RTGS-filtered tab,
    narrowed to RECEIPT STATUS == "Updated" — confirmed against the
    reference workbook (RTGS.xlsx): its pivot table has a report filter
    reading "RECEIPT STATUS: Updated", so Pending/Bounced-or-Cancelled RTGS
    receipts are excluded from turn-around-time performance. rtgs_tab
    itself is left untouched (still all RTGS-mode receipts, any status) —
    only this summary's matrix is narrowed, so the raw "RTGS receipts"
    count elsewhere in the app doesn't silently change.
    The 'Online Payment — Receipt Source' mini-block is a secondary KPI on
    the same dashboard and is computed from the full month's DCR data (it
    would always be empty if computed from rtgs_tab, since Mode there is
    always 'RTGS').
    """
    updated_only = rtgs_tab[rtgs_tab["RECEIPT STATUS"] == "Updated"]
    return {
        "matrix": compute_zone_tat_matrix(updated_only),
        "online_source_block": compute_online_receipt_source_block(dcr_tab),
    }


def _dcr_with_ageing(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    dcr_tab (every payment mode, every status) plus Ageing (days between
    txn and receipt-entry) and the old day-range TAT bucket, with the
    Pending/Cxn overlay applied. Shared by build_delay_in_rcpting_summary,
    compute_tat_exceeded, and build_zone_overview_summary so Ageing is
    computed the same way everywhere instead of three separate copies of
    the same date-math drifting apart.
    """
    full = dcr_tab.copy()
    full["Ageing"] = (
        pd.to_datetime(full["RECEIPT ENTER DATE"]) - pd.to_datetime(full["TXN DATE IN PL TAB"])
    ).dt.days
    full["TAT"] = full["Ageing"].apply(tat_bucket)
    full = _apply_tat_code_overlay(full)
    return full


def compute_tat_exceeded(full: pd.DataFrame) -> pd.Series:
    """
    Per-row TAT-exceeded flag using mode-specific limits (TAT_LIMIT_DAYS)
    instead of one fixed day-range for every mode: Cash / Airtel / Online
    Payment / RTGS must be receipted same day or next day (Ageing <= 1);
    Cheque / DD get 3 days (Ageing <= 3). Rows with no Ageing yet
    (Pending/Cancelled — no RECEIPT ENTER DATE) aren't counted as
    exceeded, since there's no completed turnaround to measure; an
    unmapped/unknown mode has no configured limit and is also excluded
    rather than guessed at.

    `full` must already have an "Ageing" column — pass the output of
    _dcr_with_ageing().
    """
    limit = full["Mode"].map(TAT_LIMIT_DAYS)
    ageing = pd.to_numeric(full["Ageing"], errors="coerce")
    return (ageing.notna() & limit.notna() & (ageing > limit)).fillna(False)


def build_zone_overview_summary(dcr_tab: pd.DataFrame) -> dict:
    """
    Zone-wise overview feeding the "Zone Overview" dashboard image:
    per-zone receipt count for every payment mode, per-zone TAT-exceeded
    count (mode-specific limits, see compute_tat_exceeded), and per-zone
    Bounced / Cancelled counts. One row per zone, sorted by total volume,
    plus a grand-total row — same shape of aggregation as
    compute_zone_tat_matrix but keyed on these specific figures instead
    of the day-bucket matrix.
    """
    full = _dcr_with_ageing(dcr_tab)
    full["TAT Exceeded"] = compute_tat_exceeded(full)
    g = full.dropna(subset=["ZONE"]).copy()

    modes = [m for m in MODE_ORDER if m in g["Mode"].unique()]

    def _zone_row(zdf: pd.DataFrame) -> dict:
        return {
            "by_mode": {m: int((zdf["Mode"] == m).sum()) for m in modes},
            "tat_exceeded": int(zdf["TAT Exceeded"].sum()),
            "bounced": int((zdf["Status"] == "Bounced").sum()),
            "cancelled": int((zdf["Status"] == "Cxn").sum()),
            "total": int(len(zdf)),
        }

    zones = []
    for zone in sorted(g["ZONE"].unique()):
        zones.append({"zone": zone, **_zone_row(g[g["ZONE"] == zone])})
    zones.sort(key=lambda z: z["total"], reverse=True)
    grand = _zone_row(g)
    return {"zones": zones, "modes": modes, "grand_total": grand}


def build_delay_in_rcpting_summary(dcr_tab: pd.DataFrame) -> dict:
    """
    Zone x Payment Mode x TAT — deliberately Payment Mode (Airtel/Cash/
    Cheque/DD/Online Payment/RTGS) rather than Receipt Type here, since
    this summary covers every payment mode (unlike RTGS Summary, which is
    RTGS-only and keeps the Receipt Type breakdown).
    """
    full = _dcr_with_ageing(dcr_tab)
    return {
        "matrix": compute_zone_tat_matrix(
            full, breakdown_col="Mode", breakdown_order=MODE_ORDER, breakdown_display=MODE_DISPLAY,
        ),
        "online_source_block": compute_online_receipt_source_block(full),
    }


def _status_table(df: pd.DataFrame, status_groups: list[str], status_cols: list[str]) -> dict:
    """
    RECEIPT STATUS (row group) x Mode (row) x Status (column) counts,
    used for both halves of 'Receipt made summary'. Uses the full set of
    modes present anywhere in `df` (not just within each status
    subgroup), so every group lists the same consistent rows — matching
    the reference report where AIRTEL/CASH, CHQ/DD, ONLINE_PAYMENT and
    RTGS each appear under every RECEIPT STATUS group, with 0/blank cells
    shown rather than the row disappearing entirely.
    """
    all_modes = [m for m in MODE_ORDER if m in df["Mode"].unique()]
    all_modes += [m for m in df["Mode"].unique() if m not in all_modes]

    groups = []
    grand = {c: 0 for c in status_cols}
    grand_total = 0
    for status in status_groups:
        sdf = df[df["RECEIPT STATUS"] == status]
        if sdf.empty:
            continue
        rows = []
        group_totals = {c: 0 for c in status_cols}
        for mode in all_modes:
            mdf = sdf[sdf["Mode"] == mode]
            counts = {c: int((mdf["Status"] == c).sum()) for c in status_cols}
            row_total = sum(counts.values())
            rows.append({"mode": mode, "counts": counts, "total": row_total})
            for c in status_cols:
                group_totals[c] += counts[c]
                grand[c] += counts[c]
            grand_total += row_total
        groups.append({"status": status, "rows": rows, "totals": group_totals})
    return {"groups": groups, "grand_totals": grand, "grand_total": grand_total}


def build_receipt_made_summary(dcr_tab: pd.DataFrame) -> dict:
    """
    Two views, both grouped from the same RECEIPT STATUS field:
    left = Updated/Pending x (Cleared, Deposited, Pending);
    right = Updated/Bounced-or-Cancelled x (Cleared, Deposited, Pending,
    Bounced, Cxn).

    'right' used to omit the Pending column (only Cleared/Deposited/
    Bounced/Cxn) — any receipt whose Status IN TAB was "NA" (Pending)
    under Updated or Bounced-or-Cancelled was silently excluded from that
    table's counts entirely (not just uncategorized — genuinely dropped
    from every row/grand total), since _status_table only counts a row
    against the status_cols it's given. Pending is now included here too,
    so nothing gets dropped.
    Used for the "Receipt Made Summary" export.
    """
    status_relabel = {
        "Updated": "UPDATED",
        "Updation Pending": "PENDING",
        "Bounced-or-Cancelled": "BOUNCED/CANCELLED",
    }
    d = dcr_tab.copy()
    d["RECEIPT STATUS"] = d["RECEIPT STATUS"].map(status_relabel).fillna(d["RECEIPT STATUS"])
    left = _status_table(d, ["UPDATED", "PENDING"], ["Cleared", "Deposit", "Pending"])
    right = _status_table(d, ["UPDATED", "BOUNCED/CANCELLED"], _RECEIPT_MADE_STATUS_COLS)
    # "full": all three RECEIPT STATUS groups together with all 5 status
    # columns — the single reconciled view used by the "Receipt Made
    # Summary" export (receipt_made_status_table_to_dataframe /
    # receipt_made_status_summary_to_excel_bytes / _image_bytes all expect
    # this exact {groups, grand_totals, grand_total} shape). left/right
    # split the same data two ways for the on-screen tabs; full is the
    # export's single source of truth.
    full = _status_table(d, ["UPDATED", "PENDING", "BOUNCED/CANCELLED"], _RECEIPT_MADE_STATUS_COLS)
    return {"left": left, "right": right, "full": full}


_RECEIPT_MADE_STATUS_COLS = ["Cleared", "Deposit", "Pending", "Bounced", "Cxn"]
_RECEIPT_MADE_STATUS_LABELS = {"Cleared": "CLEARED", "Deposit": "DEPOSITED", "Pending": "PENDING", "Bounced": "BOUNCED", "Cxn": "CXN"}


def _visible_rows(group: dict) -> list:
    """
    Mode rows to actually display for a RECEIPT STATUS group — drops rows
    whose total is 0 (e.g. ONLINE_PAYMENT / RTGS under PENDING or
    BOUNCED/CANCELLED when nothing fell into that mode this period).
    Purely a display filter: grand_totals/grand_total are computed over
    every row before this runs, so hiding an empty row here never changes
    any total anywhere.
    """
    return [r for r in group["rows"] if r["total"] > 0]


def receipt_made_status_table_to_dataframe(table: dict) -> pd.DataFrame:
    """
    Flattens the 'Receipt Made Report' grouped structure (RECEIPT STATUS x
    Payment Mode x Cleared/Deposited/Pending/Bounced/CXN) into a plain
    DataFrame for the Streamlit preview — Receipt Status repeats per mode
    row rather than being visually merged, since st.dataframe can't merge
    cells, but the numbers and row order match the reference report
    exactly. Zero counts display as "-" (matching the Excel/image
    convention for cells with no data), while Grand Total stays numeric.
    """
    rows = []
    for group in table["groups"]:
        for r in _visible_rows(group):
            row = {"Receipt Status": group["status"], "Payment Mode": r["mode"]}
            for c in _RECEIPT_MADE_STATUS_COLS:
                val = r["counts"][c]
                row[_RECEIPT_MADE_STATUS_LABELS[c]] = val if val else "-"
            row["Grand Total"] = r["total"]
            rows.append(row)
    cols = ["Receipt Status", "Payment Mode"] + [_RECEIPT_MADE_STATUS_LABELS[c] for c in _RECEIPT_MADE_STATUS_COLS] + ["Grand Total"]
    df = pd.DataFrame(rows, columns=cols)
    grand = {"Receipt Status": "GRAND TOTAL", "Payment Mode": ""}
    for c in _RECEIPT_MADE_STATUS_COLS:
        val = table["grand_totals"][c]
        grand[_RECEIPT_MADE_STATUS_LABELS[c]] = val if val else "-"
    grand["Grand Total"] = table["grand_total"]
    return pd.concat([df, pd.DataFrame([grand])], ignore_index=True)


def receipt_made_status_summary_to_excel_bytes(table: dict) -> io.BytesIO:
    """
    Single-sheet .xlsx replicating the reference 'Receipt Made Report'
    layout exactly: RECEIPT STATUS row groups (merged vertically) x
    Payment Mode rows x Cleared/Deposited/Bounced/CXN columns, dark-green
    headers, light-blue sub-header row, and a gold GRAND TOTAL row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt Made Report"

    dark_green = PatternFill("solid", fgColor="1F3D2B")
    light_blue = PatternFill("solid", fgColor="D9E1F2")
    gold = PatternFill("solid", fgColor="F0AD3D")
    white_bold = Font(bold=True, color="FFFFFF", size=11)
    header_sub_font = Font(bold=True, color="1F3D2B", size=10)
    group_font = Font(bold=True, size=10)
    total_font = Font(bold=True, size=10)
    thin = Side(style="thin", color="9AA5AD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_indent = Alignment(horizontal="left", vertical="center", indent=1)

    status_cols = _RECEIPT_MADE_STATUS_COLS
    last_col = 2 + len(status_cols) + 1  # A=status, B=mode, then status_cols, then grand total

    ws.append(["Receipt Made Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    h1, h2 = 3, 4

    ws.merge_cells(start_row=h1, start_column=1, end_row=h2, end_column=1)
    ws.cell(h1, 1, "RECEIPT STATUS").font = white_bold
    ws.cell(h1, 1).alignment = center
    ws.cell(h1, 1).fill = dark_green
    ws.cell(h2, 1).fill = dark_green

    ws.merge_cells(start_row=h1, start_column=2, end_row=h2, end_column=2)
    ws.cell(h1, 2, "PAYMENT MODE").font = white_bold
    ws.cell(h1, 2).alignment = center
    ws.cell(h1, 2).fill = dark_green
    ws.cell(h2, 2).fill = dark_green

    ws.merge_cells(start_row=h1, start_column=3, end_row=h1, end_column=2 + len(status_cols))
    ws.cell(h1, 3, "RECEIPT STATUS").font = white_bold
    ws.cell(h1, 3).alignment = center
    for col in range(3, 2 + len(status_cols) + 1):
        ws.cell(h1, col).fill = dark_green

    for k, c in enumerate(status_cols):
        cell = ws.cell(h2, 3 + k, _RECEIPT_MADE_STATUS_LABELS[c])
        cell.font = header_sub_font
        cell.fill = light_blue
        cell.alignment = center

    ws.merge_cells(start_row=h1, start_column=last_col, end_row=h2, end_column=last_col)
    ws.cell(h1, last_col, "GRAND TOTAL").font = white_bold
    ws.cell(h1, last_col).alignment = center
    ws.cell(h1, last_col).fill = dark_green
    ws.cell(h2, last_col).fill = dark_green

    row = h2 + 1
    for group in table["groups"]:
        first_row = row
        for r in _visible_rows(group):
            ws.cell(row, 2, r["mode"]).alignment = left_indent
            for k, c in enumerate(status_cols):
                val = r["counts"][c]
                cell = ws.cell(row, 3 + k, val)
                cell.alignment = center
                cell.number_format = '#,##0;-#,##0;"-"'
            tot_cell = ws.cell(row, last_col, r["total"])
            tot_cell.font = total_font
            tot_cell.alignment = center
            tot_cell.number_format = "#,##0"
            row += 1
        if row > first_row:
            gcell = ws.cell(first_row, 1, group["status"])
            gcell.font = group_font
            gcell.alignment = left_indent
            if row - 1 > first_row:
                ws.merge_cells(start_row=first_row, start_column=1, end_row=row - 1, end_column=1)

    ws.cell(row, 1, "GRAND TOTAL").font = Font(bold=True, color="1F3D2B")
    ws.cell(row, 1).fill = gold
    ws.cell(row, 2).fill = gold
    for k, c in enumerate(status_cols):
        cell = ws.cell(row, 3 + k, table["grand_totals"][c])
        cell.font = Font(bold=True)
        cell.fill = gold
        cell.alignment = center
        cell.number_format = "#,##0"
    gt_cell = ws.cell(row, last_col, table["grand_total"])
    gt_cell.font = Font(bold=True)
    gt_cell.fill = gold
    gt_cell.alignment = center
    gt_cell.number_format = "#,##0"

    for r in ws.iter_rows(min_row=h1, max_row=row, min_col=1, max_col=last_col):
        for cell in r:
            cell.border = border

    widths = [20, 18] + [12] * len(status_cols) + [14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"C{h2 + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def receipt_made_status_summary_to_image_bytes(table: dict) -> io.BytesIO:
    """
    Dashboard-style PNG for the Receipt Made Report, grouped exactly like
    the reference: a KPI strip (grand totals) up top, then one coloured
    section per RECEIPT STATUS group (Updated / Bounced-or-Cancelled),
    each containing a soft-shadowed card per payment mode with
    Cleared/Deposited/Bounced/CXN chips — zero-count chips render muted
    grey (a dash), matching the report's blank cells. Same card-dashboard
    visual language as the other summaries, just restructured around the
    RECEIPT STATUS grouping instead of a flat mode list.
    """
    from PIL import Image, ImageDraw, ImageFont, ImageFilter

    font = _img_font

    status_colors = {
        "Cleared": (39, 174, 96), "Deposit": (41, 128, 185), "Pending": (22, 160, 133),
        "Bounced": (192, 57, 43), "Cxn": (230, 126, 34),
    }
    mode_colors = {
        "AIRTEL": (26, 188, 156), "CASH": (46, 204, 113), "CHEQUE": (142, 68, 173), "DD": (155, 89, 182),
        "ONLINE_PAYMENT": (52, 152, 219), "RTGS": (44, 62, 80), UNMAPPED_LABEL: (127, 140, 141),
    }
    group_colors = {"UPDATED": (39, 174, 96), "BOUNCED/CANCELLED": (192, 57, 43)}
    navy = (15, 32, 60)
    status_cols = _RECEIPT_MADE_STATUS_COLS
    grand_total = table["grand_total"] or 1

    width = 1160
    pad = 40
    header_h = 130
    gap = 28
    kpi_h = 120
    section_header_h = 46
    card_h = 118
    card_gap = 16

    height = header_h + gap + kpi_h + gap
    for group in table["groups"]:
        n_visible = len(_visible_rows(group))
        if n_visible == 0:
            continue
        height += section_header_h + n_visible * (card_h + card_gap) + 14
    height += 50

    base = Image.new("RGB", (width, height), (243, 245, 248))

    def shadowed_card(box, radius=16, fill="white"):
        x0, y0, x1, y1 = box
        shadow = Image.new("RGBA", base.size, (0, 0, 0, 0))
        sd = ImageDraw.Draw(shadow)
        sd.rounded_rectangle([x0, y0 + 6, x1, y1 + 6], radius=radius, fill=(15, 32, 60, 55))
        shadow = shadow.filter(ImageFilter.GaussianBlur(7))
        composited = Image.alpha_composite(base.convert("RGBA"), shadow).convert("RGB")
        base.paste(composited, (0, 0))
        ImageDraw.Draw(base).rounded_rectangle(box, radius=radius, fill=fill)

    draw = ImageDraw.Draw(base)
    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 26), "RECEIPT PARTICULARS", font=font(30), fill="white")
    draw.text((pad, 68), "By Receipt Status & Payment Mode", font=font(14), fill=(180, 197, 222))
    draw.text(
        (pad, 92),
        f"Source: RECEIPT STATUS x STATUS IN TAB  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        font=font(12), fill=(140, 160, 190),
    )

    y = header_h + gap
    kpi_items = [("GRAND TOTAL", grand_total, navy)]
    kpi_items += [(_RECEIPT_MADE_STATUS_LABELS[c], table["grand_totals"][c], status_colors[c]) for c in status_cols]
    n_kpi = len(kpi_items)
    kpi_gap = 18
    kpi_w = (width - 2 * pad - kpi_gap * (n_kpi - 1)) // n_kpi
    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        shadowed_card(box, radius=14, fill="white")
        d = ImageDraw.Draw(base)
        d.rounded_rectangle([x + 16, y + 16, x + 42, y + 42], radius=7, fill=color)
        d.text((x + 16, y + 56), f"{val:,}", font=font(24), fill=(25, 25, 25))
        d.text((x + 16, y + 92), label, font=font(11), fill=(120, 120, 120))
        pct = "of receipts" if label == "GRAND TOTAL" else f"{val / grand_total * 100:.1f}%"
        d.text((x + kpi_w - 68, y + 20), pct, font=font(11), fill=color)
        x += kpi_w + kpi_gap

    y += kpi_h + gap

    for group in table["groups"]:
        visible = _visible_rows(group)
        if not visible:
            continue
        gcolor = group_colors.get(group["status"], navy)
        d = ImageDraw.Draw(base)
        d.rounded_rectangle([pad, y, width - pad, y + section_header_h], radius=10, fill=gcolor)
        d.text((pad + 18, y + 12), str(group["status"]), font=font(18), fill="white")
        gt = sum(r["total"] for r in visible)
        gt_text = f"GROUP TOTAL {gt:,}"
        gt_font = font(13)
        gt_w = d.textlength(gt_text, font=gt_font)
        d.text((width - pad - 18 - gt_w, y + 15), gt_text, font=gt_font, fill="white")
        y += section_header_h + 14

        for r in visible:
            mode = r["mode"]
            accent = mode_colors.get(mode, (100, 100, 100))
            box = [pad, y, width - pad, y + card_h]
            shadowed_card(box, radius=16, fill="white")
            d = ImageDraw.Draw(base)
            d.rounded_rectangle([pad, y, pad + 8, y + card_h], radius=4, fill=accent)
            d.text((pad + 30, y + 16), mode, font=font(18), fill=(25, 25, 25))

            total_val = r["total"]
            pill_text = f"TOTAL {total_val:,}"
            pill_font = font(13)
            pill_w = d.textlength(pill_text, font=pill_font) + 26
            d.rounded_rectangle([width - pad - pill_w - 20, y + 14, width - pad - 20, y + 42], radius=15, fill=navy)
            d.text((width - pad - pill_w / 2 - 20, y + 28), pill_text, font=pill_font, fill="white", anchor="mm")

            cx, cy = pad + 30, y + 56
            chip_gap = 14
            chip_w = max(140, (width - 2 * pad - 60 - chip_gap * (len(status_cols) - 1)) // len(status_cols))
            for c in status_cols:
                val = r["counts"][c]
                if val:
                    color = status_colors[c]
                    d.rounded_rectangle([cx, cy, cx + chip_w, cy + 46], radius=10, fill=color)
                    d.text((cx + 14, cy + 6), f"{val:,}", font=font(17), fill="white")
                    d.text((cx + 14, cy + 27), _RECEIPT_MADE_STATUS_LABELS[c], font=font(10), fill="white")
                else:
                    d.rounded_rectangle([cx, cy, cx + chip_w, cy + 46], radius=10, fill=(230, 232, 235))
                    d.text((cx + 14, cy + 6), "-", font=font(17), fill=(160, 160, 160))
                    d.text((cx + 14, cy + 27), _RECEIPT_MADE_STATUS_LABELS[c], font=font(10), fill=(160, 160, 160))
                cx += chip_w + chip_gap

            y += card_h + card_gap
        y += 14

    ImageDraw.Draw(base).text(
        (pad, height - 32), "Generated automatically — DHC Working Automation", font=font(11), fill=(150, 150, 150)
    )

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


# Friendlier display names for the 3 codes that actually appear in STATUS
# IN TAB day to day. "Bounced"/"Deposit" are kept as-is if they ever show
# up — they're real STATUS_MAP outputs too, just rarer in practice.
STATUS_DISPLAY_RENAME = {"Cxn": "Cancelled"}
STATUS_DISPLAY_ORDER = ["Cleared", "Pending", "Cancelled", "Bounced", "Deposit", UNMAPPED_LABEL]


def build_receipt_made_summary_by_status(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Clean Mode x Status breakdown, sourced directly from STATUS IN TAB
    (C -> Cleared, NA -> Pending, X -> Cancelled), independent of the
    separate 'RECEIPT STATUS' field the legacy two-table replica above
    splits on. Every receipt is counted exactly once: each Mode row's
    Total always equals the true row count for that mode, and the Grand
    Total always equals len(dcr_tab) — so this always reconciles against
    a plain filter on the raw data, with nothing silently dropped.
    """
    d = dcr_tab.copy()
    d["Status Display"] = d["Status"].replace(STATUS_DISPLAY_RENAME)

    modes = [m for m in MODE_ORDER if m in d["Mode"].unique()]
    modes += [m for m in d["Mode"].unique() if m not in modes]  # safety net for any surprise value

    statuses = [s for s in STATUS_DISPLAY_ORDER if s in d["Status Display"].unique()]
    statuses += [s for s in d["Status Display"].unique() if s not in statuses]  # same safety net

    rows = []
    for mode in modes:
        mdf = d[d["Mode"] == mode]
        row = {"Mode": mode}
        for s in statuses:
            row[s] = int((mdf["Status Display"] == s).sum())
        row["Total"] = len(mdf)
        rows.append(row)

    out = pd.DataFrame(rows, columns=["Mode"] + statuses + ["Total"])
    grand = {"Mode": "Grand Total"}
    for s in statuses:
        grand[s] = int(out[s].sum())
    grand["Total"] = int(out["Total"].sum())
    out = pd.concat([out, pd.DataFrame([grand])], ignore_index=True)
    return out


def receipt_made_summary_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    """Single-sheet, formatted .xlsx containing just this one summary table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt Made Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B7C0C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(df.columns)
    ws.append(["Receipt Made Summary — by Payment Mode & Status"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}  ·  Source: STATUS IN TAB (C / NA / X)"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(list(df.columns))
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for _, r in df.iterrows():
        ws.append(list(r))

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        is_total = row[0].value == "Grand Total"
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.column_letter != "A" else "left")
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    for i, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(col)) + 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def receipt_made_summary_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Dashboard-style PNG: a row of KPI stat boxes up top (grand totals),
    then one soft-shadowed card per payment mode, each containing its own
    colour-coded status boxes. Deliberately built as a card dashboard
    rather than a row-per-line table, so it reads as a professional
    summary graphic rather than a spreadsheet screenshot.
    """
    from PIL import Image, ImageDraw, ImageFont, ImageFilter

    font = _img_font

    status_colors = {
        "Cleared": (39, 174, 96),
        "Pending": (243, 156, 18),
        "Cancelled": (231, 76, 60),
        "Bounced": (192, 57, 43),
        "Deposit": (41, 128, 185),
        UNMAPPED_LABEL: (127, 140, 141),
    }
    mode_colors = {
        "AIRTEL": (26, 188, 156),
        "CASH": (46, 204, 113),
        "CHEQUE": (142, 68, 173),
        "DD": (155, 89, 182),
        "ONLINE_PAYMENT": (52, 152, 219),
        "RTGS": (44, 62, 80),
        UNMAPPED_LABEL: (127, 140, 141),
    }
    navy = (15, 32, 60)

    cols = list(df.columns)
    status_cols = cols[1:-1]
    mode_rows = df[df["Mode"] != "Grand Total"].reset_index(drop=True)
    grand = df[df["Mode"] == "Grand Total"].iloc[0]
    grand_total = int(grand["Total"]) or 1

    width = 1120
    pad = 40
    header_h = 130
    gap = 30
    kpi_h = 130
    card_h = 130
    card_gap = 18
    height = header_h + gap + kpi_h + gap + 34 + len(mode_rows) * (card_h + card_gap) + 50

    base = Image.new("RGB", (width, height), (243, 245, 248))

    def shadowed_card(box, radius=16, fill="white"):
        """Draws a soft drop-shadow behind a rounded rect, then the rect itself."""
        x0, y0, x1, y1 = box
        shadow = Image.new("RGBA", base.size, (0, 0, 0, 0))
        sd = ImageDraw.Draw(shadow)
        sd.rounded_rectangle([x0, y0 + 6, x1, y1 + 6], radius=radius, fill=(15, 32, 60, 55))
        shadow = shadow.filter(ImageFilter.GaussianBlur(7))
        composited = Image.alpha_composite(base.convert("RGBA"), shadow).convert("RGB")
        base.paste(composited, (0, 0))
        ImageDraw.Draw(base).rounded_rectangle(box, radius=radius, fill=fill)

    draw = ImageDraw.Draw(base)

    # ── Header banner ──
    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 26), "RECEIPT MADE SUMMARY", font=font(30), fill="white")
    draw.text((pad, 68), "By Payment Mode & Receipt Status", font=font(14), fill=(180, 197, 222))
    draw.text(
        (pad, 92),
        f"Source: STATUS IN TAB (C/NA/X)  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        font=font(12), fill=(140, 160, 190),
    )

    # ── KPI boxes (grand totals) ──
    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", grand_total, navy)]
    kpi_items += [(sc.upper(), int(grand[sc]), status_colors.get(sc, (100, 100, 100))) for sc in status_cols]
    n_kpi = len(kpi_items)
    kpi_gap = 18
    kpi_w = (width - 2 * pad - kpi_gap * (n_kpi - 1)) // n_kpi

    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        shadowed_card(box, radius=14, fill="white")
        d = ImageDraw.Draw(base)
        d.rounded_rectangle([x + 16, y + 16, x + 42, y + 42], radius=7, fill=color)
        d.text((x + 16, y + 56), f"{val:,}", font=font(26), fill=(25, 25, 25))
        d.text((x + 16, y + 94), label, font=font(11), fill=(120, 120, 120))
        pct = "of total" if label == "TOTAL RECEIPTS" else f"{val / grand_total * 100:.1f}%"
        d.text((x + kpi_w - 62, y + 20), pct, font=font(12), fill=color)
        x += kpi_w + kpi_gap

    # ── Section label ──
    y += kpi_h + gap
    ImageDraw.Draw(base).text((pad, y), "BREAKDOWN BY PAYMENT MODE", font=font(15), fill=(70, 70, 70))
    y += 34

    # ── Per-mode cards ──
    for _, r in mode_rows.iterrows():
        mode = str(r["Mode"])
        accent = mode_colors.get(mode, (100, 100, 100))
        box = [pad, y, width - pad, y + card_h]
        shadowed_card(box, radius=16, fill="white")
        d = ImageDraw.Draw(base)

        d.rounded_rectangle([pad, y, pad + 8, y + card_h], radius=4, fill=accent)
        d.text((pad + 30, y + 18), mode, font=font(20), fill=(25, 25, 25))

        total_val = int(r["Total"])
        pill_text = f"TOTAL {total_val:,}"
        pill_font = font(13)
        pill_w = d.textlength(pill_text, font=pill_font) + 26
        d.rounded_rectangle([width - pad - pill_w - 20, y + 16, width - pad - 20, y + 46], radius=15, fill=navy)
        d.text((width - pad - pill_w / 2 - 20, y + 31), pill_text, font=pill_font, fill="white", anchor="mm")

        cx, cy = pad + 30, y + 62
        chip_gap = 14
        chip_w = max(140, (width - 2 * pad - 60 - chip_gap * (len(status_cols) - 1)) // len(status_cols))
        for sc in status_cols:
            val = int(r[sc])
            color = status_colors.get(sc, (100, 100, 100))
            d.rounded_rectangle([cx, cy, cx + chip_w, cy + 48], radius=10, fill=color)
            d.text((cx + 14, cy + 7), f"{val:,}", font=font(18), fill="white")
            d.text((cx + 14, cy + 29), str(sc), font=font(11), fill=(255, 255, 255))
            cx += chip_w + chip_gap

        y += card_h + card_gap

    ImageDraw.Draw(base).text(
        (pad, height - 32), "Generated automatically — DHC Working Automation", font=font(11), fill=(150, 150, 150)
    )

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


_FONT_SIZE_SCALE = 1.4  # applied uniformly across every PNG summary — bumped from 1.15 to make text noticeably larger
_FONT_CACHE: dict = {}
_FONT_CANDIDATES = [
    "arial.ttf",                                       # Windows
    "Arial.ttf",                                        # macOS (relative)
    "/System/Library/Fonts/Supplemental/Arial.ttf",     # macOS (absolute)
    "DejaVuSans.ttf",                                   # Linux (Pillow ships this one)
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Linux (absolute)
]


def _img_font(size: int):
    """
    Shared font loader for every PNG summary builder below (RTGS, Delay,
    RCPT CXN, Cash Mode, Receipt Made Report). Fixes two things:

    1. Blur: this used to call ImageFont.load_default(size=...)
       everywhere. That's Pillow's own bundled font, and on some
       Pillow/OS combinations it silently ignores the size argument (a
       "30px heading" quietly renders at a fixed ~11px), then
       resize_png_for_mail downsamples the whole image again for the
       email on top of that — exactly what produced the blur. This now
       tries real scalable TrueType fonts first (Arial on Windows/macOS,
       DejaVu Sans on Linux), which anti-alias cleanly at any size, and
       only falls back to load_default() if none of those font files
       exist at all.
    2. Size: every requested size is scaled up by _FONT_SIZE_SCALE
       (currently 40%) before loading, so text across all five dashboard
       images is noticeably larger and easier to read.

    Cached per final pixel size, since this runs inside tight per-row
    drawing loops.
    """
    scaled = max(1, round(size * _FONT_SIZE_SCALE))
    if scaled in _FONT_CACHE:
        return _FONT_CACHE[scaled]
    from PIL import ImageFont
    loaded = None
    for name in _FONT_CANDIDATES:
        try:
            loaded = ImageFont.truetype(name, scaled)
            break
        except (OSError, IOError):
            continue
    if loaded is None:
        try:
            loaded = ImageFont.load_default(size=scaled)
        except TypeError:
            loaded = ImageFont.load_default()
    _FONT_CACHE[scaled] = loaded
    return loaded


_TAT_COLORS = {
    "Less then 4": (39, 174, 96),
    "5 TO 10": (243, 156, 18),
    "Great then 10": (231, 76, 60),
    "Pending": (149, 165, 166),
    "Cxn": (127, 140, 141),
}


def rtgs_summary_to_image_bytes(summary: dict) -> io.BytesIO:
    """
    RTGS Intelligence dashboard PNG: navy/gold 'control tower' theme. A KPI
    strip up top (count totals per TAT bucket + total value), then zones
    ranked by volume as horizontal stacked bars (green/amber/red = TAT mix),
    and a navy leaderboard panel for online payment channels. Deliberately
    bar-chart-led — visually distinct from the card-grid used for Receipt
    Made and the heatmap used for Delay, so each export reads as its own
    dashboard rather than a re-skin of the others.
    """
    from PIL import Image, ImageDraw

    navy = (13, 27, 51)
    gold = (197, 160, 89)

    matrix = summary["matrix"]
    zones = sorted(matrix["zones"], key=lambda b: b["subtotal"]["total_count"], reverse=True)
    grand = matrix["grand_total"]
    online = summary.get("online_source_block", {"rows": [], "total": 0})

    MAX_ZONES = 10
    shown_zones = zones[:MAX_ZONES]
    extra = len(zones) - len(shown_zones)

    width = 1160
    pad = 40
    header_h = 130
    kpi_h = 110
    gap = 26
    row_h = 46
    row_gap = 10
    online_rows = online["rows"] or [("No online receipts this period", 0)]
    online_h = 50 + len(online_rows) * 32 + 16

    height = (
        header_h + gap + kpi_h + gap + 30
        + len(shown_zones) * (row_h + row_gap)
        + (26 if extra else 0)
        + 34 + gap + online_h + 60
    )

    base = Image.new("RGB", (width, height), (244, 246, 249))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 24), "RTGS INTELLIGENCE DASHBOARD", font=_img_font(30), fill="white")
    draw.text((pad, 66), "Zone-wise Turn-Around-Time Performance", font=_img_font(14), fill=(200, 210, 225))
    draw.text(
        (pad, 92),
        f"{datetime.now().strftime('%d-%b-%Y %H:%M')}  ·  {len(zones)} zones",
        font=_img_font(12), fill=gold,
    )

    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", f"{grand['total_count']:,}", navy)]
    kpi_items += [(TAT_DISPLAY[b].upper(), f"{grand['by_bucket'][b]['count']:,}", _TAT_COLORS[b]) for b in TAT_ORDER]
    kpi_items.append(("TOTAL VALUE (CR)", f"{grand['total_value']:.1f}", gold))
    n = len(kpi_items)
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (n - 1)) // n
    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        draw.rounded_rectangle(box, radius=12, fill="white", outline=(226, 232, 240))
        draw.rectangle([x, y, x + 6, y + kpi_h], fill=color)
        draw.text((x + 18, y + 18), str(val), font=_img_font(22), fill=(20, 20, 20))
        draw.text((x + 18, y + 56), label, font=_img_font(10), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    draw.text((pad, y), "ZONE PERFORMANCE — RANKED BY VOLUME", font=_img_font(14), fill=(70, 70, 70))
    y += 30

    max_zone_count = max((b["subtotal"]["total_count"] for b in shown_zones), default=1) or 1
    label_w = 190
    bar_x0 = pad + label_w
    bar_max_w = width - pad - bar_x0 - 90

    for block in shown_zones:
        zname = block["zone"]
        st = block["subtotal"]
        total = st["total_count"]
        draw.text((pad, y + 12), str(zname), font=_img_font(13), fill=(30, 30, 30))

        bar_w = int(bar_max_w * (total / max_zone_count))
        by0, by1 = y + 6, y + row_h - 6
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + bar_max_w, by1], radius=6, fill=(232, 236, 241))

        if total:
            label_font = _img_font(11)
            counts = {b: st["by_bucket"][b]["count"] for b in TAT_ORDER}
            # Proportional width first, but every non-zero bucket is floored
            # to whatever width its own number needs to actually fit —
            # otherwise a small-but-real Yellow/Red count just disappears.
            raw_w = {b: bar_w * (counts[b] / total) for b in TAT_ORDER}
            min_w = {
                b: (draw.textlength(f"{counts[b]:,}", font=label_font) + 12) if counts[b] > 0 else 0
                for b in TAT_ORDER
            }
            seg_w = {b: max(raw_w[b], min_w[b]) for b in TAT_ORDER}

            # If the floors push total width past the bar, claw the excess
            # back from whichever segment(s) have the most room to spare
            # (almost always the dominant "< 4 Days" segment), never below
            # that segment's own text-minimum.
            overflow = sum(seg_w.values()) - bar_w
            for b in sorted(TAT_ORDER, key=lambda k: seg_w[k], reverse=True):
                if overflow <= 0:
                    break
                spare = seg_w[b] - min_w[b]
                take = min(spare, overflow)
                seg_w[b] -= take
                overflow -= take

            seg_x = bar_x0
            for b in TAT_ORDER:
                w = seg_w[b]
                if w > 0:
                    draw.rectangle([seg_x, by0, seg_x + w, by1], fill=_TAT_COLORS[b])
                    draw.text(
                        (seg_x + w / 2, (by0 + by1) / 2), f"{counts[b]:,}",
                        font=label_font, fill="white", anchor="mm",
                    )
                    seg_x += w
        draw.text((bar_x0 + bar_max_w + 12, (by0 + by1) / 2), f"{total:,}", font=_img_font(14), fill=(20, 20, 20), anchor="lm")
        y += row_h + row_gap

    if extra:
        draw.text((pad, y), f"+ {extra} more zone(s) not shown", font=_img_font(11), fill=(140, 140, 140))
        y += 26

    lx = pad
    for b in TAT_ORDER:
        draw.rectangle([lx, y, lx + 14, y + 14], fill=_TAT_COLORS[b])
        draw.text((lx + 20, y - 2), TAT_DISPLAY[b], font=_img_font(11), fill=(70, 70, 70))
        lx += 150
    y += 34

    y += gap
    draw.rounded_rectangle([pad, y, width - pad, y + online_h], radius=14, fill=navy)
    draw.text((pad + 20, y + 16), "ONLINE PAYMENT CHANNELS", font=_img_font(15), fill=gold)
    ry = y + 50
    max_val = max((c for _, c in online_rows), default=1) or 1
    for name, cnt in online_rows:
        draw.text((pad + 20, ry), str(name), font=_img_font(12), fill="white")
        bw = int(240 * (cnt / max_val)) if max_val and cnt else 0
        if bw > 0:
            draw.rounded_rectangle([pad + 300, ry + 2, pad + 300 + max(bw, 2), ry + 16], radius=6, fill=gold)
        draw.text((pad + 300 + 250, ry), f"{cnt:,}", font=_img_font(12), fill="white")
        ry += 32

    draw.text((pad, height - 32), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def zone_overview_to_image_bytes(summary: dict) -> io.BytesIO:
    """
    Zone Overview dashboard PNG: per-zone payment-mode split as horizontal
    stacked bars (same proportional-bar technique as the RTGS/Delay
    dashboards, so it still reads cleanly once shrunk to mail width),
    ranked by volume, with TAT Exceeded / Bounced / Cancelled shown as
    small coloured chips at the end of each zone's row, plus a KPI strip
    of the grand totals. Takes the output of build_zone_overview_summary.
    """
    from PIL import Image, ImageDraw

    slate = (30, 41, 59)
    teal = (13, 148, 136)
    mode_colors = {
        "AIRTEL": (26, 188, 156), "CASH": (46, 204, 113),
        "CHEQUE": (142, 68, 173), "DD": (155, 89, 182),
        "ONLINE_PAYMENT": (52, 152, 219), "RTGS": (44, 62, 80),
        UNMAPPED_LABEL: (127, 140, 141),
    }
    modes = summary["modes"]
    zones = summary["zones"]
    grand = summary["grand_total"]

    MAX_ZONES = 12
    shown_zones = zones[:MAX_ZONES]
    extra = len(zones) - len(shown_zones)

    width = 1000
    pad = 36
    header_h = 110
    kpi_h = 90
    gap = 22
    row_h = 44
    row_gap = 10
    chip_col_w = 210

    height = (
        header_h + gap + kpi_h + gap + 28
        + len(shown_zones) * (row_h + row_gap)
        + (24 if extra else 0)
        + 30 + 46
    )

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=slate)
    draw.text((pad, 20), "ZONE OVERVIEW", font=_img_font(26), fill="white")
    draw.text((pad, 58), "Pay-mode split, TAT exceeded & Bounced/Cancelled — by zone", font=_img_font(13), fill=(180, 210, 205))
    draw.text((pad, 80), datetime.now().strftime("%d-%b-%Y %H:%M"), font=_img_font(11), fill=teal)

    y = header_h + gap
    kpi_items = [
        ("TOTAL RECEIPTS", f"{grand['total']:,}", slate),
        ("TAT EXCEEDED", f"{grand['tat_exceeded']:,}", (230, 126, 34)),
        ("BOUNCED", f"{grand['bounced']:,}", (192, 57, 43)),
        ("CANCELLED", f"{grand['cancelled']:,}", (127, 140, 141)),
    ]
    kpi_gap = 14
    kpi_w = (width - 2 * pad - kpi_gap * (len(kpi_items) - 1)) // len(kpi_items)
    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        draw.rounded_rectangle(box, radius=10, fill="white", outline=(220, 226, 232))
        draw.rectangle([x, y, x + 5, y + kpi_h], fill=color)
        draw.text((x + 14, y + 16), val, font=_img_font(20), fill=(20, 20, 20))
        draw.text((x + 14, y + 52), label, font=_img_font(9), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    y += kpi_h + gap
    draw.text((pad, y), "ZONE-WISE MODE SPLIT — RANKED BY VOLUME", font=_img_font(13), fill=(70, 70, 70))
    y += 28

    max_zone_count = max((z["total"] for z in shown_zones), default=1) or 1
    label_w = 90
    bar_x0 = pad + label_w
    bar_max_w = width - pad - bar_x0 - chip_col_w

    for block in shown_zones:
        draw.text((pad, y + 12), str(block["zone"]), font=_img_font(13), fill=(30, 30, 30))
        total = block["total"]
        bar_w = int(bar_max_w * (total / max_zone_count))
        by0, by1 = y + 6, y + row_h - 6
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + bar_max_w, by1], radius=6, fill=(232, 236, 241))

        if total and bar_w:
            label_font = _img_font(10)
            counts = block["by_mode"]
            raw_w = {m: bar_w * (counts.get(m, 0) / total) for m in modes}
            min_w = {
                m: (draw.textlength(f"{counts.get(m, 0):,}", font=label_font) + 10) if counts.get(m, 0) > 0 else 0
                for m in modes
            }
            seg_w = {m: max(raw_w[m], min_w[m]) for m in modes}
            overflow = sum(seg_w.values()) - bar_w
            for m in sorted(modes, key=lambda k: seg_w[k], reverse=True):
                if overflow <= 0:
                    break
                spare = seg_w[m] - min_w[m]
                take = min(spare, overflow)
                seg_w[m] -= take
                overflow -= take

            seg_x = bar_x0
            for m in modes:
                w = seg_w[m]
                if w > 0:
                    color = mode_colors.get(m, (127, 140, 141))
                    draw.rectangle([seg_x, by0, seg_x + w, by1], fill=color)
                    draw.text((seg_x + w / 2, (by0 + by1) / 2), f"{counts.get(m, 0):,}", font=label_font, fill="white", anchor="mm")
                    seg_x += w

        chip_x = bar_x0 + bar_max_w + 14
        chip_y = (by0 + by1) / 2 if total else y + row_h / 2
        for label, cnt, color in [
            ("TAT", block["tat_exceeded"], (230, 126, 34)),
            ("BND", block["bounced"], (192, 57, 43)),
            ("CXN", block["cancelled"], (127, 140, 141)),
        ]:
            chip_fill = color if cnt else (225, 228, 232)
            chip_text = "white" if cnt else (150, 150, 150)
            draw.rounded_rectangle([chip_x, chip_y - 11, chip_x + 58, chip_y + 11], radius=10, fill=chip_fill)
            draw.text((chip_x + 29, chip_y), f"{label} {cnt}", font=_img_font(10), fill=chip_text, anchor="mm")
            chip_x += 64

        draw.text((pad, y + row_h + row_gap - 6), "", font=_img_font(1))  # keep row spacing consistent
        y += row_h + row_gap

    if extra:
        draw.text((pad, y), f"+ {extra} more zone(s) not shown", font=_img_font(11), fill=(140, 140, 140))
        y += 24

    lx = pad
    for m in modes:
        color = mode_colors.get(m, (127, 140, 141))
        draw.rectangle([lx, y, lx + 12, y + 12], fill=color)
        draw.text((lx + 18, y - 2), MODE_DISPLAY.get(m, m), font=_img_font(10), fill=(70, 70, 70))
        lx += 120

    draw.text((pad, height - 24), "Generated automatically — DHC Working Automation", font=_img_font(10), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def delay_summary_to_image_bytes(summary: dict) -> io.BytesIO:
    """
    Delay-in-RCPTING dashboard PNG: slate/teal theme, zones ranked by
    volume as horizontal stacked bars (same proportional-bar technique as
    the RTGS dashboard) plus a callout panel calling out Cancelled/Pending
    receipts specifically.

    Also includes a PAYMENT MODE — RANKED BY VOLUME section (Cash,
    Cheque/DD, Online Payment, RTGS), drawn the same way as the zone
    section. The data for this was already being computed — the matrix
    is built with breakdown_col="Mode" in build_delay_in_rcpting_summary
    — it just wasn't being drawn anywhere; this adds that missing panel
    by summing each zone's per-mode breakdown into month-wide mode
    totals. Modes with zero receipts this period are hidden entirely,
    same rule as the Receipt Made Report's zero rows.

    Previously the zone section was a fixed-width heatmap grid (one
    fixed-width column per TAT bucket). That worked fine back when there
    were only 3 buckets, but it was never updated when Cancelled/Pending
    became their own TAT buckets alongside the day-ranges — with 5
    columns the grid needed ~1400px against a hardcoded 1160px canvas, so
    it was already clipping off the right-hand columns, and shrinking
    that dense a grid down to mail width (see resize_png_for_mail) made
    whatever survived unreadable. A proportional bar's width doesn't
    depend on how many buckets there are, so it doesn't have that
    failure mode, and it reads cleanly all the way down to mail width —
    same reason the RTGS dashboard was never affected.
    """
    from PIL import Image, ImageDraw

    slate = (30, 41, 59)
    teal = (13, 148, 136)
    matrix = summary["matrix"]
    zones = sorted(matrix["zones"], key=lambda b: b["subtotal"]["total_count"], reverse=True)
    grand = matrix["grand_total"]

    MAX_ZONES = 12
    shown_zones = zones[:MAX_ZONES]
    extra = len(zones) - len(shown_zones)

    # Month-wide totals per payment mode, summed across every zone's
    # already-computed breakdown — this is the data the new Payment Mode
    # section draws.
    mode_totals: dict = {}
    for block in matrix["zones"]:
        for code, label, stats in block["breakdown"]:
            entry = mode_totals.setdefault(
                code, {"label": label, "total_count": 0, "by_bucket": {b: {"count": 0} for b in TAT_ORDER}},
            )
            entry["total_count"] += stats["total_count"]
            for b in TAT_ORDER:
                entry["by_bucket"][b]["count"] += stats["by_bucket"][b]["count"]
    mode_rows = [
        mode_totals[code] for code in MODE_ORDER
        if code in mode_totals and mode_totals[code]["total_count"] > 0
    ]

    width = 1160
    pad = 40
    header_h = 120
    kpi_h = 100
    gap = 26
    row_h = 44
    row_gap = 10
    callout_h = 96
    section_label_h = 30

    height = (
        header_h + gap + kpi_h + gap + section_label_h
        + len(shown_zones) * (row_h + row_gap)
        + (26 if extra else 0)
        + gap + section_label_h
        + len(mode_rows) * (row_h + row_gap)
        + 34 + gap + callout_h + 60
    )

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=slate)
    draw.text((pad, 22), "DELAY IN RCPTING — TURN-AROUND TIME", font=_img_font(28), fill="white")
    draw.text((pad, 64), "Receipt processing time, Zone x TAT bucket — full month, all payment modes", font=_img_font(14), fill=(180, 210, 205))
    draw.text(
        (pad, 88),
        f"{datetime.now().strftime('%d-%b-%Y %H:%M')}  ·  {len(zones)} zones",
        font=_img_font(12), fill=teal,
    )

    y = header_h + gap
    kpi_items = [("TOTAL RECEIPTS", f"{grand['total_count']:,}", slate)]
    kpi_items += [(TAT_DISPLAY[b].upper(), f"{grand['by_bucket'][b]['count']:,}", _TAT_COLORS[b]) for b in TAT_ORDER]
    n = len(kpi_items)
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (n - 1)) // n
    x = pad
    for label, val, color in kpi_items:
        box = [x, y, x + kpi_w, y + kpi_h]
        draw.rounded_rectangle(box, radius=12, fill="white", outline=(220, 226, 232))
        draw.rectangle([x, y, x + 6, y + kpi_h], fill=color)
        draw.text((x + 16, y + 18), val, font=_img_font(20), fill=(20, 20, 20))
        draw.text((x + 16, y + 56), label, font=_img_font(9), fill=(110, 110, 110))
        x += kpi_w + kpi_gap

    label_w = 190
    bar_x0 = pad + label_w
    bar_max_w = width - pad - bar_x0 - 90

    def draw_ranked_bar(y: int, name: str, total: int, by_bucket: dict, max_count: int) -> int:
        """One horizontal proportional bar (name label + TAT-coloured stacked segments + total). Shared by the Zone and Payment Mode sections."""
        draw.text((pad, y + 12), str(name), font=_img_font(13), fill=(30, 30, 30))
        bar_w = int(bar_max_w * (total / max_count)) if max_count else 0
        by0, by1 = y + 6, y + row_h - 6
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + bar_max_w, by1], radius=6, fill=(232, 236, 241))
        if total:
            label_font = _img_font(11)
            counts = {b: by_bucket[b]["count"] for b in TAT_ORDER}
            raw_w = {b: bar_w * (counts[b] / total) for b in TAT_ORDER}
            min_w = {
                b: (draw.textlength(f"{counts[b]:,}", font=label_font) + 12) if counts[b] > 0 else 0
                for b in TAT_ORDER
            }
            seg_w = {b: max(raw_w[b], min_w[b]) for b in TAT_ORDER}
            overflow = sum(seg_w.values()) - bar_w
            for b in sorted(TAT_ORDER, key=lambda k: seg_w[k], reverse=True):
                if overflow <= 0:
                    break
                spare = seg_w[b] - min_w[b]
                take = min(spare, overflow)
                seg_w[b] -= take
                overflow -= take
            seg_x = bar_x0
            for b in TAT_ORDER:
                w = seg_w[b]
                if w > 0:
                    draw.rectangle([seg_x, by0, seg_x + w, by1], fill=_TAT_COLORS[b])
                    draw.text(
                        (seg_x + w / 2, (by0 + by1) / 2), f"{counts[b]:,}",
                        font=label_font, fill="white", anchor="mm",
                    )
                    seg_x += w
        draw.text((bar_x0 + bar_max_w + 12, (by0 + by1) / 2), f"{total:,}", font=_img_font(14), fill=(20, 20, 20), anchor="lm")
        return y + row_h + row_gap

    y += kpi_h + gap
    draw.text((pad, y), "ZONE-WISE PROCESSING TIME — RANKED BY VOLUME", font=_img_font(14), fill=(70, 70, 70))
    y += section_label_h

    max_zone_count = max((b["subtotal"]["total_count"] for b in shown_zones), default=1) or 1
    for block in shown_zones:
        y = draw_ranked_bar(y, block["zone"], block["subtotal"]["total_count"], block["subtotal"]["by_bucket"], max_zone_count)

    if extra:
        draw.text((pad, y), f"+ {extra} more zone(s) not shown", font=_img_font(11), fill=(140, 140, 140))
        y += 26

    y += gap
    draw.text((pad, y), "PAYMENT MODE — RANKED BY VOLUME", font=_img_font(14), fill=(70, 70, 70))
    y += section_label_h

    max_mode_count = max((m["total_count"] for m in mode_rows), default=1) or 1
    for m in sorted(mode_rows, key=lambda r: r["total_count"], reverse=True):
        y = draw_ranked_bar(y, m["label"], m["total_count"], m["by_bucket"], max_mode_count)

    lx = pad
    for b in TAT_ORDER:
        draw.rectangle([lx, y, lx + 14, y + 14], fill=_TAT_COLORS[b])
        draw.text((lx + 20, y - 2), TAT_DISPLAY[b], font=_img_font(11), fill=(70, 70, 70))
        lx += 150
    y += 34

    y += gap
    cxn_count = grand["by_bucket"].get("Cxn", {}).get("count", 0)
    pending_count = grand["by_bucket"].get("Pending", {}).get("count", 0)
    total_count = grand["total_count"] or 1
    draw.rounded_rectangle([pad, y, width - pad, y + callout_h], radius=14, fill=slate)
    draw.text((pad + 20, y + 16), "CANCELLED & PENDING — OF THIS MONTH'S RECEIPTS", font=_img_font(13), fill=teal)
    stat_w = (width - 2 * pad - 40) // 2
    for i, (label, cnt, color) in enumerate([
        ("Cancelled", cxn_count, _TAT_COLORS["Cxn"]),
        ("Pending", pending_count, _TAT_COLORS["Pending"]),
    ]):
        sx = pad + 20 + i * stat_w
        draw.rectangle([sx, y + 44, sx + 6, y + 76], fill=color)
        draw.text((sx + 16, y + 44), f"{cnt:,}", font=_img_font(22), fill="white")
        draw.text((sx + 16, y + 72), f"{label}  ·  {cnt / total_count * 100:.1f}% of total", font=_img_font(11), fill=(190, 200, 215))

    draw.text((pad, height - 26), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def cash_mode_validation_summary_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Cash Mode Compliance — one-page executive summary PNG. Previously this
    rendered one 'ticket' per violating CIF (a row-level list); this is a
    genuine aggregate summary instead: a donut showing what share of the
    total violation value is high-risk, a violations-by-zone bar chart,
    and a Top 5 Highest-Value CIFs leaderboard — the shape of the problem
    at a glance, not a scrollable list of every row.
    """
    from PIL import Image, ImageDraw

    navy = (20, 33, 61)
    gold = (191, 149, 63)
    teal = (23, 110, 100)
    crimson = (176, 42, 42)
    ink = (30, 30, 34)
    width = 1080

    if df.empty:
        height = 320
        base = Image.new("RGB", (width, height), (245, 248, 247))
        draw = ImageDraw.Draw(base)
        draw.rectangle([0, 0, width, 10], fill=teal)
        draw.ellipse([width / 2 - 60, 90, width / 2 + 60, 210], outline=teal, width=6)
        draw.line([width / 2 - 26, 150, width / 2 - 6, 172], fill=teal, width=6)
        draw.line([width / 2 - 6, 172, width / 2 + 34, 122], fill=teal, width=6)
        draw.text((width / 2, 232), "FULLY COMPLIANT", font=_img_font(26), fill=teal, anchor="mm")
        draw.text((width / 2, 264), "No cash-mode violations detected this period", font=_img_font(13), fill=(100, 100, 100), anchor="mm")
        draw.text((40, height - 28), f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(11), fill=(150, 150, 150))
        buf = io.BytesIO()
        base.save(buf, format="PNG")
        buf.seek(0)
        return buf

    total_count = len(df)
    total_amount = float(df["Grand Total"].sum())
    reason = df["Violation Reason"].fillna("")
    per_day_mask = reason.str.contains("Per-Day", na=False)
    grand_total_mask = reason.str.contains("Grand Total", na=False)
    disabled_mask = reason.str.contains("Disabled", na=False)
    hv_amount = float(df.loc[grand_total_mask, "Grand Total"].sum())
    hv_amount_pct = (hv_amount / total_amount * 100) if total_amount else 0.0

    zone_col = "Zone2" if "Zone2" in df.columns else None
    zone_breakdown = (
        df.groupby(zone_col)["Grand Total"].sum().sort_values(ascending=False).head(6)
        if zone_col else pd.Series(dtype=float)
    )
    max_zone_amount = float(zone_breakdown.max()) if len(zone_breakdown) else 1.0

    top5 = df.sort_values("Grand Total", ascending=False).head(5)
    medals = ["🥇", "🥈", "🥉", "4", "5"]

    pad = 40
    header_h = 110
    kpi_h = 86
    panel_h = 300
    top5_row_h = 46
    top5_h = 46 + len(top5) * top5_row_h + 20
    gap = 20
    height = header_h + gap + kpi_h + gap + panel_h + gap + top5_h + 50

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    # --- header ---
    draw.rectangle([0, 0, width, header_h], fill=navy)
    draw.text((pad, 22), "CASH MODE COMPLIANCE — MONTHLY SUMMARY", font=_img_font(23), fill="white")
    draw.text((pad, 60), "Per-day limit, Grand Total limit & disabled-CIF cash receipts", font=_img_font(12), fill=(190, 200, 215))
    draw.text((pad, 82), datetime.now().strftime("%d-%b-%Y %H:%M"), font=_img_font(10), fill=(150, 165, 185))
    pill_text = f"{hv_amount_pct:.0f}% OF VALUE IS GRAND-TOTAL BREACH"
    ptw = draw.textlength(pill_text, font=_img_font(12))
    px1 = width - pad - ptw - 30
    draw.rounded_rectangle([px1, 34, width - pad, 68], radius=17, fill=gold)
    draw.text(((px1 + width - pad) / 2, 51), pill_text, font=_img_font(12), fill=navy, anchor="mm")

    # --- KPI row ---
    y = header_h + gap
    kpis = [
        ("TOTAL VIOLATIONS", f"{total_count:,}", navy),
        ("PER-DAY > ₹1.95L", f"{int(per_day_mask.sum()):,}", teal),
        ("GRAND TOTAL > ₹10L", f"{int(grand_total_mask.sum()):,}", gold),
        ("DISABLED CIF RECEIPTED", f"{int(disabled_mask.sum()):,}", crimson),
    ]
    n_kpi = len(kpis)
    kpi_gap = 16
    kpi_w = (width - 2 * pad - kpi_gap * (n_kpi - 1)) // n_kpi
    x = pad
    for label, val, color in kpis:
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=10, fill="white", outline=(225, 225, 230))
        draw.rectangle([x, y, x + 5, y + kpi_h], fill=color)
        draw.text((x + 18, y + 16), val, font=_img_font(19), fill=ink)
        draw.text((x + 18, y + kpi_h - 26), label, font=_img_font(10), fill=(120, 120, 120))
        x += kpi_w + kpi_gap

    # --- donut (left) + zone breakdown (right) ---
    y += kpi_h + gap
    col_w = (width - 2 * pad - gap) // 2

    donut_box = [pad, y, pad + col_w, y + panel_h]
    draw.rounded_rectangle(donut_box, radius=12, fill="white", outline=(225, 225, 230))
    draw.text((pad + 20, y + 18), "VIOLATION VALUE — RISK SPLIT", font=_img_font(13), fill=(70, 70, 70))
    cx, cy = pad + col_w // 2 - 70, y + panel_h // 2 + 15
    r_outer, r_inner = 85, 50
    hv_angle = 360 * hv_amount_pct / 100
    bbox = [cx - r_outer, cy - r_outer, cx + r_outer, cy + r_outer]
    draw.pieslice(bbox, -90, -90 + hv_angle, fill=crimson)
    draw.pieslice(bbox, -90 + hv_angle, 270, fill=teal)
    draw.ellipse([cx - r_inner, cy - r_inner, cx + r_inner, cy + r_inner], fill="white")
    draw.text((cx, cy - 8), f"{hv_amount_pct:.0f}%", font=_img_font(26), fill=crimson, anchor="mm")
    draw.text((cx, cy + 18), "of value", font=_img_font(10), fill=(120, 120, 120), anchor="mm")
    lx, ly = cx + r_outer + 30, cy - 34
    draw.rectangle([lx, ly, lx + 14, ly + 14], fill=crimson)
    draw.text((lx + 22, ly - 1), "Grand Total Breach", font=_img_font(12), fill=ink)
    draw.text((lx + 22, ly + 15), f"₹{hv_amount:,.0f}", font=_img_font(11), fill=(110, 110, 110))
    ly += 46
    draw.rectangle([lx, ly, lx + 14, ly + 14], fill=teal)
    draw.text((lx + 22, ly - 1), "Other Reasons", font=_img_font(12), fill=ink)
    draw.text((lx + 22, ly + 15), f"₹{total_amount - hv_amount:,.0f}", font=_img_font(11), fill=(110, 110, 110))

    zone_box = [pad + col_w + gap, y, width - pad, y + panel_h]
    draw.rounded_rectangle(zone_box, radius=12, fill="white", outline=(225, 225, 230))
    zx0 = pad + col_w + gap + 20
    draw.text((zx0, y + 18), "VIOLATIONS BY ZONE", font=_img_font(13), fill=(70, 70, 70))
    zy = y + 50
    label_w = 90
    amt_w = 130
    bar_x0 = zx0 + label_w
    bar_max_w = (width - pad) - bar_x0 - amt_w - 20
    zone_palette = [navy, teal, gold, crimson, (110, 90, 150), (110, 110, 110)]
    if len(zone_breakdown) == 0:
        draw.text((zx0, zy), "No zone data available", font=_img_font(12), fill=(140, 140, 140))
    for i, (zone_name, amt) in enumerate(zone_breakdown.items()):
        color = zone_palette[i % len(zone_palette)]
        bar_w = max(4, int(bar_max_w * (float(amt) / max_zone_amount))) if max_zone_amount else 4
        draw.text((zx0, zy + 6), str(zone_name), font=_img_font(12), fill=ink)
        draw.rounded_rectangle([bar_x0, zy, bar_x0 + bar_max_w, zy + 22], radius=4, fill=(238, 238, 240))
        draw.rounded_rectangle([bar_x0, zy, bar_x0 + bar_w, zy + 22], radius=4, fill=color)
        draw.text((bar_x0 + bar_max_w + 12, zy + 4), f"₹{float(amt):,.0f}", font=_img_font(11), fill=ink)
        zy += 34

    # --- Top 5 Highest-Value CIFs leaderboard ---
    y += panel_h + gap
    draw.rounded_rectangle([pad, y, width - pad, y + top5_h], radius=12, fill="white", outline=(225, 225, 230))
    draw.text((pad + 20, y + 16), "TOP 5 HIGHEST-VALUE CIFs", font=_img_font(13), fill=(70, 70, 70))
    ry = y + 44
    for i, (_, r) in enumerate(top5.iterrows()):
        rank = medals[i] if i < 3 else str(i + 1)
        draw.text((pad + 24, ry + top5_row_h / 2), rank, font=_img_font(18 if i < 3 else 13), fill=gold if i < 3 else (140, 140, 140), anchor="lm")
        draw.text((pad + 70, ry + 10), str(r.get("CIF", "")), font=_img_font(14), fill=ink)
        detail = f"{r.get('Zone2', '')} / {r.get('Sub Region2', '')} · Slab {r.get('Slab', '')}"
        reason_text = str(r.get("Violation Reason", ""))
        if reason_text:
            detail += f"  ·  {reason_text}"
        draw.text((pad + 70, ry + 28), detail, font=_img_font(10), fill=(130, 130, 130))
        amt_text = f"₹{float(r.get('Grand Total', 0)):,.0f}"
        draw.text((width - pad - 24, ry + top5_row_h / 2), amt_text, font=_img_font(15), fill=crimson, anchor="rm")
        if i < len(top5) - 1:
            draw.line([pad + 20, ry + top5_row_h - 4, width - pad - 20, ry + top5_row_h - 4], fill=(238, 238, 240), width=1)
        ry += top5_row_h

    remaining = total_count - len(top5)
    footer_note = f"+ {remaining:,} more violation(s) in the full export" if remaining else "Full detail available in the Excel export"
    draw.text((pad, height - 30), f"{footer_note}  ·  Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(11), fill=(150, 150, 150))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


def rcpt_cxn_to_image_bytes(df: pd.DataFrame) -> io.BytesIO:
    """
    Cancelled-receipts PNG: same cream 'ledger' theme (dark header, rotated
    red stamp) as before, but now broken down BY PAYMENT MODE — how much
    of the cancellation volume is Cash, Cheque/DD, Online, RTGS — instead
    of listing individual rows. Each mode gets its own bar showing count
    and cancelled amount.
    """
    from PIL import Image, ImageDraw

    cream = (250, 246, 235)
    ink = (40, 35, 30)
    red = (176, 42, 42)
    width = 1100

    if df.empty:
        height = 320
        base = Image.new("RGB", (width, height), cream)
        draw = ImageDraw.Draw(base)
        draw.rectangle([0, 0, width, 10], fill=red)
        draw.text((width / 2, 130), "CLEAN REGISTER", font=_img_font(28), fill=(39, 120, 96), anchor="mm")
        draw.text((width / 2, 168), "No cancelled receipts this period", font=_img_font(13), fill=(100, 95, 85), anchor="mm")
        draw.text((40, height - 28), f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(11), fill=(150, 145, 135))
        buf = io.BytesIO()
        base.save(buf, format="PNG")
        buf.seek(0)
        return buf

    mode_colors = {
        "AIRTEL / CASH": (27, 138, 122),
        "CHQ / DD": (122, 61, 145),
        "ONLINE_PAYMENT": (39, 100, 156),
        "RTGS": (58, 47, 38),
        "Cashfree": (39, 100, 156),
        "Payment Gateway": (122, 61, 145),
        "QR Code": (27, 138, 122),
        "CashfreeQR": (176, 42, 42),
        UNMAPPED_LABEL: (140, 120, 95),
    }

    by_mode = (
        df.assign(Amount=pd.to_numeric(df["Amount"], errors="coerce").fillna(0))
        .groupby("PaymentMode", dropna=False)
        .agg(count=("Amount", "size"), amount=("Amount", "sum"))
        .reset_index()
        .sort_values("count", ascending=False)
    )
    total_count = int(len(df))
    total_amount = float(by_mode["amount"].sum())
    max_count = max(int(by_mode["count"].max()), 1)

    pad = 40
    header_h = 110
    kpi_h = 100
    gap = 26
    bar_row_h = 62
    section_label_h = 30
    height = header_h + gap + kpi_h + gap + section_label_h + len(by_mode) * bar_row_h + 70

    base = Image.new("RGB", (width, height), cream)
    draw = ImageDraw.Draw(base)

    draw.rectangle([0, 0, width, header_h], fill=(58, 47, 38))
    draw.text((pad, 22), "CANCELLED RECEIPTS — BY PAYMENT MODE", font=_img_font(24), fill=cream)
    draw.text((pad, 62), f"{total_count} cancelled receipts  ·  {datetime.now().strftime('%d-%b-%Y %H:%M')}", font=_img_font(13), fill=(210, 195, 175))

    stamp = Image.new("RGBA", (200, 82), (0, 0, 0, 0))
    sd = ImageDraw.Draw(stamp)
    sd.rounded_rectangle([4, 4, 196, 78], radius=10, outline=red, width=5)
    sd.text((100, 41), "CANCELLED", font=_img_font(20), fill=red, anchor="mm")
    stamp = stamp.rotate(-14, expand=True)
    base.paste(stamp, (width - pad - stamp.width + 26, 10), stamp)

    # KPI strip: total count + total amount
    y = header_h + gap
    kpi_w = (width - 2 * pad - 20) // 2
    for i, (label, val, is_amount) in enumerate([
        ("TOTAL CANCELLED RECEIPTS", total_count, False),
        ("TOTAL CANCELLED AMOUNT", total_amount, True),
    ]):
        x = pad + i * (kpi_w + 20)
        draw.rounded_rectangle([x, y, x + kpi_w, y + kpi_h], radius=12, outline=(210, 198, 178), width=2, fill=(255, 253, 246))
        draw.rectangle([x, y, x + 6, y + kpi_h], fill=red)
        val_str = f"₹{val:,.0f}" if is_amount else f"{val:,}"
        draw.text((x + 24, y + 26), val_str, font=_img_font(26), fill=ink)
        draw.text((x + 24, y + 64), label, font=_img_font(11), fill=(120, 110, 95))

    y += kpi_h + gap
    draw.text((pad, y), "BREAKDOWN BY PAYMENT MODE", font=_img_font(14), fill=(90, 78, 60))
    y += section_label_h

    label_w = 190
    bar_x0 = pad + label_w
    bar_max_w = width - pad - bar_x0 - 130

    for _, r in by_mode.iterrows():
        mode = str(r["PaymentMode"]) if pd.notna(r["PaymentMode"]) else UNMAPPED_LABEL
        cnt, amt = int(r["count"]), float(r["amount"])
        color = mode_colors.get(mode, (140, 120, 95))

        draw.text((pad, y + bar_row_h / 2 - 8), mode, font=_img_font(14), fill=ink, anchor=None)
        by0, by1 = y + 12, y + bar_row_h - 22
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + bar_max_w, by1], radius=6, fill=(238, 231, 214))
        w = int(bar_max_w * (cnt / max_count))
        w = max(w, 4)
        draw.rounded_rectangle([bar_x0, by0, bar_x0 + w, by1], radius=6, fill=color)
        draw.text((bar_x0 + bar_max_w + 14, (by0 + by1) / 2), f"{cnt:,}", font=_img_font(15), fill=ink, anchor="lm")
        draw.text((bar_x0, by1 + 4), f"₹{amt:,.0f} cancelled", font=_img_font(10), fill=(120, 110, 95))
        y += bar_row_h

    draw.text((pad, height - 30), "Generated automatically — DHC Working Automation", font=_img_font(11), fill=(150, 145, 135))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


CASH_DAILY_LIMIT = 195000        # ₹1.95 lakh — per CIF, per day
CASH_GRAND_TOTAL_LIMIT = 1000000  # ₹10 lakh — per CIF, across the whole period


def build_cash_mode_validation_summary(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Compliance check on cash-mode (Airtel/Cash) receipts. A CIF is
    flagged if ANY of these three independent conditions hold — each
    checked separately, not combined into one number:

      1. Per-day limit: some SINGLE day's cumulative cash collected from
         that CIF crosses Rs 1,95,000 (Income Tax Act S.269ST buffer). A
         CIF who stays under Rs 1,95,000 every day is NOT flagged for
         this even if the days add up to a large total over time — e.g.
         Rs 1.9L on day 1 + Rs 1L on day 2 is fine, since neither day
         alone crosses the limit.
      2. Grand Total limit: the CIF's cumulative cash total across the
         WHOLE period exceeds Rs 10,00,000, regardless of how it's
         spread across days.
      3. Disabled CIF still receipted: the CIF is on the Ag Level or CIF
         Level disable list and has ANY cash receipt at all, any amount
         — a disabled customer paying in cash is itself the compliance
         issue, independent of the amount.

    The 'Violation Reason' column records exactly which of the three
    triggered (a CIF can trigger more than one at once), so the report
    is self-explanatory rather than a bare numbers table.

    Airtel counts as cash here too (cash-collection channel, not a
    separate payment method), using the mapped Mode column plus a
    raw-text fallback for any Airtel/Cash-worded MODEOFPAYMENT variant
    MODE_MAP doesn't recognise yet.
    """
    is_cash_mode = dcr_tab["Mode"].isin(["AIRTEL", "CASH"])
    raw_mode_text = dcr_tab["MODEOFPAYMENT"].astype(str).str.strip().str.upper()
    is_cash_like_raw = raw_mode_text.str.contains("AIRTEL", na=False) | raw_mode_text.str.contains("CASH", na=False)

    cash = dcr_tab[is_cash_mode | is_cash_like_raw].copy()
    cols = ["CIF", "Zone2", "Sub Region2", "Slab", "Grand Total", "Violation Reason"]
    if cash.empty:
        return pd.DataFrame(columns=cols)

    cash["Receipt Date"] = pd.to_datetime(cash["RECEIPT ENTER DATE"]).dt.normalize()
    cash["_is_disabled"] = cash["Ag Level cash mode"].notna() | cash["CIF LEVEL"].notna()

    group_cols = ["CIF", "Zone", "Sub Region", "Slab"]
    pivot = pd.pivot_table(
        cash, index=group_cols, columns="Receipt Date", values="AMOUNTPAID", aggfunc="sum",
    )
    date_cols = list(pivot.columns)
    pivot["Grand Total"] = pivot.sum(axis=1, skipna=True)

    disabled_flag = cash.groupby(group_cols)["_is_disabled"].any().reindex(pivot.index, fill_value=False)
    exceeds_daily = (pivot[date_cols].fillna(0) > CASH_DAILY_LIMIT).any(axis=1)
    exceeds_grand_total = pivot["Grand Total"] > CASH_GRAND_TOTAL_LIMIT

    is_violation = exceeds_daily | exceeds_grand_total | disabled_flag
    pivot = pivot[is_violation].copy()

    reasons = []
    for idx in pivot.index:
        r = []
        if exceeds_daily.loc[idx]:
            r.append(f"Per-Day > ₹{CASH_DAILY_LIMIT:,}")
        if exceeds_grand_total.loc[idx]:
            r.append(f"Grand Total > ₹{CASH_GRAND_TOTAL_LIMIT:,}")
        if disabled_flag.loc[idx]:
            r.append("Disabled CIF — Cash Received")
        reasons.append("; ".join(r))
    pivot["Violation Reason"] = reasons

    pivot = pivot.reset_index().rename(columns={"Zone": "Zone2", "Sub Region": "Sub Region2"})
    return pivot


def cash_mode_mail_counts(dcr_tab: pd.DataFrame) -> dict:
    """
    The four Cash Mode Validation headline counts used in the daily mail
    (matches the reference report's exact four bullets):
      1. 'monthly_195' — CIFs whose cash Grand Total for the period
         crosses Rs 1,95,000 (an informational monthly cut at the same
         Rs 1.95L level as the daily limit — broader than #3, since a CIF
         can cross this by accumulating over many small-value days).
      2. 'disabled' — CIFs on the Ag Level/CIF Level disable list with
         ANY cash receipt at all.
      3. 'daily_195' — CIFs with some single day's cash crossing
         Rs 1,95,000 (Income Tax Act S.269ST buffer).
      4. 'monthly_10L' — CIFs whose cash Grand Total for the period
         crosses Rs 10,00,000.
    Shares the same cash-mode filter (mapped Mode plus raw-text
    Airtel/Cash fallback) as build_cash_mode_validation_summary, so the
    two never disagree on which rows count as "cash".
    """
    is_cash_mode = dcr_tab["Mode"].isin(["AIRTEL", "CASH"])
    raw_mode_text = dcr_tab["MODEOFPAYMENT"].astype(str).str.strip().str.upper()
    is_cash_like_raw = raw_mode_text.str.contains("AIRTEL", na=False) | raw_mode_text.str.contains("CASH", na=False)
    cash = dcr_tab[is_cash_mode | is_cash_like_raw].copy()

    zero_counts = {"monthly_195": 0, "disabled": 0, "daily_195": 0, "monthly_10L": 0}
    if cash.empty:
        return zero_counts

    cash["Receipt Date"] = pd.to_datetime(cash["RECEIPT ENTER DATE"]).dt.normalize()
    cash["_is_disabled"] = cash["Ag Level cash mode"].notna() | cash["CIF LEVEL"].notna()

    group_cols = ["CIF", "Zone", "Sub Region", "Slab"]
    pivot = pd.pivot_table(cash, index=group_cols, columns="Receipt Date", values="AMOUNTPAID", aggfunc="sum")
    date_cols = list(pivot.columns)
    pivot["Grand Total"] = pivot.sum(axis=1, skipna=True)
    disabled_flag = cash.groupby(group_cols)["_is_disabled"].any().reindex(pivot.index, fill_value=False)

    return {
        "monthly_195": int((pivot["Grand Total"] > CASH_DAILY_LIMIT).sum()),
        "disabled": int(disabled_flag.sum()),
        "daily_195": int((pivot[date_cols].fillna(0) > CASH_DAILY_LIMIT).any(axis=1).sum()),
        "monthly_10L": int((pivot["Grand Total"] > CASH_GRAND_TOTAL_LIMIT).sum()),
    }


def build_rcpt_cxn(cancellation_report: pd.DataFrame) -> pd.DataFrame:
    """
    Cancelled-receipt register, sourced directly from the dedicated
    Receipt Cancellation Report — no longer derived from the DCR extract's
    Status == 'Cxn' rows. This report already carries every cancelled
    receipt with its real PaymentMode (Cashfree, Payment Gateway, QR Code,
    CashfreeQR, RTGS, ...), so nothing needs mapping or re-deriving here.

    'Status' (the cancellation reason, e.g. Duplicate Receipt) isn't
    present in this source report, so it's left blank alongside 'Remarks'
    for mam to fill in by hand — this sheet is a judgment log, not a
    formula output.
    """
    out = cancellation_report.copy()
    out["Status"] = ""
    out["Remarks"] = ""
    cols = ["ReceiptNo", "ReceiptDate", "Amount", "ReceiptStatus", "ReceiptType",
            "PaymentMode", "Zone", "AgreementNo", "CustomerName", "ReceiptCreatedDate",
            "Status", "Remarks"]
    return out[[c for c in cols if c in out.columns]]


def _stats_row(label: str, stats: dict) -> list:
    """Helper to flatten a stats dict into a row for preview DataFrame."""
    row = [label]
    for bucket in TAT_ORDER:
        row += [stats["by_bucket"][bucket]["count"], round(stats["by_bucket"][bucket]["value"], 2)]
    row += [stats["total_count"], round(stats["total_value"], 2)]
    return row


def zone_tat_matrix_to_dataframe(matrix: dict) -> pd.DataFrame:
    """Flattens the Zone x <breakdown> x TAT structure into a preview-friendly DataFrame."""
    cols = ["Zone / Category"]
    for bucket in TAT_ORDER:
        cols += [f"{TAT_DISPLAY[bucket]} Count", f"{TAT_DISPLAY[bucket]} Value"]
    cols += ["Total Count", "Total Value (Cr)"]
    rows = []
    for block in matrix["zones"]:
        rows.append(_stats_row(block["zone"], block["subtotal"]))
        for code, label, stats in block["breakdown"]:
            rows.append(_stats_row(f"   {label}", stats))
    rows.append(_stats_row("GRAND TOTAL", matrix["grand_total"]))
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
# Email summary — plain-text (subject, body) pair for the "Email Summary"
# mailto button. This is a quick heads-up, not a re-print of the tables, so
# every section is kept to a handful of short lines with blank lines around
# each one — meant to be skimmed on a phone, not read like a report.
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Email summary — plain-text (subject, body) pair for the "Email Summary"
# mailto button, formatted to match the team's existing "Daily Health Check"
# mail exactly (same section headings/bullet wording as the reference mail).
# Every figure the pipeline can actually compute from the DCR is filled in
# automatically; the handful of items that come from somewhere other than
# the DCR file (an LMS spot-check, Delq/CFE allocation status, mobile-number
# validation, FVC charge rollout, and the deposition side of the last
# section — there's no deposit-date field anywhere in this pipeline, only
# receipting-date TAT) are left as a clearly marked [Fill manually] so nothing
# gets silently invented.
# ---------------------------------------------------------------------------
FILL_MANUALLY = "[Fill manually]"


def _tat_exceeded_count(dcr_tab: pd.DataFrame) -> int:
    """
    Mode-specific TAT breach count (see TAT_LIMIT_DAYS / compute_tat_exceeded):
    Cash/Airtel/Online/RTGS over 1 day, Cheque/DD over 3 days. Replaces the
    old fixed "5-10 Days + >10 Days" bucket count, which applied the same
    day-range to every payment mode regardless of its actual TAT policy.
    """
    full = _dcr_with_ageing(dcr_tab)
    return int(compute_tat_exceeded(full).sum())


def _bounced_chq_dd_count(receipt_made_summary: dict) -> int:
    """Cheque/DD rows marked Bounced, across every RECEIPT STATUS group."""
    total = 0
    for group in receipt_made_summary["full"]["groups"]:
        for row in group["rows"]:
            if row["mode"] in ("CHEQUE", "DD"):
                total += row["counts"].get("Bounced", 0)
    return total


def build_mail_summary(
    receipt_made_summary: dict,
    rtgs_summary: dict,
    cash_mode_validation_df: pd.DataFrame,
    delay_summary: dict,
    rcpt_cxn_df: pd.DataFrame,
    dcr_tab: pd.DataFrame,
    report_date: str | None = None,
    zone_label: str = "Overall",
    chq_brs: dict | None = None,
    air_brs: dict | None = None,
    cash_brs: dict | None = None,
) -> tuple[str, str]:
    """
    Builds (subject, body) in the exact "Daily Health Check" format the
    team already sends manually — same headings, same bullet wording —
    with every figure the pipeline can compute pulled live from the same
    summary dicts/DataFrames already in st.session_state, so the numbers
    always match the exported Excel files. See FILL_MANUALLY above for
    what still needs a human to fill in.

    chq_brs / air_brs are the dicts from process_brs_for_mail (Cheque/DD
    and Airtel Cash BRS respectively) — when supplied, they fill in the
    Delay in Receipting / Deposition section for real instead of leaving
    it as [Fill manually]. The Airtel Cash BRS file covers both the "Cash
    Mode" and "Airtel Deposit" lines (there's no separate physical-cash
    BRS — Airtel and Cash are reconciled together in one "Airtel Cash BRS"
    report), so both pairs of lines share the same air_brs numbers.

    dcr_tab should already be filtered to the desired zone before calling
    this (see filter_by_broad_zone) — zone_label is just for the
    subject/heading text, it doesn't filter anything itself.
    """
    date_str = report_date or datetime.now().strftime("%d-%b-%Y")
    try:
        month_label = datetime.strptime(date_str, "%d-%b-%Y").strftime("%b'%y")
    except ValueError:
        month_label = datetime.now().strftime("%b'%y")
    zone_suffix = "" if zone_label.strip().upper() == "OVERALL" else f" — {zone_label} Zone"
    subject = f"Daily Health Check — {date_str}{zone_suffix}"

    rms_full = receipt_made_summary["full"]
    online = rtgs_summary.get("online_source_block", {"rows": [], "total": 0})
    auto_manual_all = compute_auto_manual_block(dcr_tab)
    rtgs_auto_manual = compute_rtgs_auto_manual_block(dcr_tab)
    receipts_overview = compute_receipts_overview_block(dcr_tab)
    sync_block = compute_receipt_status_sync_block(dcr_tab)
    cmv = cash_mode_mail_counts(dcr_tab)
    tat_exceeded = _tat_exceeded_count(dcr_tab)
    bounced_chq_dd = _bounced_chq_dd_count(receipt_made_summary)
    cxn_count = len(rcpt_cxn_df)

    online_lines = [f"   - {name}: {cnt:,}" for name, cnt in online["rows"]] or ["   - No online receipts this period"]
    auto_manual_all_lines = [f"   - {name}: {cnt:,}" for name, cnt in auto_manual_all["rows"]]
    auto_manual_lines = [f"   - {name}: {cnt:,}" for name, cnt in rtgs_auto_manual["rows"]]
    status_lines = [f"   - {s.title()}: {sync_block['totals'][s]:,}" for s in sync_block["status_order"]]

    body = "\n".join([
        "Dear Sir,",
        "",
        f"Please find the Daily Health Check on below key Parameters{zone_suffix},",
        "",
        "Receipts Overview :-",
        "",
        f"   - No of Receipts: {receipts_overview['total']:,}",
        f"   - Auto Receipts: {receipts_overview['auto']:,}",
        f"   - Manual Receipts: {receipts_overview['manual']:,}",
        *[f"   - {label}: {cnt:,}" for label, cnt in receipts_overview["receipt_status_rows"]],
        f"   - Sync Success: {receipts_overview['sync_success']:,}",
        f"   - Sync Awaited: {receipts_overview['sync_awaited']:,}",
        "",
        "Receipt Sync Status :-",
        "",
        *status_lines,
        f"   - Sync Success (Approved/Batched/Realized): {sync_block['success_total']:,}",
        f"   - Sync Awaiting (Batched/Pending): {sync_block['awaiting_total']:,}",
        "",
        "Pay Mode — Auto Receipt / Manual (All Modes) :-",
        "",
        *auto_manual_all_lines,
        f"   - Total: {auto_manual_all['total']:,}",
        "",
        "- Random check is done in LMS to identify any deviations in EMI OS / Charges due / Cash mode enable.",
        f"- {bounced_chq_dd} Cheque payment got updated as bounced in system ( Same got updated in SOA / CCP )",
        f"- Delq accounts Allocation to CFE. ( {FILL_MANUALLY} )",
        f"- CIF Level Same DPD Accounts Allocation to Single CFE. ( {FILL_MANUALLY} )",
        "",
        "Receipt Particulars :",
        "",
        f"1. TAT Exceeded Receipt Count is {tat_exceeded}",
        f"2. Receipt Didn't got flown into LMS :- {FILL_MANUALLY}",
        f"3. {online['total']:,} Online Receipts – Payment Source Bifurcation is given below.",
        *online_lines,
        "",
        "RTGS Summary :",
        "",
        f"- {rtgs_auto_manual['total']:,} RTGS Payments received this period.",
        "",
        "Mobile Number Validation :-",
        "",
        f"- Employee Mobile Number got updated for Online Bitly Payments made by Customer – {FILL_MANUALLY}.",
        f"- Employee Mobile Number Used for Receipting Purpose – {FILL_MANUALLY}.",
        f"- Similar mobile number updated in multiple Receipts – {FILL_MANUALLY}.",
        "",
        "Cash Mode Validation :-",
        "",
        f"- CIF level Per Month Cash Collection exceeds 1.95 L – {cmv['monthly_195']}",
        f"- Cash Collection is done on the Cash mode disabled Accounts – {cmv['disabled']} ({month_label}).",
        f"- CIF Level Per Day cash collection exceeds 1.95 L – {cmv['daily_195']}",
        f"- CIF Level Per Month cash collection exceeds 10 L – {cmv['monthly_10L']}",
        "",
        "FVC Charge Implementation :-",
        "",
        f"- FVC Charge Implementation – {FILL_MANUALLY}.",
        "",
        "Receipt Cancellation :-",
        "",
        f"- Receipt Cancellation Count is {cxn_count if cxn_count else 'Nil'}",
        "",
        "Pay Mode RTGS — Auto Receipt / Manual :-",
        "",
        *auto_manual_lines,
        f"   - Total: {rtgs_auto_manual['total']:,}",
        "",
        "Delay in Receipting / Deposition :-",
        "",
        f"- Cash Mode – Delay in Deposition Count is {cash_brs['delay']['delay_deposition'] if cash_brs else (air_brs['delay']['delay_deposition'] if air_brs else FILL_MANUALLY)}.",
        f"- Cash Mode – Delay in Receipting Count is {cash_brs['delay']['delay_receipting'] if cash_brs else (air_brs['delay']['delay_receipting'] if air_brs else FILL_MANUALLY)}.",
        f"- Airtel Deposit – Delay in Deposition Count is {air_brs['delay']['delay_deposition'] if air_brs else FILL_MANUALLY}.",
        f"- Airtel Deposit – Delay in Receipting Count is {air_brs['delay']['delay_receipting'] if air_brs else FILL_MANUALLY}.",
        f"- CHQ / DD – Delay in Deposition Count is {chq_brs['delay']['delay_deposition'] if chq_brs else FILL_MANUALLY}.",
        f"- CHQ / DD – Delay in Receipting Count is {chq_brs['delay']['delay_receipting'] if chq_brs else FILL_MANUALLY}.",
        "",
        "Regards,",
    ])
    return subject, body


def resize_png_for_mail(png_bytes: bytes, max_width: int = 560) -> bytes:
    """
    Downscales a dashboard PNG (natively ~1100-1160px wide, sized for a
    full-screen Streamlit preview) to a mail-friendly width, preserving
    aspect ratio. Only ever shrinks — if the source is already narrower
    than max_width it's returned untouched. Used for the Outlook "with
    charts" mail so images don't dominate the message the way they do at
    full size.
    """
    from PIL import Image

    img = Image.open(io.BytesIO(png_bytes))
    if img.width <= max_width:
        return png_bytes
    ratio = max_width / img.width
    new_size = (max_width, max(1, round(img.height * ratio)))
    resized = img.resize(new_size, Image.LANCZOS)
    out = io.BytesIO()
    resized.save(out, format="PNG")
    return out.getvalue()


def build_mail_html_summary(
    receipt_made_summary: dict,
    rtgs_summary: dict,
    cash_mode_validation_df: pd.DataFrame,
    delay_summary: dict,
    rcpt_cxn_df: pd.DataFrame,
    dcr_tab: pd.DataFrame,
    receipt_made_cid: str = "receipt_made_report",
    rtgs_cid: str = "rtgs_summary",
    delay_cid: str = "delay_summary",
    rcpt_cxn_cid: str = "rcpt_cxn",
    zone_overview_cid: str = "zone_overview",
    overview_cid: str = "receipts_overview",
    sync_cid: str = "receipt_sync_status",
    chq_brs_cid: str = "chq_brs_status",
    air_brs_cid: str = "air_brs_status",
    cash_brs_cid: str = "cash_brs_status",
    img_width: int = 560,
    report_date: str | None = None,
    zone_label: str = "Overall",
    chq_brs: dict | None = None,
    air_brs: dict | None = None,
    cash_brs: dict | None = None,
) -> tuple[str, str]:
    """
    HTML counterpart of build_mail_summary — same "Daily Health Check"
    section headings/wording, as HTML, with each relevant section's own
    dashboard image sitting right under its bullet points (Receipt Made
    Report under the Receipt Made section, Delay-in-Rcpting heatmap under
    the Delay in Receipting / Deposition section, and the cancelled-
    receipts register under Receipt Cancellation — RTGS's own image
    covers the Online Payment Channels split referenced in bullet 3).

    Every dynamic value is passed through html.escape() before going into
    the template. This is the actual fix for "Delay in Receipting and
    Receipt Cancellation images not showing": TAT_DISPLAY has entries
    like "< 4 Days" and "> 10 Days" with real unescaped angle brackets in
    them, and the old version dropped those straight into the HTML
    (`f"<li>{TAT_DISPLAY[b]}: ..."`). "< 4 Days" is a `<` immediately
    followed by a space, which most browsers silently recover from as
    literal text — but Outlook's classic HTML renderer (it's Word's
    engine, not a browser engine, and it is notoriously less forgiving of
    malformed markup) can lose the rest of the document from that point
    on once it hits enough of these, which is exactly why the two
    sections *after* the one repeating this pattern first (RTGS, itself
    fine) were the ones going missing — Delay in Receipting reuses the
    same TAT_DISPLAY labels a second time, and everything from there
    down, including its own `<img>` tag and the whole Receipt
    Cancellation section after it, silently disappeared. html.escape()
    turns "< 4 Days" into "&lt; 4 Days" so there's no stray "<" left
    anywhere in the body for the renderer to trip on.

    img_width is both the CSS width and the HTML width attribute on each
    <img> — Outlook's classic HTML renderer honours the width attribute
    more reliably than max-width alone, so both are set.

    The <img> tags reference cid: URIs — the caller (see
    compose_outlook_mail_with_images) must attach the four PNGs, already
    resized with resize_png_for_mail, with matching Content-IDs
    (receipt_made_cid / rtgs_cid / delay_cid / rcpt_cxn_cid) for the
    images to actually show up once this is set as an Outlook MailItem's
    HTMLBody.
    """
    subject, _ = build_mail_summary(
        receipt_made_summary, rtgs_summary, cash_mode_validation_df, delay_summary, rcpt_cxn_df, dcr_tab, report_date, zone_label,
        chq_brs, air_brs, cash_brs,
    )
    date_str = report_date or datetime.now().strftime("%d-%b-%Y")
    try:
        month_label = datetime.strptime(date_str, "%d-%b-%Y").strftime("%b'%y")
    except ValueError:
        month_label = datetime.now().strftime("%b'%y")

    e = html.escape  # local alias — every dynamic value below goes through this

    auto_manual_all = compute_auto_manual_block(dcr_tab)
    cmv = cash_mode_mail_counts(dcr_tab)
    fill = f'<i style="color:#b45309;">{e(FILL_MANUALLY)}</i>'
    cash_violation_total = cmv["monthly_195"] + cmv["disabled"] + cmv["daily_195"] + cmv["monthly_10L"]

    def _auto_manual_box(title: str, block: dict, width: int = 260) -> str:
        rows_html = "".join(
            f'<tr><td style="padding:3px 10px 3px 0;color:#374151;">{e(name)}</td>'
            f'<td style="padding:3px 0;text-align:right;font-weight:600;color:#0f172a;">{cnt:,}</td></tr>'
            for name, cnt in block["rows"]
        )
        return f"""
        <table width="{width}" cellpadding="0" cellspacing="0" style="width:{width}px;border:1px solid #e2e8f0;border-radius:8px;background:#f8fafc;">
          <tr><td style="padding:10px 14px;">
            <div style="font-weight:600;font-size:13px;color:#0f172a;margin-bottom:6px;">{e(title)}</div>
            <table width="100%" cellpadding="0" cellspacing="0" style="width:100%;border-collapse:collapse;font-size:13px;">
              {rows_html}
              <tr style="border-top:1px solid #e2e8f0;">
                <td style="padding:5px 10px 0 0;font-weight:700;color:#0f172a;">Total</td>
                <td style="padding:5px 0 0 0;text-align:right;font-weight:700;color:#0f172a;">{block['total']:,}</td>
              </tr>
            </table>
          </td></tr>
        </table>"""

    auto_manual_all_box = _auto_manual_box("Pay Mode — Auto Receipt / Manual (All Modes)", auto_manual_all)

    body_font = "font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#1f2937;line-height:1.7;"
    section = "margin:22px 0;"
    heading = "margin:0 0 8px 0;color:#0f172a;"

    def img_tag(cid: str, alt: str, width: int = img_width) -> str:
        style = f"width:{width}px;max-width:100%;border:1px solid #e2e8f0;border-radius:8px;margin-top:12px;display:block;"
        return f'<img src="cid:{cid}" width="{width}" style="{style}" alt="{e(alt)}" />'

    pair_gap = 12
    pair_w = max(320, int(img_width * 0.68))

    html_body = f"""
    <div style="{body_font}">
      <p>Dear Sir,</p>
      <p>Please find the Daily Health Check on below key Parameters{'' if zone_label.strip().upper() == 'OVERALL' else f' — {e(zone_label)} Zone'},</p>

      <div style="{section}margin-top:0;">
        <table cellpadding="0" cellspacing="0" style="margin-top:0;">
          <tr>
            <td style="padding-right:{pair_gap}px;vertical-align:top;">
              {img_tag(overview_cid, "Receipts Overview", pair_w)}
            </td>
            <td style="vertical-align:top;">
              {img_tag(sync_cid, "Receipt Sync Status", pair_w)}
            </td>
          </tr>
        </table>
      </div>

      <div style="{section}">
        {auto_manual_all_box}
      </div>

      <div style="{section}">
        <h3 style="{heading}">Receipt Particulars :</h3>
        <table cellpadding="0" cellspacing="0" style="margin-top:12px;">
          <tr>
            <td style="padding-right:{pair_gap}px;vertical-align:top;">
              {img_tag(receipt_made_cid, "Receipt Particulars", pair_w)}
            </td>
            <td style="vertical-align:top;">
              {img_tag(zone_overview_cid, "Zone Overview — Mode split, TAT exceeded, Bounced/Cancelled", pair_w)}
            </td>
          </tr>
        </table>
      </div>

      <div style="{section}">
        <h3 style="{heading}">Receipt Cancellation :-</h3>
        {img_tag(rcpt_cxn_cid, "Cancelled Receipts by Mode")}
      </div>

      <div style="{section}">
        <h3 style="{heading}">Delay in Receipting / Deposition :-</h3>
        {img_tag(chq_brs_cid, "Cheque/DD BRS Status") if chq_brs else ""}
        {img_tag(air_brs_cid, "Airtel Cash BRS Status") if air_brs else ""}
        {img_tag(cash_brs_cid, "Cash BRS Status") if cash_brs else ""}
      </div>

      <div style="{section}">
        <h3 style="{heading}">Mobile Number Validation :-</h3>
        <ul style="margin:0;padding-left:20px;">
          <li>Employee Mobile Number got updated for Online Bitly Payments made by Customer – {fill}.</li>
          <li>Employee Mobile Number Used for Receipting Purpose – {fill}.</li>
          <li>Similar mobile number updated in multiple Receipts – {fill}.</li>
        </ul>
      </div>

      <div style="{section}">
        <h3 style="{heading}">Cash Mode Validation :-</h3>
        <div style="padding:10px 14px;border-radius:8px;border:1px solid {'#fecaca' if cash_violation_total else '#bbf7d0'};background:{'#fef2f2' if cash_violation_total else '#f0fdf4'};margin-bottom:10px;">
          <b style="color:{'#b91c1c' if cash_violation_total else '#15803d'};">{f'⚠ Cash Mode Violation(s) Found — {cash_violation_total} total' if cash_violation_total else '✓ No Cash Mode Violations'}</b>
        </div>
        <ul style="margin:0;padding-left:20px;">
          <li>CIF level Per Month Cash Collection exceeds 1.95 L – <b style="color:{'#b91c1c' if cmv['monthly_195'] else '#0f172a'};">{cmv['monthly_195']}</b></li>
          <li>Cash Collection is done on the Cash mode disabled Accounts – <b style="color:{'#b91c1c' if cmv['disabled'] else '#0f172a'};">{cmv['disabled']}</b> ({e(month_label)}).</li>
          <li>CIF Level Per Day cash collection exceeds 1.95 L – <b style="color:{'#b91c1c' if cmv['daily_195'] else '#0f172a'};">{cmv['daily_195']}</b></li>
          <li>CIF Level Per Month cash collection exceeds 10 L – <b style="color:{'#b91c1c' if cmv['monthly_10L'] else '#0f172a'};">{cmv['monthly_10L']}</b></li>
        </ul>
      </div>

      <p>Regards,</p>
    </div>
    """
    return subject, html_body




def compose_outlook_mail_with_images(subject: str, html_body: str, image_paths: list[tuple[str, str]]) -> None:
    """
    Opens (does NOT send) a new mail draft directly in classic Outlook via
    COM automation, with one or more images embedded inline in the HTML
    body. image_paths is a list of (content_id, file_path) pairs — each
    content_id must match a "cid:{content_id}" reference already inside
    html_body (see build_mail_html_summary).

    Deliberately uses win32com's "Outlook.Application" COM object rather
    than a mailto: link: mailto can't attach or embed anything, and — as
    a side effect — the classic Outlook.Application COM object is only
    exposed by classic (desktop) Outlook, not the new Outlook app, so this
    also sidesteps the "which mail app opens" ambiguity a mailto: link
    runs into on machines with both installed.

    Requires pywin32 (`pip install pywin32`) and a Windows machine with
    classic Outlook installed — raises ImportError / whatever COM error
    Outlook raises otherwise, so the caller can show a clear message
    instead of this failing silently.

    Calls pythoncom.CoInitialize() first: Streamlit runs each button's
    callback on a worker thread (not the main thread), and COM requires
    the calling thread to be initialized before any COM object is
    created on it — win32com.client.Dispatch doesn't do this itself, so
    skipping it is what produces "CoInitialize has not been called."
    CoUninitialize() in `finally` releases that thread's COM apartment
    again once Outlook has the draft open.
    """
    import pythoncom
    import win32com.client as win32  # local import: only needed by this one Windows-only path

    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0 = olMailItem
        mail.Subject = subject
        mail.BodyFormat = 2  # olFormatHTML — set explicitly before HTMLBody so Outlook doesn't guess
        mail.HTMLBody = html_body
        for content_id, path in image_paths:
            attachment = mail.Attachments.Add(path)
            # PR_ATTACH_CONTENT_ID (MAPI) — lets the HTML body's cid: reference
            # find this attachment and render it inline instead of as a
            # separate file.
            attachment.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F", content_id
            )
            # PR_ATTACHMENT_HIDDEN — keeps it purely inline, so it never also
            # shows up a second time as a regular file attachment underneath.
            attachment.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True
            )
        mail.Display()  # opens the draft for the user to review/attach the workbook/send — never sends automatically
    finally:
        pythoncom.CoUninitialize()


def receipt_made_table_to_dataframe(table: dict, status_cols: list[str]) -> pd.DataFrame:
    """Flattens the receipt status x mode table into a preview DataFrame."""
    cols = ["Status", "Mode"] + [c.upper() for c in status_cols] + ["Grand Total"]
    rows = []
    for group in table["groups"]:
        for i, r in enumerate(group["rows"]):
            label = group["status"] if i == 0 else ""
            rows.append([label, r["mode"]] + [r["counts"][c] for c in status_cols] + [r["total"]])
    rows.append(["GRAND TOTAL", ""] + [table["grand_totals"][c] for c in status_cols] + [table["grand_total"]])
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
SUBTOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
GRANDTOTAL_FILL = PatternFill("solid", start_color="2E5395")
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=NAVY)
SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
GRANDTOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="B7C5D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COUNT_FMT = '#,##0;-#,##0;"-"'
VALUE_FMT = '#,##0.00;-#,##0.00;"-"'
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_INDENT = Alignment(horizontal="left", indent=1)


def _set(ws, row, col, value, font=BODY_FONT, fill=None, fmt=None, align=None, border=BOX):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    for j, col in enumerate(df.columns, start=1):
        col_label = col.strftime("%d-%b") if isinstance(col, pd.Timestamp) else str(col)
        _set(ws, start_row, j, col_label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        width = max(10, min(28, len(col_label) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            fmt = None
            if isinstance(val, (np.integer,)):
                val = int(val)
                fmt = COUNT_FMT
            elif isinstance(val, (np.floating, float)):
                if pd.isna(val):
                    val = None
                else:
                    val = float(val)
                    fmt = VALUE_FMT
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
                fmt = "dd-mmm-yyyy"
            _set(ws, i, j, val, fmt=fmt)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    return start_row + 1 + len(df)


def _write_zone_tat_sheet(ws, title: str, summary: dict):
    """Writes the merged-header Zone x Receipt Type x TAT pivot, with an
    online-receipt-source mini block on the left, exactly mirroring the
    layout found in the original 'RTGS Summary' / 'Delay in RCPTING
    Summary' tabs.

    Column layout is entirely derived from TAT_ORDER's length (was
    hardcoded for exactly 3 buckets, which broke — MergedCell write error
    — the moment TAT_ORDER grew to 5 with the Pending/Cxn buckets, because
    the fixed "Total Count"/"Total Value" columns (K/L) landed on top of
    cells the bucket loop had already merged)."""
    matrix = summary["matrix"]
    source_block = summary["online_source_block"]

    zone_col = 4                                   # D: "ZONE / RECEIPT TOWARDS"
    first_bucket_col = zone_col + 1                 # E
    n_buckets = len(TAT_ORDER)
    total_count_col = first_bucket_col + n_buckets * 2   # right after the last bucket's 2 cols
    total_value_col = total_count_col + 1
    last_col = total_value_col

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _set(ws, 1, 1, title, font=TITLE_FONT, fill=None, align=Alignment(horizontal="left"), border=None)

    # --- online receipt source mini-block (cols A:B) ---
    ws.merge_cells("A2:B2")
    _set(ws, 2, 1, "Online Payment — Receipt Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 1, "Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 2, "Receipt Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    r = 4
    for name, cnt in source_block["rows"]:
        _set(ws, r, 1, name, align=LEFT_INDENT)
        _set(ws, r, 2, cnt, fmt=COUNT_FMT)
        r += 1
    _set(ws, r, 1, "GRAND TOTAL", font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL)
    _set(ws, r, 2, source_block["total"], font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL, fmt=COUNT_FMT)

    # --- main Zone x Receipt Type x TAT matrix ---
    ws.merge_cells(start_row=2, start_column=zone_col, end_row=3, end_column=zone_col)
    _set(ws, 2, zone_col, "ZONE / CATEGORY", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    col = first_bucket_col
    for bucket in TAT_ORDER:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        _set(ws, 2, col, TAT_DISPLAY[bucket], font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, col + 1, None, fill=HEADER_FILL)
        _set(ws, 3, col, "Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 3, col + 1, "Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        col += 2
    ws.merge_cells(start_row=2, start_column=total_count_col, end_row=3, end_column=total_count_col)
    ws.merge_cells(start_row=2, start_column=total_value_col, end_row=3, end_column=total_value_col)
    _set(ws, 2, total_count_col, "Total Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 2, total_value_col, "Total Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    def _write_stat_row(row, label_col, label, stats, font, fill, indent=False):
        align = LEFT_INDENT if indent else Alignment(horizontal="left")
        _set(ws, row, label_col, label, font=font, fill=fill, align=align)
        col = label_col + 1
        for bucket in TAT_ORDER:
            _set(ws, row, col, stats["by_bucket"][bucket]["count"], font=font, fill=fill, fmt=COUNT_FMT)
            _set(ws, row, col + 1, stats["by_bucket"][bucket]["value"], font=font, fill=fill, fmt=VALUE_FMT)
            col += 2
        _set(ws, row, col, stats["total_count"], font=font, fill=fill, fmt=COUNT_FMT)
        _set(ws, row, col + 1, stats["total_value"], font=font, fill=fill, fmt=VALUE_FMT)

    row = 4
    for block in matrix["zones"]:
        _write_stat_row(row, zone_col, block["zone"], block["subtotal"], SUBTOTAL_FONT, SUBTOTAL_FILL)
        row += 1
        for code, display, stats in block["breakdown"]:
            _write_stat_row(row, zone_col, display, stats, BODY_FONT, None, indent=True)
            row += 1
    _write_stat_row(row, zone_col, "GRAND TOTAL", matrix["grand_total"], GRANDTOTAL_FONT, GRANDTOTAL_FILL)

    fixed_widths = {1: 20, 2: 14, 3: 3, 4: 26, total_count_col: 12, total_value_col: 14}  # A,B,C(spacer),D,Total Count,Total Value
    for c in range(1, last_col + 1):
        if c in fixed_widths:
            width = fixed_widths[c]
        else:
            width = 9 if (c - first_bucket_col) % 2 == 0 else 12  # bucket pairs: Count(9) then Value(12)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = ws.cell(row=4, column=first_bucket_col).coordinate


def _write_receipt_made_summary_sheet(ws, summary: dict):
    def _write_table(start_col, title, status_cols, table):
        n = len(status_cols)
        last_col = start_col + 2 + n  # label cols (2) + status cols (n) + grand total (1)
        from openpyxl.utils import get_column_letter as gcl
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=last_col)
        _set(ws, 1, start_col, title, font=TITLE_FONT, align=Alignment(horizontal="left"), border=None)
        _set(ws, 2, start_col, "RECEIPT STATUS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, start_col + 1, "PAYMENT MODE", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        for k, c in enumerate(status_cols):
            _set(ws, 2, start_col + 2 + k, c.upper(), font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, last_col, "GRAND TOTAL", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

        row = 3
        for group in table["groups"]:
            first_row = row
            for r in group["rows"]:
                _set(ws, row, start_col, None)
                _set(ws, row, start_col + 1, r["mode"], align=LEFT_INDENT)
                for k, c in enumerate(status_cols):
                    _set(ws, row, start_col + 2 + k, r["counts"][c], fmt=COUNT_FMT)
                _set(ws, row, last_col, r["total"], font=SUBTOTAL_FONT, fmt=COUNT_FMT)
                row += 1
            if row > first_row:
                _set(ws, first_row, start_col, group["status"], font=SUBTOTAL_FONT, align=Alignment(horizontal="left"))
                if row - 1 > first_row:
                    ws.merge_cells(start_row=first_row, start_column=start_col, end_row=row - 1, end_column=start_col)
        _set(ws, row, start_col, "GRAND TOTAL", font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL)
        _set(ws, row, start_col + 1, None, fill=GRANDTOTAL_FILL)
        for k, c in enumerate(status_cols):
            _set(ws, row, start_col + 2 + k, table["grand_totals"][c], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        _set(ws, row, last_col, table["grand_total"], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        return last_col

    last = _write_table(1, "Receipt Made Summary — Updated / Pending",
                         ["Cleared", "Deposit", "Pending"], summary["left"])
    _write_table(last + 2, "Receipt Made Summary — Updated / Bounced or Cancelled",
                 _RECEIPT_MADE_STATUS_COLS, summary["right"])

    for letter, width in zip("ABCDEFGHIJKLMN", [20, 16, 11, 11, 11, 13, 3, 20, 16, 11, 11, 11, 11, 13]):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A3"


def write_output_workbook(
    rtgs_summary: dict,
    delay_summary: dict,
    receipt_made_summary: dict,
    cash_mode_validation_summary: pd.DataFrame,
    rcpt_cxn: pd.DataFrame,
    extra_tabs: dict[str, pd.DataFrame] | None = None,
) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Receipt made summary")
    _write_receipt_made_summary_sheet(ws, receipt_made_summary)

    ws = wb.create_sheet("RTGS Summary")
    _write_zone_tat_sheet(ws, "RTGS Summary", rtgs_summary)

    ws = wb.create_sheet("Cash Mode Validat Summary")
    _write_cash_mode_validation_sheet(ws, cash_mode_validation_summary)

    ws = wb.create_sheet("Delay in RCPTING Summary")
    _write_zone_tat_sheet(ws, "Delay in RCPTING Summary", delay_summary)

    ws = wb.create_sheet("RCPT CXN")
    _write_rcpt_cxn_sheet(ws, rcpt_cxn)

    for name, df in (extra_tabs or {}).items():
        ws = wb.create_sheet(name[:31])
        _write_df(ws, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# Single-sheet Excel exports for the "Export a Specific Summary" picker —
# each reuses the same sheet-writing logic as the full workbook so the
# numbers always match, just packaged as a standalone one-sheet .xlsx.
# ---------------------------------------------------------------------------
def rtgs_summary_to_excel_bytes(summary: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RTGS Summary")
    _write_zone_tat_sheet(ws, "RTGS Summary", summary)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def delay_summary_to_excel_bytes(summary: dict) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Delay in RCPTING Summary")
    _write_zone_tat_sheet(ws, "Delay in RCPTING Summary", summary)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_cash_mode_validation_sheet(ws, df: pd.DataFrame):
    """
    Cash Mode Validation Summary — shows every flagged CIF (each already
    matched at least one of the three violation rules in
    build_cash_mode_validation_summary: per-day > ₹1,95,000, Grand Total
    > ₹10,00,000, or a disabled CIF that still received cash), sorted by
    Grand Total descending, with a 'Violation Reason' column spelling out
    which rule(s) triggered for that row.
    """
    ws.cell(row=1, column=1, value="Cash Mode Validation Summary").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value=(
            f"Per-Day > ₹{CASH_DAILY_LIMIT:,}  ·  Grand Total > ₹{CASH_GRAND_TOTAL_LIMIT:,}  ·  "
            f"Disabled CIF still receipted  ·  Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}"
        ),
    ).font = Font(italic=True, size=9, color="666666")

    if df.empty:
        ws.cell(row=4, column=1, value="No violations detected — fully compliant").font = BODY_FONT
        return

    ordered = df.sort_values("Grand Total", ascending=False)
    reason = ordered["Violation Reason"].fillna("")
    reason_counts = {
        "Per-Day limit": int(reason.str.contains("Per-Day", na=False).sum()),
        "Grand Total limit": int(reason.str.contains("Grand Total", na=False).sum()),
        "Disabled CIF receipted": int(reason.str.contains("Disabled", na=False).sum()),
    }
    summary_text = "  ·  ".join(f"{k}: {v}" for k, v in reason_counts.items())

    _set(
        ws, 4, 1,
        f"VIOLATIONS ({len(ordered)} customers)  —  {summary_text}",
        font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        fill=PatternFill("solid", start_color="C0392B"),
    )
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(ordered.columns))
    _write_df(ws, ordered, start_row=5)


def cash_mode_validation_summary_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    """Standalone export — same high-value-only sheet as the combined workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Cash Mode Validation")
    _write_cash_mode_validation_sheet(ws, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_rcpt_cxn_sheet(ws, df: pd.DataFrame):
    """
    Cancelled receipts: a Mode-wise summary block up top (how many/how
    much cancelled per PaymentMode), then the detail rows grouped under a
    header band per payment mode with a subtotal — used by both the
    combined workbook and the standalone RCPT CXN export.
    """
    detail_cols = [c for c in df.columns if c != "PaymentMode"]
    n_cols = max(len(detail_cols), 5)

    ws.cell(row=1, column=1, value="Cancelled Receipts — by Payment Mode").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d-%b-%Y %H:%M')}").font = Font(italic=True, size=9, color="666666")

    if df.empty:
        ws.cell(row=4, column=1, value="No cancelled receipts this period").font = BODY_FONT
        return

    amt = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    by_mode = df.assign(_amt=amt).groupby("PaymentMode", dropna=False).agg(
        count=("_amt", "size"), amount=("_amt", "sum")
    ).reset_index().sort_values("count", ascending=False)

    # ── Summary block ──
    r = 4
    _set(ws, r, 1, "SUMMARY BY PAYMENT MODE", font=HEADER_FONT, fill=HEADER_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    for label, col in [("Payment Mode", 1), ("Cancelled Count", 2), ("Cancelled Amount", 3)]:
        _set(ws, r, col, label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    r += 1
    for _, row in by_mode.iterrows():
        mode = row["PaymentMode"] if pd.notna(row["PaymentMode"]) else UNMAPPED_LABEL
        _set(ws, r, 1, str(mode), align=LEFT_INDENT)
        _set(ws, r, 2, int(row["count"]), fmt=COUNT_FMT, align=CENTER)
        _set(ws, r, 3, float(row["amount"]), fmt=VALUE_FMT, align=CENTER)
        r += 1
    _set(ws, r, 1, "GRAND TOTAL", font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL)
    _set(ws, r, 2, int(len(df)), font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT, align=CENTER)
    _set(ws, r, 3, float(amt.sum()), font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=VALUE_FMT, align=CENTER)
    r += 3

    # ── Detail, grouped per mode ──
    for _, mode_row in by_mode.iterrows():
        mode = mode_row["PaymentMode"] if pd.notna(mode_row["PaymentMode"]) else UNMAPPED_LABEL
        group = df[df["PaymentMode"] == mode_row["PaymentMode"]] if pd.notna(mode_row["PaymentMode"]) else df[df["PaymentMode"].isna()]

        _set(ws, r, 1, f"{mode}  ({len(group)} receipt{'s' if len(group) != 1 else ''})", font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        r += 1

        for j, col in enumerate(detail_cols, start=1):
            _set(ws, r, j, str(col), font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        r += 1

        for _, row in group[detail_cols].iterrows():
            for j, col in enumerate(detail_cols, start=1):
                val = row[col]
                fmt = None
                if col == "Amount":
                    val = float(val) if pd.notna(val) else 0.0
                    fmt = VALUE_FMT
                elif pd.isna(val):
                    val = ""
                _set(ws, r, j, val, fmt=fmt)
            r += 1
        r += 2

    for j in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16


def rcpt_cxn_to_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    """Standalone export — same Mode-wise summary + grouped detail as the combined workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RCPT CXN")
    _write_rcpt_cxn_sheet(ws, df)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# BRS Status (Bank Reconciliation Statement) — Cheque/DD and Airtel Cash.
# Ported in from the standalone Collection Ops app so its "Delay in
# Deposition" / "Delay in Receipting" counts can feed the Daily Health
# Check mail directly, instead of leaving those two lines as
# [Fill manually]. Two separate BRS MIS files feed this: one for Cheque/DD,
# one for Airtel Cash — each uploaded and processed independently.
# ---------------------------------------------------------------------------
CHQ_TAT_BUCKETS = ["0-1", "2-4", "5-6", "Above 6"]
AIR_TAT_BUCKETS = ["0-1", "2-3", "4-5", "Above 5"]
OPS_BRS_ORDER = ["Credit Not Received", "Not Tally", "Tally"]
CHQ_RECEIPT_ROWS = {
    "Credit Not Received": ["Credit Not Received"],
    "Not Tally": ["Waiting for Re-Credit", "Amount Mismatch", "Challan No Mismatch"],
    "Tally": ["Delay in Deposition", "Delay in Receipting", "Nil Query"],
}
AIR_RECEIPT_ROWS = {
    "Tally": ["Delay In Receipting", "Delay In Deposition", "Nil Query"],
    "Credit Not Received": [],
    "Not Tally": [],
}


def _aging_to_bucket_chq(aging) -> str:
    try:
        a = int(float(aging))
    except (TypeError, ValueError):
        return "0-1"
    if a <= 1:
        return "0-1"
    elif a <= 4:
        return "2-4"
    elif a <= 6:
        return "5-6"
    return "Above 6"


def _aging_to_bucket_air(aging) -> str:
    try:
        a = int(float(aging))
    except (TypeError, ValueError):
        return "0-1"
    if a <= 1:
        return "0-1"
    elif a <= 3:
        return "2-3"
    elif a <= 5:
        return "4-5"
    return "Above 5"


def _aging_to_bucket_cash(aging) -> str:
    """Fallback only — _brs_pivot prefers the file's own TAT column text over this."""
    try:
        a = int(float(aging))
    except (TypeError, ValueError):
        return "0-1"
    if a <= 1:
        return "0-1"
    elif a <= 3:
        return "2-3"
    elif a <= 5:
        return "4-5"
    return "Above 6"


def _brs_normalize(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip()).upper()


def _canon_ops_brs_status(raw: str) -> str:
    """
    Canonicalizes an OPS BRS STATUS value to one of the three labels every
    downstream consumer (OPS_BRS_ORDER, CASH_RECEIPT_ROWS/CHQ_.../AIR_...,
    brs_status_image_bytes_v2's pivot.get("Tally", ...) lookups) expects:
    "Tally" / "Not Tally" / "Credit Not Received".

    Raw files spell these inconsistently -- e.g. the Cash BRS "DATA" sheet
    has "TALLY" in all caps -- and _brs_pivot previously kept whatever
    casing the file used. That mismatch (pivot key "TALLY" vs. every
    lookup using "Tally") silently zeroed out the zone-filtered Cash BRS
    image every time, regardless of which zone was picked, since the
    "Overall" mail path (load_cash_mis) had its own separate copy of this
    same canonicalization and wasn't affected. Centralizing it here so
    _brs_pivot always emits load_cash_mis-compatible casing fixes it for
    Cash and guards Cheque/Airtel against the same class of bug.
    """
    s = str(raw).strip().upper()
    if s == "TALLY":
        return "Tally"
    if s in ("NOT TALLY", "NOT_TALLY"):
        return "Not Tally"
    if "CREDIT" in s:
        return "Credit Not Received"
    return str(raw).strip()


def _brs_find_col(df: pd.DataFrame, *candidates: str) -> str | None:
    """Case/whitespace-insensitive column lookup, with a loose contains-match fallback."""
    norm_map = {_brs_normalize(c): c for c in df.columns}
    for cand in candidates:
        key = _brs_normalize(cand)
        if key in norm_map:
            return norm_map[key]
    for cand in candidates:
        key = _brs_normalize(cand)
        for norm, orig in norm_map.items():
            if key in norm or norm in key:
                return orig
    return None


def load_brs_raw(file) -> pd.DataFrame:
    """Reads a BRS MIS .xlsx, stripping column whitespace."""
    df = pd.read_excel(file)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _canonical_bucket(raw: str, canonical_buckets: list[str]) -> str:
    """Case/whitespace-insensitive match of a raw TAT-column value to one of
    the canonical bucket labels (e.g. " above 6 " -> "Above 6"). If it
    doesn't match any known label, the cleaned original text is kept as
    its own bucket rather than being silently dropped or miscounted."""
    norm = re.sub(r"\s+", "", str(raw).strip()).upper()
    if not norm or norm in {"NAN", "NONE"}:
        return "Unspecified"
    for b in canonical_buckets:
        if re.sub(r"\s+", "", b).upper() == norm:
            return b
    return str(raw).strip()


def _brs_pivot(df: pd.DataFrame, bucket_fn, canonical_buckets: list[str]) -> tuple[dict, dict, list[str]]:
    """
    Returns ({ ops_brs_status: { receipt_status: { bucket: count } } },
    grand_totals_by_bucket, buckets_in_display_order).

    Prefers the file's own "TAT" column (column I in the real MIS export)
    as the source of truth for each row's bucket — that's what Ops
    actually assigned it to. Only falls back to recomputing a bucket from
    AGING via bucket_fn if there's no TAT column at all. Recomputing from
    AGING when a TAT column already exists was the bug: this pipeline's
    own AGING-based bucketing could disagree with Ops' own TAT
    assignment, which is exactly why 2-4 / 5-6 / Above 6 weren't
    reflecting correctly — those rows were being re-bucketed instead of
    read as-given.

    Any TAT value that doesn't match one of the canonical labels (after
    case/whitespace normalization) is still counted, as its own bucket
    appended after the canonical ones — never silently dropped — so
    totals always reconcile with the row count.
    """
    ops_col = _brs_find_col(df, "OPS BRS STATUS", "OPS_BRS_STATUS")
    tally_col = _brs_find_col(df, "RECEIPT STATUS\n(Tally)", "RECEIPT STAUS (Tally)", "RECEIPT STATUS (Tally)", "RECEIPT STATUS")
    tat_col = _brs_find_col(df, "TAT")
    aging_col = _brs_find_col(df, "AGING", "Aging", "AGEING")
    if not all([ops_col, tally_col]):
        raise ValueError(f"Missing required BRS columns. Found: ops={ops_col}, tally={tally_col}")
    if tat_col is None and aging_col is None:
        raise ValueError("Missing both a TAT column and an AGING column — can't determine the TAT bucket.")

    d = df.copy()
    d["_OPS"] = d[ops_col].apply(_canon_ops_brs_status)
    d["_TALLY"] = d[tally_col].astype(str).str.strip()

    if tat_col is not None:
        d["_BUCKET"] = d[tat_col].apply(lambda v: _canonical_bucket(v, canonical_buckets))
    else:
        d["_BUCKET"] = d[aging_col].apply(bucket_fn)

    extra_buckets = [b for b in d["_BUCKET"].unique() if b not in canonical_buckets]
    buckets = canonical_buckets + sorted(extra_buckets)

    result: dict = {}
    for ops in d["_OPS"].unique():
        sub = d[d["_OPS"] == ops]
        result[ops] = {}
        for tally in sub["_TALLY"].unique():
            tsub = sub[sub["_TALLY"] == tally]
            result[ops][tally] = {b: int((tsub["_BUCKET"] == b).sum()) for b in buckets}

    grand = {b: int((d["_BUCKET"] == b).sum()) for b in buckets}
    return result, grand, buckets


def brs_delay_counts(pivot: dict) -> dict:
    """
    Pulls the two headline counts out of a BRS pivot: Delay in Deposition
    and Delay in Receipting, both counted only within the "Tally" OPS BRS
    group (that's where those receipt-status rows live) — case-insensitive
    since Cheque/DD rows spell it "Delay in Deposition" and Airtel rows
    spell it "Delay In Deposition".
    """
    tally = pivot.get("Tally", {})
    deposition = receipting = 0
    for receipt_status, buckets in tally.items():
        label = receipt_status.strip().upper()
        total = sum(buckets.values())
        if label == "DELAY IN DEPOSITION":
            deposition += total
        elif label == "DELAY IN RECEIPTING":
            receipting += total
    return {"delay_deposition": deposition, "delay_receipting": receipting}


def process_brs_for_mail(df: pd.DataFrame, mode: str) -> dict:
    """
    mode = "Cheque" or "Airtel". Returns pivot/grand/buckets/ops_rows/title
    plus the two headline delay counts, ready to slot into the Daily
    Health Check mail's "Delay in Receipting / Deposition" section.
    """
    if mode == "Cheque":
        pivot, grand, buckets = _brs_pivot(df, _aging_to_bucket_chq, CHQ_TAT_BUCKETS)
        ops_rows, title = CHQ_RECEIPT_ROWS, "CHEQUE/DD BRS Status"
    else:
        pivot, grand, buckets = _brs_pivot(df, _aging_to_bucket_air, AIR_TAT_BUCKETS)
        ops_rows, title = AIR_RECEIPT_ROWS, "Airtel Cash BRS Status"

    return {
        "pivot": pivot, "grand": grand, "buckets": buckets, "ops_rows": ops_rows,
        "title": title, "mode": mode, "delay": brs_delay_counts(pivot),
    }


def _write_brs_status_sheet(ws, result: dict):
    """
    BRS Status workbook sheet: title band, a simple KPI strip (one cell
    per OPS BRS category + total), then the full detail pivot — every OPS
    BRS category x receipt-status row, broken down by TAT bucket with a
    row total, subtotalled per category — same visual language (NAVY
    title, HEADER_FILL headers, SUBTOTAL/GRANDTOTAL bands) as the other
    sheets in this workbook.
    """
    pivot, grand, buckets, ops_rows, title = result["pivot"], result["grand"], result["buckets"], result["ops_rows"], result["title"]

    ops_totals = {}
    for ops in OPS_BRS_ORDER:
        rows = ops_rows.get(ops) or list(pivot.get(ops, {}).keys())
        ops_totals[ops] = sum(sum(pivot.get(ops, {}).get(r, {}).values()) for r in rows)
    grand_total = sum(ops_totals.values())

    label_col = 1
    first_bucket_col = 2
    n_buckets = len(buckets)
    total_col = first_bucket_col + n_buckets
    last_col = total_col

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _set(ws, 1, 1, title.upper(), font=TITLE_FONT, border=None)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    _set(ws, 2, 1, f"OPS BRS Status x TAT Bucket  ·  Generated {datetime.now():%d-%b-%Y %H:%M}",
         font=Font(name="Calibri", italic=True, size=9, color="666666"), border=None)

    # ── simple KPI strip ──
    kpi_fills = {
        "TOTAL RECEIPTS": PatternFill("solid", start_color=NAVY),
        "TALLY": PatternFill("solid", start_color="27AE60"),
        "NOT TALLY": PatternFill("solid", start_color="E67E22"),
        "CREDIT NOT RECEIVED": PatternFill("solid", start_color="C0392B"),
    }
    kpi_values = {
        "TOTAL RECEIPTS": grand_total,
        "TALLY": ops_totals.get("Tally", 0),
        "NOT TALLY": ops_totals.get("Not Tally", 0),
        "CREDIT NOT RECEIVED": ops_totals.get("Credit Not Received", 0),
    }
    kpi_row = 4
    n_kpi = len(kpi_fills)
    col = 1
    for i, (label, fill) in enumerate(kpi_fills.items()):
        end_col = last_col if i == n_kpi - 1 else max(col, round((i + 1) * last_col / n_kpi))
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=end_col)
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=end_col)
        _set(ws, kpi_row, col, label, font=Font(name="Calibri", bold=True, size=9, color="FFFFFF"), fill=fill, border=None)
        _set(ws, kpi_row + 1, col, kpi_values[label], font=Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
             fill=fill, fmt=COUNT_FMT, border=None)
        col = end_col + 1

    header_row = kpi_row + 3
    _set(ws, header_row, label_col, "OPS BRS STATUS / RECEIPT STATUS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    for j, b in enumerate(buckets):
        _set(ws, header_row, first_bucket_col + j, b, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, header_row, total_col, "TOTAL", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    row = header_row + 1
    first_data_row = row
    for ops in OPS_BRS_ORDER:
        rows = ops_rows.get(ops) or list(pivot.get(ops, {}).keys())
        _set(ws, row, label_col, ops.upper(), font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL)
        for j, b in enumerate(buckets):
            _set(ws, row, first_bucket_col + j, sum(pivot.get(ops, {}).get(r, {}).get(b, 0) for r in rows),
                 font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL, fmt=COUNT_FMT)
        _set(ws, row, total_col, ops_totals[ops], font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL, fmt=COUNT_FMT)
        row += 1
        for r in rows:
            row_total = sum(pivot.get(ops, {}).get(r, {}).values())
            if row_total == 0 and r not in pivot.get(ops, {}):
                continue
            _set(ws, row, label_col, r, font=BODY_FONT, align=LEFT_INDENT)
            for j, b in enumerate(buckets):
                _set(ws, row, first_bucket_col + j, pivot.get(ops, {}).get(r, {}).get(b, 0),
                     font=BODY_FONT, fmt=COUNT_FMT)
            _set(ws, row, total_col, row_total, font=BODY_FONT, fmt=COUNT_FMT)
            row += 1
    last_data_row = row - 1

    _set(ws, row, label_col, "GRAND TOTAL", font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL)
    for j, b in enumerate(buckets):
        _set(ws, row, first_bucket_col + j, grand.get(b, 0), font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
    _set(ws, row, total_col, grand_total, font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)

    ws.auto_filter.ref = f"{get_column_letter(label_col)}{header_row}:{get_column_letter(last_col)}{last_data_row}"

    ws.column_dimensions[get_column_letter(label_col)].width = 34
    for j in range(n_buckets):
        ws.column_dimensions[get_column_letter(first_bucket_col + j)].width = 15
    ws.column_dimensions[get_column_letter(total_col)].width = 16
    ws.freeze_panes = ws.cell(row=first_data_row, column=first_bucket_col).coordinate

    delay = result["delay"]
    r2 = row + 3
    _set(ws, r2, 1, "HEADLINE (used in mail)", font=HEADER_FONT, fill=HEADER_FILL, border=None)
    r2 += 1
    for label, val in [("Delay in Deposition", delay["delay_deposition"]), ("Delay in Receipting", delay["delay_receipting"])]:
        _set(ws, r2, 1, label, font=BODY_FONT, align=LEFT_INDENT)
        _set(ws, r2, 2, val, font=BODY_FONT, fmt=COUNT_FMT)
        r2 += 1


def brs_status_to_excel_bytes(result: dict) -> io.BytesIO:
    """Standalone BRS Status Excel export — one sheet, same detail as the dashboard PNG."""
    wb = Workbook()
    wb.remove(wb.active)
    safe_title = re.sub(r'[\\/*?:\[\]]', '-', result["title"])[:31]
    ws = wb.create_sheet(safe_title)
    _write_brs_status_sheet(ws, result)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def brs_status_image_bytes(result: dict) -> io.BytesIO:
    """
    BRS Status summary PNG — card-table + donut layout (matching the
    reference "Downloaded Type Summary" style): left is a banner-headed
    card table (Tally's receipt-status sub-rows on top, then the three
    OPS BRS categories as bold colour-bordered cards with count + % pill),
    right is a donut chart of the OPS BRS split with the grand total in
    the middle and a dotted legend underneath.
    """
    from PIL import Image, ImageDraw
    import math

    navy = (16, 40, 96)
    green = (39, 174, 96)
    red = (211, 47, 47)
    amber = (243, 146, 0)
    ink = (30, 34, 45)
    muted = (110, 118, 132)
    width = 1080

    pivot, grand, buckets, ops_rows, title = result["pivot"], result["grand"], result["buckets"], result["ops_rows"], result["title"]
    ops_colors = {"Tally": green, "Not Tally": amber, "Credit Not Received": red}
    ops_light = {"Tally": (232, 247, 237), "Not Tally": (255, 243, 224), "Credit Not Received": (253, 232, 232)}

    ops_totals = {}
    for ops in OPS_BRS_ORDER:
        rows = ops_rows.get(ops) or list(pivot.get(ops, {}).keys())
        ops_totals[ops] = sum(sum(pivot.get(ops, {}).get(r, {}).values()) for r in rows)
    grand_total = sum(ops_totals.values()) or 1

    tally_rows = [(r, sum(pivot.get("Tally", {}).get(r, {}).values())) for r in ops_rows.get("Tally", [])]

    pad = 30
    left_w = 500
    right_x0 = pad + left_w + 30
    right_w = width - right_x0 - pad

    banner_h = 54
    sub_row_h = 50
    main_row_h = 62
    gap_between_groups = 14
    left_content_h = (
        banner_h + 16
        + len(tally_rows) * (sub_row_h + 10)
        + gap_between_groups
        + len(OPS_BRS_ORDER) * (main_row_h + 10)
    )
    right_content_h = 480
    height = max(left_content_h, right_content_h) + pad * 2 + 30

    base = Image.new("RGB", (width, height), (247, 248, 250))
    draw = ImageDraw.Draw(base)

    # ── LEFT: card table ──
    lx0, ly0 = pad, pad
    draw.rounded_rectangle([lx0, ly0, lx0 + left_w, ly0 + banner_h], radius=8, fill=navy)
    draw.ellipse([lx0 + 12, ly0 + 12, lx0 + 12 + 30, ly0 + 12 + 30], fill=(255, 255, 255))
    draw.text((lx0 + 27, ly0 + 27), "\u2193", font=_img_font(16), fill=navy, anchor="mm")
    draw.text((lx0 + 56, ly0 + banner_h / 2), "BRS STATUS SUMMARY", font=_img_font(16), fill="white", anchor="lm")

    y = ly0 + banner_h + 16

    def pct(n):
        return f"{n / grand_total * 100:.1f}%"

    def icon(cx, cy, r, color):
        draw.ellipse([cx - r, cy - r, cx + r, cy + r], outline=color, width=3)
        draw.ellipse([cx - r * 0.45, cy - r * 0.65, cx + r * 0.45, cy - r * 0.05], fill=color)
        draw.pieslice([cx - r * 0.7, cy - r * 0.1, cx + r * 0.7, cy + r * 0.9], 180, 360, fill=color)

    tally_colors = {"DELAY IN DEPOSITION": amber, "DELAY IN RECEIPTING": red, "NIL QUERY": green}

    # Sub-rows: Tally's own receipt-status breakdown, small & light
    for label, cnt in tally_rows:
        row_color = tally_colors.get(label.strip().upper(), green)
        draw.rounded_rectangle([lx0, y, lx0 + left_w, y + sub_row_h], radius=8, fill="white", outline=(226, 230, 236))
        icon(lx0 + 30, y + sub_row_h / 2, 13, row_color)
        draw.text((lx0 + 58, y + sub_row_h / 2), label, font=_img_font(13), fill=ink, anchor="lm")
        draw.text((lx0 + left_w - 130, y + sub_row_h / 2), f"{cnt:,}", font=_img_font(15), fill=navy, anchor="lm")
        pct_font = _img_font(11)
        pct_text = pct(cnt)
        pw = max(62, draw.textlength(pct_text, font=pct_font) + 24)
        draw.rounded_rectangle([lx0 + left_w - 58, y + sub_row_h / 2 - 13, lx0 + left_w - 58 + pw, y + sub_row_h / 2 + 13],
                                radius=12, fill=(230, 237, 248))
        draw.text((lx0 + left_w - 58 + pw / 2, y + sub_row_h / 2), pct_text, font=pct_font, fill=navy, anchor="mm")
        y += sub_row_h + 10

    y += gap_between_groups

    # Main OPS BRS category cards, bold colour border
    for ops in OPS_BRS_ORDER:
        color = ops_colors[ops]
        light = ops_light[ops]
        draw.rounded_rectangle([lx0, y, lx0 + left_w, y + main_row_h], radius=10, fill=light, outline=color, width=2)
        icon(lx0 + 34, y + main_row_h / 2, 16, color)
        draw.text((lx0 + 64, y + main_row_h / 2), ops.upper(), font=_img_font(14), fill=ink, anchor="lm")
        draw.text((lx0 + left_w - 150, y + main_row_h / 2), f"{ops_totals[ops]:,}", font=_img_font(18), fill=color, anchor="lm")
        ops_pct_font = _img_font(12)
        ops_pct_text = pct(ops_totals[ops])
        pw = max(66, draw.textlength(ops_pct_text, font=ops_pct_font) + 24)
        draw.rounded_rectangle([lx0 + left_w - 76, y + main_row_h / 2 - 15, lx0 + left_w - 76 + pw, y + main_row_h / 2 + 15],
                                radius=14, fill=color)
        draw.text((lx0 + left_w - 76 + pw / 2, y + main_row_h / 2), ops_pct_text, font=ops_pct_font, fill="white", anchor="mm")
        y += main_row_h + 10

    # ── RIGHT: donut chart — Tally's own breakdown (Delay in Deposition /
    # Delay in Receipting / Nil Query), with counts shown, not just %.
    rx0, ry0 = right_x0, pad
    dash = "\u2022"
    heading = f"{dash} DELAY BREAKDOWN (TALLY) {dash}"
    draw.text((rx0 + right_w / 2, ry0 + 18), heading, font=_img_font(16), fill=navy, anchor="mm")

    tally_total = sum(c for _, c in tally_rows) or 1

    cx, cy = rx0 + right_w / 2, ry0 + 230
    r_outer, r_inner = 150, 88
    start = -90
    legend_items = []
    for label, cnt in tally_rows:
        color = tally_colors.get(label.strip().upper(), muted)
        frac = cnt / tally_total
        sweep = 360 * frac
        if cnt > 0:
            draw.pieslice([cx - r_outer, cy - r_outer, cx + r_outer, cy + r_outer], start, start + sweep, fill=color)
        mid_angle = start + sweep / 2
        if cnt > 0 and frac > 0.02:
            lx = cx + (r_outer + 34) * math.cos(math.radians(mid_angle))
            ly_ = cy + (r_outer + 34) * math.sin(math.radians(mid_angle))
            label_text = f"{cnt:,}"
            tw = draw.textlength(label_text, font=_img_font(13))
            draw.rounded_rectangle([lx - tw / 2 - 8, ly_ - 14, lx + tw / 2 + 8, ly_ + 14], radius=8, fill=color)
            draw.text((lx, ly_), label_text, font=_img_font(13), fill="white", anchor="mm")
        start += sweep
        legend_items.append((label, color, cnt))

    draw.ellipse([cx - r_inner, cy - r_inner, cx + r_inner, cy + r_inner], fill="white")
    icon(cx, cy - 22, 20, navy)
    draw.text((cx, cy + 14), f"{tally_total:,}", font=_img_font(26), fill=navy, anchor="mm")
    draw.text((cx, cy + 40), "TALLY RECEIPTS", font=_img_font(10), fill=muted, anchor="mm")

    ly = cy + r_outer + 50
    lx = rx0 + 20
    for label, color, cnt in legend_items:
        draw.ellipse([lx, ly - 6, lx + 12, ly + 6], fill=color)
        draw.text((lx + 20, ly), label, font=_img_font(12), fill=ink, anchor="lm")
        sub = f"{cnt:,}  ({cnt / tally_total * 100:.1f}%)"
        draw.text((lx + 20, ly + 16), sub, font=_img_font(10), fill=muted, anchor="lm")
        ly += 44

    draw.text((pad, height - 22), f"{title} \u00b7 Generated {datetime.now().strftime('%d-%b-%Y %H:%M')} \u00b7 DHC Working Automation",
              font=_img_font(10), fill=(160, 160, 165))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf


# =====================================================================
# CASH MIS — parser + unified BRS image
# =====================================================================

CASH_TAT_BUCKETS = ["0-1", "2-3", "4-5", "Above 6"]
CASH_RECEIPT_ROWS = {
    "Tally": ["Delay In Receipting", "Delay In Deposition", "Nil Query"],
    "Not Tally": [],
    "Credit Not Received": ["Credit Not Received"],
}

_TALLY_COLORS_IMG = {
    "DELAY IN DEPOSITION":  (243, 146,  0),   # amber
    "DELAY IN RECEIPTING":  (211,  47, 47),    # red
    "NIL QUERY":            ( 39, 174, 96),    # green
    "DELAY IN RECEIPTING":  (211,  47, 47),
}


def load_cash_brs_raw(file) -> pd.DataFrame:
    """
    Reads the Cash BRS MIS's raw per-row "DATA" sheet (not the pre-built
    "MIS" pivot sheet load_cash_mis() reads) — this is the sheet with a
    Zone column, needed to support zone-filtered mail drafts, since the
    "MIS" sheet is already a collapsed summary with no per-row zone to
    filter by.
    """
    df = pd.read_excel(file, sheet_name="DATA")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def process_cash_brs_for_mail(df: pd.DataFrame) -> dict:
    """
    Cash BRS built from the DATA sheet's raw rows via the same _brs_pivot
    approach as Cheque/Airtel — used only for zone-filtered mail drafts.
    The unfiltered "Overall" mail keeps using load_cash_mis()'s pre-built
    "MIS" sheet numbers unchanged, since that sheet may carry manual
    adjustments beyond what's in DATA; this path exists purely because
    "MIS" has no per-row zone to filter by.
    """
    pivot, grand, buckets = _brs_pivot(df, _aging_to_bucket_cash, CASH_TAT_BUCKETS)
    return {
        "pivot": pivot, "grand": grand, "buckets": buckets,
        "ops_rows": CASH_RECEIPT_ROWS, "title": "Cash BRS Status",
        "mode": "Cash", "delay": brs_delay_counts(pivot),
    }


def load_cash_mis(file) -> dict:
    """
    Parses the Cash BRS MIS pre-formatted Excel (no raw-row structure —
    it's already a summary table).  Returns the same dict shape as
    process_brs_for_mail so downstream image/mail functions work
    identically for all three modes.
    """
    df = pd.read_excel(file, header=None)

    # Find the header row: the row that contains "BRS Status" AND
    # "Receipting Status" (row 3 in the sample, 0-indexed row 3)
    header_row_idx = None
    for i, row in df.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v)]
        if "BRS Status" in vals and "Receipting Status" in vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Could not locate the header row in the Cash BRS MIS file.")

    # Title is the last non-empty string value before the header row
    title = "Cash BRS Status"
    for i in range(header_row_idx - 1, -1, -1):
        vals = [str(v).strip() for v in df.iloc[i] if pd.notna(v) and str(v).strip()]
        if vals:
            title = vals[0]
            break

    # Bucket labels are the cells after "Receipting Status" in the header row
    header_vals = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[header_row_idx]]
    rs_col = next((ci for ci, v in enumerate(header_vals) if v == "Receipting Status"), None)
    if rs_col is None:
        raise ValueError("Could not find 'Receipting Status' column in header row.")

    brs_col = rs_col - 1        # "BRS Status" column
    first_bucket_col = rs_col + 1
    # Bucket labels until we hit "Grand Total" or end
    buckets = []
    for ci in range(first_bucket_col, len(header_vals)):
        v = header_vals[ci]
        if v == "Grand Total" or v == "":
            break
        buckets.append(v)
    grand_total_col = first_bucket_col + len(buckets)

    # Canonicalize buckets to CASH_TAT_BUCKETS where possible
    def _canon(b):
        nb = re.sub(r"\s+", "", b).upper()
        for cb in CASH_TAT_BUCKETS:
            if re.sub(r"\s+", "", cb).upper() == nb:
                return cb
        return b
    buckets = [_canon(b) for b in buckets]

    # Parse data rows: until we hit "Total No. Receipt" or a row with
    # only NaN/empty values after the header
    pivot = {}
    current_brs = None
    for i in range(header_row_idx + 1, len(df)):
        row = df.iloc[i]
        brs_val  = str(row.iloc[brs_col]).strip()  if pd.notna(row.iloc[brs_col])  else ""
        rs_val   = str(row.iloc[rs_col]).strip()   if pd.notna(row.iloc[rs_col])   else ""
        first_bkt= str(row.iloc[first_bucket_col]).strip() if pd.notna(row.iloc[first_bucket_col]) else ""

        # Stop rows
        if brs_val.upper().startswith("TOTAL") or (not brs_val and not rs_val and not first_bkt):
            if not brs_val and not rs_val:
                continue   # blank spacer row, keep going
            break

        if brs_val:
            current_brs = brs_val.strip()
            # Canonicalize
            if current_brs.upper() == "TALLY":
                current_brs = "Tally"
            elif current_brs.upper() in ("NOT TALLY", "NOT_TALLY"):
                current_brs = "Not Tally"
            elif "CREDIT" in current_brs.upper():
                current_brs = "Credit Not Received"

        if current_brs not in pivot:
            pivot[current_brs] = {}

        # Receipt-status label: rs_val if present, else use brs label itself
        receipt_label = rs_val if rs_val and rs_val != "-" else (current_brs if current_brs else "—")

        bucket_counts = {}
        for bi, b in enumerate(buckets):
            raw = str(row.iloc[first_bucket_col + bi]).strip() if pd.notna(row.iloc[first_bucket_col + bi]) else "-"
            try:
                bucket_counts[b] = int(float(raw))
            except (ValueError, TypeError):
                bucket_counts[b] = 0

        pivot[current_brs][receipt_label] = bucket_counts

    grand = {b: sum(pivot.get(ops, {}).get(r, {}).get(b, 0)
                    for ops in pivot for r in pivot[ops])
             for b in buckets}

    delay_dep = sum(
        sum(pivot.get("Tally", {}).get(r, {}).values())
        for r in pivot.get("Tally", {})
        if r.strip().upper() == "DELAY IN DEPOSITION"
    )
    delay_rec = sum(
        sum(pivot.get("Tally", {}).get(r, {}).values())
        for r in pivot.get("Tally", {})
        if r.strip().upper() == "DELAY IN RECEIPTING"
    )

    return {
        "pivot": pivot,
        "grand": grand,
        "buckets": buckets,
        "ops_rows": CASH_RECEIPT_ROWS,
        "title": title,
        "mode": "Cash",
        "delay": {"delay_deposition": delay_dep, "delay_receipting": delay_rec},
    }


def brs_status_image_bytes_v2(result: dict) -> io.BytesIO:
    """
    Redesigned BRS Status summary PNG.

    LEFT panel — three "receipt-status" cards (Delay in Deposition /
    Delay in Receipting / Nil Query), each showing count + % pill,
    coloured amber / red / green respectively.  This replaces the old
    mixed card/OPS-category layout that was confusing and didn't match
    what the user wanted.

    RIGHT panel — donut chart of those same three items inside "Tally",
    with count bubbles on the segments and a dotted legend below.

    A header bar showing the dynamic title
    ("CHEQUE/DD BRS Status / Airtel Cash BRS Status / Cash BRS Status")
    spans the full width at the top so the report is self-labelling.
    """
    from PIL import Image, ImageDraw
    import math

    navy  = (16, 40, 96)
    green = (39, 174, 96)
    red   = (211, 47, 47)
    amber = (243, 146, 0)
    ink   = (30, 34, 45)
    muted = (110, 118, 132)
    bg    = (247, 248, 250)

    pivot    = result["pivot"]
    ops_rows = result["ops_rows"]
    title    = result["title"]
    buckets  = result["buckets"]

    # Tally sub-row items: Delay in Deposition, Delay in Receipting, Nil Query
    tally_row_labels = ops_rows.get("Tally", [])
    tally_rows = []
    for label in tally_row_labels:
        cnt = sum(pivot.get("Tally", {}).get(label, {}).values())
        # case-insensitive fallback
        if cnt == 0:
            for k, v in pivot.get("Tally", {}).items():
                if k.strip().upper() == label.strip().upper():
                    cnt = sum(v.values())
        tally_rows.append((label, cnt))

    tally_total = sum(c for _, c in tally_rows) or 1

    # OPS totals for grand total
    all_ops_total = sum(
        sum(sum(v.values()) for v in pivot.get(ops, {}).values())
        for ops in OPS_BRS_ORDER
    ) or 1

    TALLY_COLORS = {
        "DELAY IN DEPOSITION": amber,
        "DELAY IN RECEIPTING": red,
        "NIL QUERY":           green,
    }
    TALLY_LIGHT = {
        "DELAY IN DEPOSITION": (255, 243, 224),
        "DELAY IN RECEIPTING": (253, 232, 232),
        "NIL QUERY":           (232, 247, 237),
    }

    width  = 1100
    header_h = 60
    pad    = 30
    left_w = 460
    right_x0 = pad + left_w + 40
    right_w  = width - right_x0 - pad

    card_h  = 82
    card_gap = 16

    left_content_h = len(tally_rows) * (card_h + card_gap) + 30
    right_content_h = 660
    content_h = max(left_content_h, right_content_h)
    height = header_h + pad + content_h + pad + 28

    base = Image.new("RGB", (width, height), bg)
    draw = ImageDraw.Draw(base)

    # ── Full-width header bar ─────────────────────────────────────────
    draw.rounded_rectangle([0, 0, width, header_h], radius=0, fill=navy)
    draw.text((width / 2, header_h / 2), title.upper(),
              font=_img_font(23), fill="white", anchor="mm")

    y_start = header_h + pad

    # ── LEFT: receipt-status cards ───────────────────────────────────
    y = y_start
    for label, cnt in tally_rows:
        color = TALLY_COLORS.get(label.strip().upper(), muted)
        light = TALLY_LIGHT.get(label.strip().upper(), (240, 240, 240))
        pct_val = f"{cnt / tally_total * 100:.1f}%"

        draw.rounded_rectangle([pad, y, pad + left_w, y + card_h],
                                radius=12, fill=light, outline=color, width=2)

        # Person icon
        cx_icon, cy_icon = pad + 40, y + card_h // 2
        r = 16
        draw.ellipse([cx_icon - r, cy_icon - r, cx_icon + r, cy_icon + r],
                     outline=color, width=3)
        draw.ellipse([cx_icon - r * 0.45, cy_icon - r * 0.65,
                      cx_icon + r * 0.45, cy_icon - r * 0.05], fill=color)
        draw.pieslice([cx_icon - r * 0.7, cy_icon - r * 0.1,
                       cx_icon + r * 0.7, cy_icon + r * 0.9], 180, 360, fill=color)

        # % pill — sized from real text, pinned to the card's right edge
        pct_pill_font = _img_font(15)
        pw = max(84, draw.textlength(pct_val, font=pct_pill_font) + 28)
        pill_x = pad + left_w - 8 - pw
        draw.rounded_rectangle([pill_x, y + card_h // 2 - 19,
                                 pill_x + pw, y + card_h // 2 + 19],
                                radius=17, fill=color)
        draw.text((pill_x + pw / 2, y + card_h // 2), pct_val,
                  font=pct_pill_font, fill="white", anchor="mm")

        # Count — right-anchored just before the pill, so it can never run
        # into it regardless of how wide the pill ends up
        count_text = f"{cnt:,}"
        count_font = _img_font(23)
        count_right_x = pill_x - 14
        draw.text((count_right_x, y + card_h // 2), count_text,
                  font=count_font, fill=navy, anchor="rm")

        # Label — auto-shrinks to whatever space is left before the count,
        # so long labels (e.g. "Delay in Receipting") never overlap it
        label_start_x = pad + 74
        count_w = draw.textlength(count_text, font=count_font)
        avail_w = count_right_x - count_w - 14 - label_start_x
        label_font = _img_font(18)
        label_size = 18
        while draw.textlength(label, font=label_font) > avail_w and label_size > 11:
            label_size -= 1
            label_font = _img_font(label_size)
        draw.text((label_start_x, y + card_h // 2), label,
                  font=label_font, fill=ink, anchor="lm")

        y += card_h + card_gap

    # ── RIGHT: donut ─────────────────────────────────────────────────
    rx0, ry0 = right_x0, y_start

    dot = "\u2022"
    draw.text((rx0 + right_w / 2, ry0 + 22),
              f"{dot} DELAY BREAKDOWN (TALLY) {dot}",
              font=_img_font(19), fill=navy, anchor="mm")

    cx = rx0 + right_w / 2
    cy = ry0 + 250
    r_outer, r_inner = 155, 92

    start = -90
    legend_items = []
    for label, cnt in tally_rows:
        color = TALLY_COLORS.get(label.strip().upper(), muted)
        frac  = cnt / tally_total
        sweep = 360 * frac
        if cnt > 0:
            draw.pieslice([cx - r_outer, cy - r_outer,
                           cx + r_outer, cy + r_outer],
                          start, start + sweep, fill=color)
        mid_angle = start + sweep / 2
        if cnt > 0 and frac > 0.015:
            lx_ = cx + (r_outer + 38) * math.cos(math.radians(mid_angle))
            ly_ = cy + (r_outer + 38) * math.sin(math.radians(mid_angle))
            label_text = f"{cnt:,}"
            tw = draw.textlength(label_text, font=_img_font(17))
            draw.rounded_rectangle([lx_ - tw / 2 - 10, ly_ - 17,
                                     lx_ + tw / 2 + 10, ly_ + 17],
                                   radius=10, fill=color)
            draw.text((lx_, ly_), label_text,
                      font=_img_font(17), fill="white", anchor="mm")
        start += sweep
        legend_items.append((label, color, cnt))

    # Donut hole
    draw.ellipse([cx - r_inner, cy - r_inner,
                  cx + r_inner, cy + r_inner], fill="white")
    # Centre icon + total
    ci_cx, ci_cy = int(cx), int(cy) - 22
    ri = 20
    draw.ellipse([ci_cx - ri, ci_cy - ri, ci_cx + ri, ci_cy + ri],
                 outline=navy, width=3)
    draw.ellipse([ci_cx - ri * 0.45, ci_cy - ri * 0.65,
                  ci_cx + ri * 0.45, ci_cy - ri * 0.05], fill=navy)
    draw.pieslice([ci_cx - ri * 0.7, ci_cy - ri * 0.1,
                   ci_cx + ri * 0.7, ci_cy + ri * 0.9], 180, 360, fill=navy)
    draw.text((cx, cy + 14), f"{tally_total:,}",
              font=_img_font(32), fill=navy, anchor="mm")
    draw.text((cx, cy + 44), "TALLY RECEIPTS",
              font=_img_font(13), fill=muted, anchor="mm")

    # Legend — pushed down far enough that even a bubble sitting straight
    # below the donut (max reach = r_outer + 38 + half its own height)
    # can't collide with the first legend row
    leg_y = cy + r_outer + 38 + 17 + 20
    leg_x = rx0 + 10
    for label, color, cnt in legend_items:
        draw.ellipse([leg_x, leg_y - 7, leg_x + 15, leg_y + 8], fill=color)
        draw.text((leg_x + 24, leg_y), label,
                  font=_img_font(16), fill=ink, anchor="lm")
        sub = f"{cnt:,}  ({cnt / tally_total * 100:.1f}%)"
        draw.text((leg_x + 24, leg_y + 21), sub,
                  font=_img_font(13), fill=muted, anchor="lm")
        leg_y += 56

    # Footer
    draw.text((pad, height - 20),
              f"{title} \u00b7 Generated {datetime.now().strftime('%d-%b-%Y %H:%M')} \u00b7 DHC",
              font=_img_font(10), fill=(160, 160, 165))

    buf = io.BytesIO()
    base.save(buf, format="PNG")
    buf.seek(0)
    return buf
